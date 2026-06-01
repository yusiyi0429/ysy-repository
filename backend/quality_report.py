#!/usr/bin/env python3
"""
质量评分报告生成脚本

基于 Excel 知识库数据，按五维度计算质量评分并生成 Markdown 报告。
评分维度：完整性(25)、准确性(25)、可操作性(20)、反模式覆盖(15)、来源可溯(15)
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed"}, ensure_ascii=False))
    sys.exit(1)

try:
    import yaml
except ImportError:
    print(json.dumps({"status": "error", "message": "pyyaml not installed"}, ensure_ascii=False))
    sys.exit(1)

from field_aliases import FIELD_ALIASES, resolve_header

NON_KNOWLEDGE_SHEETS = {"场景配置", "版本追踪", "配置", "config", "来源索引"}


def _is_knowledge_sheet(ws) -> bool:
    """Check if a worksheet looks like it contains knowledge items."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
    if not any(headers):
        return False
    header_strs = {str(h).strip() for h in headers if h}
    key_fields = {"知识分类", "分类", "知识描述", "描述", "知识内容", "内容", "知识编号", "编号",
                  "知识要点", "数据规则", "专家经验", "步骤", "知识", "场景", "具体方法"}
    matches = len(header_strs & key_fields)
    return matches >= 2


def read_excel_data(excel_path: str) -> tuple:
    """读取 Excel 知识条目，支持传统格式和多sheet格式"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    records = []

    # Try legacy format first
    if "知识条目" in wb.sheetnames:
        ws = wb["知识条目"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
        if any(headers):
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or not row[0]:
                    continue
                record = {}
                for idx, header in enumerate(headers):
                    record[header] = row[idx] if idx < len(row) else ""
                records.append(record)
    else:
        # Multi-sheet format: iterate all sheets, skip non-knowledge ones
        for sn in wb.sheetnames:
            if sn in NON_KNOWLEDGE_SHEETS:
                continue
            ws = wb[sn]
            if _is_knowledge_sheet(ws):
                headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
                headers = [resolve_header(h) for h in headers_raw]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not any(row):
                        continue
                    record = {}
                    for idx, header in enumerate(headers):
                        val = row[idx] if idx < len(row) else ""
                        if header:
                            record[header] = val
                    if record.get("知识分类") or record.get("知识描述"):
                        records.append(record)

    # 读取场景配置 Sheet
    config_info = {}
    if "场景配置" in wb.sheetnames:
        ws2 = wb["场景配置"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                config_info[str(row[0])] = str(row[1])

    wb.close()
    return records, config_info


def load_config(config_path: str) -> dict:
    """加载 YAML 配置"""
    if not config_path or not os.path.exists(config_path):
        return {}
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def score_completeness(records: list, config: dict) -> dict:
    """完整性评分 (0-25)"""
    total = len(records)
    quality_cfg = config.get("quality", {})
    target = quality_cfg.get("target_per_pipeline", quality_cfg.get("min_knowledge_items", 30))

    # 条目数量分 (0-10)
    if total >= target:
        quantity_score = 10
    elif total >= target * 0.5:
        quantity_score = round(10 * total / max(target, 1), 1)
    else:
        quantity_score = round(5 * total / max(target * 0.5, 1), 1)

    # 必填字段填充率 (0-10) — 从 schema fields[].required 动态读取
    fields_cfg = config.get("fields", [])
    if fields_cfg:
        required_fields = [f["name"] for f in fields_cfg if f.get("required")]
    else:
        required_fields = ["知识编号", "知识分类", "知识描述", "适用条件"]
    filled_count = 0
    total_slots = len(records) * max(len(required_fields), 1)
    for rec in records:
        for field in required_fields:
            if str(rec.get(field, "")).strip():
                filled_count += 1
    fill_rate = filled_count / max(total_slots, 1)
    fill_score = round(10 * fill_rate, 1)

    # 分类覆盖度 (0-5)
    categories = set()
    for rec in records:
        cat = rec.get("知识分类", "")
        if cat:
            categories.add(cat)
    expected_categories = config.get("categories", ["判断规则", "操作流程", "反模式"])
    category_coverage = len(categories & set(expected_categories)) / max(len(expected_categories), 1)
    category_score = round(5 * category_coverage, 1)

    raw_score = quantity_score + fill_score + category_score
    score = min(raw_score, 25)

    return {
        "score": score,
        "max_score": 25,
        "details": {
            "quantity": {"score": quantity_score, "max": 10, "total_items": total, "target": target},
            "fill_rate": {"score": fill_score, "max": 10, "rate": round(fill_rate, 2)},
            "category_coverage": {"score": category_score, "max": 5, "covered": len(categories), "expected": len(expected_categories)},
        }
    }


def score_accuracy(records: list, config: dict = None) -> dict:
    """准确性评分 (0-25)"""
    total = len(records)

    # 置信度分布 (0-15)
    high = sum(1 for r in records if str(r.get("置信度", "")).strip() == "高")
    mid = sum(1 for r in records if str(r.get("置信度", "")).strip() == "中")
    low = sum(1 for r in records if str(r.get("置信度", "")).strip() == "低")
    high_ratio = high / max(total, 1)
    mid_ratio = mid / max(total, 1)

    conf_score = round(15 * (high_ratio * 1.0 + mid_ratio * 0.7), 1)
    conf_score = min(conf_score, 15)

    # 知识编号规范性 (0-5)
    valid_ids = sum(1 for r in records if str(r.get("知识编号", "")).strip().startswith(("KN-", "KN")))
    id_ratio = valid_ids / max(total, 1)
    id_score = round(5 * id_ratio, 1)
    # auto_generate_id 时放宽：有编号即满分（已由系统生成）
    if (config or {}).get("auto_generate_id") and valid_ids > 0:
        id_score = 5

    # 描述完整性 (0-5)
    desc_with_condition = sum(1 for r in records if str(r.get("知识描述", "")).strip() and str(r.get("适用条件", "")).strip())
    desc_ratio = desc_with_condition / max(total, 1)
    desc_score = round(5 * desc_ratio, 1)

    score = min(conf_score + id_score + desc_score, 25)

    return {
        "score": score,
        "max_score": 25,
        "details": {
            "confidence": {"score": conf_score, "max": 15, "high": high, "mid": mid, "low": low},
            "id_format": {"score": id_score, "max": 5, "valid_ratio": round(id_ratio, 2)},
            "desc_completeness": {"score": desc_score, "max": 5, "ratio": round(desc_ratio, 2)},
        }
    }


def score_operability(records: list) -> dict:
    """可操作性评分 (0-20)"""
    total = len(records)

    # 判断逻辑覆盖率 (0-10)
    with_logic = sum(1 for r in records if str(r.get("判断逻辑", "")).strip())
    logic_ratio = with_logic / max(total, 1)
    logic_score = round(10 * logic_ratio, 1)

    # 知识描述可执行性 (0-10)
    # 检查描述中是否包含条件关键词
    condition_keywords = ["当", "如果", "若", "超过", "低于", "高于", "等于", "达到", "触发"]
    with_condition = sum(1 for r in records
                         if any(kw in str(r.get("知识描述", "")) for kw in condition_keywords))
    condition_ratio = with_condition / max(total, 1)
    condition_score = round(10 * condition_ratio, 1)

    score = min(logic_score + condition_score, 20)

    return {
        "score": score,
        "max_score": 20,
        "details": {
            "logic_coverage": {"score": logic_score, "max": 10, "ratio": round(logic_ratio, 2), "count": with_logic},
            "condition_clarity": {"score": condition_score, "max": 10, "ratio": round(condition_ratio, 2), "count": with_condition},
        }
    }


def score_anti_pattern(records: list) -> dict:
    """反模式覆盖评分 (0-15)"""
    total = len(records)

    # 反模式覆盖率 (0-10)
    with_anti = sum(1 for r in records if str(r.get("反模式/踩坑提示", "")).strip())
    anti_ratio = with_anti / max(total, 1)
    anti_score = round(10 * min(anti_ratio / 0.3, 1.0), 1)  # 30%覆盖率为满分

    # 反模式分类条目数 (0-5)
    anti_category = sum(1 for r in records
                       if str(r.get("知识分类", "")).strip() == "反模式" and str(r.get("知识描述", "")).strip())
    anti_cat_score = min(round(5 * anti_category / 50, 1), 5)  # 50条反模式为满分

    score = min(anti_score + anti_cat_score, 15)

    return {
        "score": score,
        "max_score": 15,
        "details": {
            "anti_coverage": {"score": anti_score, "max": 10, "ratio": round(anti_ratio, 2), "count": with_anti},
            "anti_category": {"score": anti_cat_score, "max": 5, "count": anti_category},
        }
    }


def score_traceability(records: list) -> dict:
    """来源可溯评分 (0-15)"""
    total = len(records)

    # 来源标注率 (0-10)
    with_source = sum(1 for r in records if str(r.get("来源文档", "")).strip())
    source_ratio = with_source / max(total, 1)
    source_score = round(10 * source_ratio, 1)

    # 署名完整率 (0-5)
    with_attribution = sum(1 for r in records
                          if str(r.get("贡献专家", "")).strip() or str(r.get("确认专家", "")).strip())
    attr_ratio = with_attribution / max(total, 1)
    attr_score = round(5 * attr_ratio, 1)

    score = min(source_score + attr_score, 15)

    return {
        "score": score,
        "max_score": 15,
        "details": {
            "source_coverage": {"score": source_score, "max": 10, "ratio": round(source_ratio, 2), "count": with_source},
            "attribution": {"score": attr_score, "max": 5, "ratio": round(attr_ratio, 2), "count": with_attribution},
        }
    }


def get_grade(total_score: float) -> str:
    """根据总分判定等级"""
    if total_score >= 90:
        return "A"
    elif total_score >= 75:
        return "B"
    elif total_score >= 60:
        return "C"
    else:
        return "D"


def _compute_tacitness(records: list) -> tuple[float, float]:
    """计算隐性度指标：tacit_ratio 与 case_derived_ratio。"""
    total = max(len(records), 1)
    with_tacit = sum(
        1 for r in records
        if str(r.get("经验判断", "")).strip()
        or str(r.get("适用边界", "")).strip()
        or str(r.get("例外情形", "")).strip()
    )
    tacit_ratio = round(with_tacit / total, 3)
    with_case = sum(
        1 for r in records
        if "案例复盘" in str(r.get("来源文档", "") + r.get("来源", ""))
        or str(r.get("证据数", "")).strip().isdigit() and int(str(r.get("证据数", "")).strip()) > 0
    )
    case_ratio = round(with_case / total, 3)
    return tacit_ratio, case_ratio


def _compute_evidence_strength(records: list) -> dict:
    """计算证据强度统计。"""
    total = max(len(records), 1)
    evidence_counts = []
    break_counts = []
    for r in records:
        ev = str(r.get("证据数", "")).strip()
        br = str(r.get("突破数", "")).strip()
        evidence_counts.append(int(ev) if ev.isdigit() else 0)
        break_counts.append(int(br) if br.isdigit() else 0)
    avg_evidence = round(sum(evidence_counts) / total, 1)
    avg_breaks = round(sum(break_counts) / total, 1)
    with_breaks = sum(1 for b in break_counts if b > 0)
    return {
        "avg_evidence": avg_evidence,
        "avg_breaks": avg_breaks,
        "items_with_breaks": with_breaks,
        "total_items": total,
    }


def generate_report(records: list, config: dict, scores: dict, total_score: float, grade: str,
                    tacit_ratio: float = 0, case_ratio: float = 0,
                    evidence_stats: dict = None) -> str:
    """生成 Markdown 质量报告"""
    scenario = config.get("display_name", "未命名场景")
    total_items = len(records)

    lines = [
        f"# 知识萃取质量报告",
        "",
        f"- **场景**：{scenario}",
        f"- **知识条目数**：{total_items}",
        f"- **总分**：{total_score:.1f} / 100",
        f"- **等级**：{grade}",
        "",
        "## 各维度得分",
        "",
        "| 维度 | 得分 | 满分 | 得分率 |",
        "|------|------|------|--------|",
    ]

    for dim_name, dim_data in scores.items():
        score = dim_data["score"]
        max_score = dim_data["max_score"]
        rate = round(score / max_score * 100, 1) if max_score > 0 else 0
        lines.append(f"| {dim_name} | {score:.1f} | {max_score} | {rate}% |")

    lines.append(f"| **总计** | **{total_score:.1f}** | **100** | **{total_score:.0f}%** |")
    lines.append("")

    # 隐性化成效
    if evidence_stats:
        lines.append("## 隐性化成效")
        lines.append("")
        lines.append(f"- **隐性度（tacit_ratio）**：{tacit_ratio * 100:.1f}%（有专家经验判断/适用边界/例外情形的条目占比）")
        lines.append(f"- **案例衍生占比**：{case_ratio * 100:.1f}%（来自案例复盘或已被案例支撑的条目占比）")
        if evidence_stats:
            lines.append(f"- **平均证据数**：{evidence_stats['avg_evidence']}（每条知识被多少案例支撑）")
            lines.append(f"- **被突破条目**：{evidence_stats['items_with_breaks']}/{evidence_stats['total_items']}（专家实践中曾被突破的规则数）")
        lines.append("")

    # 改进建议
    lines.append("## 改进建议")
    lines.append("")

    for dim_name, dim_data in scores.items():
        score = dim_data["score"]
        max_score = dim_data["max_score"]
        if score < max_score * 0.6:
            lines.append(f"- **{dim_name}**（得分率 {round(score/max_score*100):.0f}%）：")
            details = dim_data.get("details", {})
            for sub_name, sub_data in details.items():
                if sub_data["score"] < sub_data["max"] * 0.5:
                    lines.append(f"  - {sub_name}：当前 {sub_data['score']}/{sub_data['max']}，需重点提升")
            lines.append("")

    return "\n".join(lines)


def quality_report(excel_path: str, config_path: str) -> dict:
    """主评分流程"""
    records, config_info = read_excel_data(excel_path)
    if not records:
        return {"status": "error", "message": f"No data found in {excel_path}"}

    config = load_config(config_path)

    scores = {
        "完整性": score_completeness(records, config),
        "准确性": score_accuracy(records, config),
        "可操作性": score_operability(records),
        "反模式覆盖": score_anti_pattern(records),
        "来源可溯": score_traceability(records),
    }

    total_score = sum(s["score"] for s in scores.values())

    # 隐性化成效（展示但不纳入总分，避免初创场景负激励）
    tacit_ratio, case_ratio = _compute_tacitness(records)
    evidence_stats = _compute_evidence_strength(records)

    grade = get_grade(total_score)

    report_md = generate_report(records, config, scores, total_score, grade,
                                tacit_ratio, case_ratio, evidence_stats)

    return {
        "status": "ok",
        "total_score": round(total_score, 1),
        "grade": grade,
        "scores": scores,
        "tacit_ratio": tacit_ratio,
        "case_derived_ratio": case_ratio,
        "evidence_strength": evidence_stats,
        "report_markdown": report_md,
    }


def main():
    parser = argparse.ArgumentParser(description="知识萃取质量评分报告生成")
    parser.add_argument("--input", required=True, help="输入 Excel 文件路径")
    parser.add_argument("--config", default="", help="场景 YAML 配置文件路径（可选）")
    parser.add_argument("--output", default="", help="Markdown 报告输出路径（可选）")
    args = parser.parse_args()

    try:
        result = quality_report(args.input, args.config)
    except FileNotFoundError as e:
        print(json.dumps({"status": "error", "message": f"文件不存在: {str(e)}"}, ensure_ascii=False))
        sys.exit(1)
    except Exception as e:
        import traceback
        print(json.dumps({"status": "error", "message": str(e), "detail": traceback.format_exc()}, ensure_ascii=False))
        sys.exit(1)

    if args.output and result.get("status") == "success":
        output_dir = os.path.dirname(args.output)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(result["report_markdown"])

    # stdout 只输出 JSON 摘要（不含 report_markdown，避免过长）
    stdout_result = {k: v for k, v in result.items() if k != "report_markdown"}
    print(json.dumps(stdout_result, ensure_ascii=False))


if __name__ == "__main__":
    main()
