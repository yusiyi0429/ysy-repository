#!/usr/bin/env python3
"""Step2 知识萃取：将 LLM 条目写入 Step1 场景模板（或标准表），仅输出 Excel。"""

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from step1_template import (
    _set_cell_value,
    detect_anchor_block_height,
    detect_data_start_row,
    detect_header_rows,
    find_anchor_columns,
)

# 英文字段 / 中文字段 → 模板列名候选（子串匹配）
FIELD_COLUMN_ALIASES = {
    "category": ("知识类型", "知识分类", "分类", "类别", "环节", "步骤", "category"),
    "content": ("具体方法", "知识描述", "知识内容", "知识引用", "方法", "content"),
    "trigger_condition": ("访谈方向", "适用条件", "触发条件", "条件", "关键输出-名称", "名称", "trigger_condition"),
    "judgment_logic": ("判断逻辑", "决策逻辑", "规则引用", "judgment_logic"),
    "anti_pattern": ("反模式", "踩坑", "反模式/踩坑提示", "关键输出-描述", "描述", "anti_pattern"),
    "source": ("来源", "来源文档", "source"),
    "confidence": ("置信度", "confidence"),
    "excerpt": ("原文摘录", "原文", "摘录", "知识引用"),
}

STANDARD_HEADERS = [
    "知识编号", "知识分类", "知识描述", "适用条件",
    "判断逻辑", "反模式/踩坑提示", "来源", "置信度",
]


def _norm(val):
    if val is None:
        return ""
    return str(val).strip()


def normalize_item(item: dict) -> dict:
    """统一中英文 LLM 字段名。"""
    if not isinstance(item, dict):
        return {}
    return {
        "category": _norm(
            item.get("category")
            or item.get("知识分类")
            or item.get("分类")
        ),
        "content": _norm(
            item.get("content")
            or item.get("知识描述")
            or item.get("知识内容")
        ),
        "trigger_condition": _norm(
            item.get("trigger_condition")
            or item.get("适用条件")
            or item.get("触发条件")
        ),
        "judgment_logic": _norm(
            item.get("judgment_logic")
            or item.get("判断逻辑")
        ),
        "anti_pattern": _norm(
            item.get("anti_pattern")
            or item.get("反模式/踩坑提示")
            or item.get("反模式")
        ),
        "source": _norm(item.get("source") or item.get("来源") or item.get("来源文档")),
        "confidence": _norm(item.get("confidence") or item.get("置信度")),
        "excerpt": _norm(item.get("excerpt") or item.get("原文摘录") or item.get("原文")),
    }


def build_header_map(ws):
    """解析第 1–2 行表头 → {表头文本: 列号}。"""
    header_map = {}
    rows = detect_header_rows(ws)
    for r in range(1, rows + 1):
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                key = _norm(v)
                if key:
                    header_map[key] = c
    return header_map


def _match_alias(header: str, aliases: tuple) -> bool:
    h = _norm(header)
    if not h:
        return False
    for alias in aliases:
        if h == alias or alias in h:
            return True
    return False


def _norm_key(text: str) -> str:
    return _norm(text).replace(" ", "").lower()


def _candidate_item_keys(key: str) -> list[str]:
    k = _norm(key)
    if not k:
        return []
    out = [k]
    # Support "关键输出-名称"/"关键输出：名称" style keys.
    for sep in ("-", "：", ":", "/", "\\"):
        if sep in k:
            out.append(k.split(sep)[-1].strip())
    return [x for x in out if x]


def _resolve_column_by_item_key(header_map: dict, item_key: str):
    key_norm = _norm_key(item_key)
    if not key_norm:
        return None
    for h, c in header_map.items():
        if _norm_key(h) == key_norm:
            return c
    return None


def resolve_field_columns(header_map: dict) -> dict:
    """字段名 → 列号；未匹配则为 None。"""
    cols = {}
    for field, aliases in FIELD_COLUMN_ALIASES.items():
        col = None
        for header, c in header_map.items():
            if _match_alias(header, aliases):
                col = c
                break
        cols[field] = col
    return cols


def _count_anchor_blocks(ws, data_start: int, block_height: int) -> int:
    total_rows = max(0, (ws.max_row or data_start) - data_start + 1)
    if block_height < 1:
        block_height = 1
    return max(1, (total_rows + block_height - 1) // block_height)


def _read_effective_cell_value(ws, row: int, col: int):
    """Read cell value, falling back to merged top-left value when needed."""
    v = ws.cell(row, col).value
    if v is not None:
        return v
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col <= col <= merged.max_col
        ):
            return ws.cell(merged.min_row, merged.min_col).value
    return None


def _unmerge_anchor_ranges(ws, anchor_cols: list, data_start: int):
    """Remove existing anchor merges before rebuilding smart merges."""
    anchor_set = set(anchor_cols)
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_col == merged.max_col
            and merged.min_col in anchor_set
            and merged.max_row >= data_start
        ):
            ws.unmerge_cells(str(merged))


def _smart_merge_anchor_columns(ws, anchor_cols: list, data_start: int, used_rows: int):
    """
    Smart-merge anchor columns based on contiguous effective values.
    This aligns the left 4 anchor columns with actually filled extraction rows.
    """
    if used_rows <= 0:
        return

    end_row = data_start + used_rows - 1
    _unmerge_anchor_ranges(ws, anchor_cols, data_start)

    for col in anchor_cols:
        values = []
        last_non_empty = None
        for r in range(data_start, end_row + 1):
            val = _read_effective_cell_value(ws, r, col)
            if _norm(val):
                last_non_empty = val
            elif last_non_empty is not None:
                # Propagate anchor values to newly duplicated/blank blocks.
                val = last_non_empty
            values.append(val)

        if not any(_norm(v) for v in values):
            continue

        group_start = data_start
        group_val = values[0]
        for idx in range(1, len(values) + 1):
            is_boundary = idx == len(values) or _norm(values[idx]) != _norm(group_val)
            if not is_boundary:
                continue

            group_end = data_start + idx - 1
            top_val = group_val if _norm(group_val) else None
            _set_cell_value(ws, group_start, col, top_val)
            for rr in range(group_start + 1, group_end + 1):
                _set_cell_value(ws, rr, col, None)
            if top_val is not None and group_end > group_start:
                ws.merge_cells(
                    start_row=group_start,
                    start_column=col,
                    end_row=group_end,
                    end_column=col,
                )

            if idx < len(values):
                group_start = data_start + idx
                group_val = values[idx]


def _ensure_blocks(ws, data_start: int, block_height: int, needed_blocks: int, anchor_cols: list):
    """条目超出已有子场景块时，在末尾复制最后一个块的版式。"""
    from step1_template import _duplicate_block

    current = _count_anchor_blocks(ws, data_start, block_height)
    max_col = ws.max_column or 1
    while current < needed_blocks:
        block_start = data_start + (current - 1) * block_height
        _duplicate_block(ws, block_start, block_height, anchor_cols, max_col)
        current += 1


def _fill_into_template_data(ws, items: list, pipeline_id: str = "") -> int:
    """在 Step1 模板数据区按行填入知识，返回写入行数。"""
    anchor = find_anchor_columns(ws)
    if len(anchor) < 4:
        return 0

    anchor_cols = sorted(anchor.values())
    data_start = detect_data_start_row(ws)
    block_height = detect_anchor_block_height(ws, data_start, anchor_cols)
    header_map = build_header_map(ws)
    field_cols = resolve_field_columns(header_map)

    # 找到知识编号列（若有）
    id_col = None
    for h, c in header_map.items():
        if _match_alias(h, ("知识编号", "knowledge_id")):
            id_col = c
            break
    prefix = pipeline_id[:8] if pipeline_id else "KN"

    if not any(field_cols.values()):
        return 0

    needed_blocks = max(1, (len(items) + block_height - 1) // block_height)
    _ensure_blocks(ws, data_start, block_height, needed_blocks, anchor_cols)

    written = 0
    for idx, raw in enumerate(items):
        item = normalize_item(raw)
        raw_item = raw if isinstance(raw, dict) else {}
        block_idx = idx // block_height
        offset = idx % block_height
        row = data_start + block_idx * block_height + offset
        row_written = False

        # 自动写入知识编号
        if id_col:
            _set_cell_value(ws, row, id_col, f"{prefix}-{row:04d}")

        # 1) Canonical field mapping (category/content/...).
        for field, col in field_cols.items():
            if not col:
                continue
            val = item.get(field, "")
            if val:
                _set_cell_value(ws, row, col, val)
                row_written = True

        # 2) Raw-key mapping for template-specific headers.
        for raw_key, raw_val in raw_item.items():
            val = _norm(raw_val)
            if not val:
                continue
            for candidate in _candidate_item_keys(raw_key):
                col = _resolve_column_by_item_key(header_map, candidate)
                if col:
                    _set_cell_value(ws, row, col, val)
                    row_written = True
                    break

        if row_written:
            written += 1

    # Smart-merge anchor columns based on actual extraction row span.
    _smart_merge_anchor_columns(ws, anchor_cols, data_start, len(items))
    return written


def _append_standard_rows(ws, items: list, pipeline_id: str = "") -> int:
    """无模板锚定列时，在表尾追加标准列行。"""
    header_map = build_header_map(ws)
    field_cols = resolve_field_columns(header_map)

    # 若表头不含任何知识列，写入标准表头
    if not any(field_cols.values()):
        if ws.max_row < 1 or not ws.cell(1, 1).value:
            for col, h in enumerate(STANDARD_HEADERS, 1):
                ws.cell(1, col, value=h)
        header_map = build_header_map(ws)
        field_cols = resolve_field_columns(header_map)

    col_id_num = None
    for h, c in header_map.items():
        if _match_alias(h, ("知识编号", "knowledge_id")):
            col_id_num = c
            break

    prefix = pipeline_id[:8] if pipeline_id else "KN"
    start_row = (ws.max_row or 1) + 1
    for idx, raw in enumerate(items):
        item = normalize_item(raw)
        row = start_row + idx
        if col_id_num:
            _set_cell_value(ws, row, col_id_num, f"{prefix}-{row:04d}")
        for field, col in field_cols.items():
            if col and item.get(field):
                _set_cell_value(ws, row, col, item[field])
    return len(items)


def _style_standard_sheet(ws):
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="C7000B", end_color="C7000B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col in range(1, len(STANDARD_HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
    ws.freeze_panes = "A2"


def write_preextract_excel(
    *,
    step1_path: Path | str | None,
    output_path: Path | str,
    items: list,
    pipeline_id: str = "",
) -> dict:
    """
    生成萃取 Excel。
    返回: {path, count, filled_rows, used_step1_template}
    """
    output_path = Path(output_path)
    items = items or []
    used_template = False
    filled_rows = 0

    if step1_path and Path(step1_path).exists():
        shutil.copy2(step1_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb[wb.sheetnames[0]]
        used_template = True
        filled_rows = _fill_into_template_data(ws, items, pipeline_id)
        if filled_rows == 0 and items:
            filled_rows = _append_standard_rows(ws, items, pipeline_id)
        wb.save(output_path)
        wb.close()
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "知识萃取"
        for col, h in enumerate(STANDARD_HEADERS, 1):
            ws.cell(1, col, value=h)
        _style_standard_sheet(ws)
        filled_rows = _append_standard_rows(ws, items, pipeline_id)
        wb.save(output_path)
        wb.close()

    return {
        "path": str(output_path),
        "count": len(items),
        "filled_rows": filled_rows,
        "used_step1_template": used_template,
    }
