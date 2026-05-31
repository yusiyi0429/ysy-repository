"""Excel workbook layout helpers for revision row alignment."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from step1_template import detect_data_start_row, detect_header_rows


def get_sheet_layout(ws) -> dict:
    """Return header/data row boundaries for a worksheet."""
    header_rows = detect_header_rows(ws)
    data_start_row = detect_data_start_row(ws)
    return {
        "header_rows": header_rows,
        "data_start_row": data_start_row,
        "max_row": ws.max_row or 0,
        "max_column": ws.max_column or 0,
    }


def _row_cells(ws, row_idx: int, max_col: int | None = None) -> dict:
    max_col = max_col or (ws.max_column or 1)
    cells = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row_idx, c).value
        if v is not None and str(v).strip():
            cells[f"col{c}"] = str(v).strip()
    return cells


def build_revision_context(excel_path: str, *, max_rows: int = 45) -> tuple[str, dict]:
    """
    Build LLM context with explicit excel_row numbering.
    Returns (summary_text, layouts_by_sheet).
    """
    path = Path(excel_path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    layouts: dict = {}
    sheet_blocks = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        layout = get_sheet_layout(ws)
        layouts[ws_name] = layout
        header_rows = layout["header_rows"]
        data_start = layout["data_start_row"]
        preview_rows = []
        limit = min(ws.max_row or 0, max_rows)
        for r in range(1, limit + 1):
            is_header = r < data_start
            data_row = (r - data_start + 1) if r >= data_start else None
            preview_rows.append({
                "excel_row": r,
                "data_row": data_row,
                "is_header": is_header,
                "cells": _row_cells(ws, r),
            })
        sheet_blocks.append({
            "sheet": ws_name,
            "header_rows": header_rows,
            "data_start_row": data_start,
            "rows": preview_rows,
        })

    wb.close()

    summary = (
        "行号约定：JSON 中的 row 必须使用下列预览里的 excel_row（Excel 物理行号，含表头）。\n"
        "表头行号 < data_start_row；首条数据行的 excel_row 等于 data_start_row。\n"
        "若你按「数据行序号」（不含表头，首条数据=1）理解，系统会自动换算，但优先使用 excel_row。\n"
    )
    ctx_json = json.dumps(sheet_blocks, ensure_ascii=False, indent=2)
    if len(ctx_json) > 7000:
        ctx_json = ctx_json[:7000] + "\n...（已截断）"
    return summary + "\n" + ctx_json, layouts


def _to_row_int(val, default: int = 0) -> int:
    if isinstance(val, int):
        return val
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default
    if s.isdigit():
        return int(s)
    digits = []
    for ch in s:
        if ch.isdigit():
            digits.append(ch)
        elif digits:
            break
    return int("".join(digits)) if digits else default


def normalize_revision_notes(expert_notes: list, layouts: dict) -> tuple[list, dict]:
    """
    Map data-relative row numbers to excel_row; drop header-targeting edits.
    Returns (notes, stats).
    """
    stats = {"adjusted": 0, "skipped_header": 0, "skipped_invalid": 0}
    out = []

    for raw in expert_notes or []:
        if not isinstance(raw, dict):
            stats["skipped_invalid"] += 1
            continue
        note = dict(raw)
        sheet = note.get("sheet") or "Sheet1"
        layout = layouts.get(sheet)
        if not layout:
            out.append(note)
            continue

        data_start = layout["data_start_row"]
        row = _to_row_int(note.get("row"), 0)
        action = str(note.get("action", "")).lower()

        if row < 1:
            if action == "add":
                row = data_start
                note["row"] = row
                stats["adjusted"] += 1
            else:
                stats["skipped_invalid"] += 1
                continue

        # Data-relative convention: row in 1..(data_start-1) → offset to physical data rows
        if 1 <= row < data_start:
            new_row = data_start + row - 1
            if new_row != row:
                note["row"] = new_row
                note["_row_convention"] = "data_relative"
                stats["adjusted"] += 1
                row = new_row

        # Do not modify header rows (add may insert at data_start+)
        if action != "add" and row < data_start:
            stats["skipped_header"] += 1
            continue

        if action != "add" and row > layout["max_row"]:
            row = layout["max_row"]
            note["row"] = row

        out.append(note)

    return out, stats


def layout_prompt_rules() -> str:
    return (
        "行号规则（必须遵守）：\n"
        "1. row 字段 = 预览 JSON 中的 excel_row（Excel 物理行号，含表头）。\n"
        "2. 禁止修改 is_header=true 的行（表头区）。\n"
        "3. add 操作必须给出明确 row（excel_row）与 col（1-based，A=1）。\n"
        "4. sheet 名称必须与预览完全一致。\n"
    )
