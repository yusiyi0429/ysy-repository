#!/usr/bin/env python3
"""根据 scenario-schema 生成 Step1 萃取模板 Excel（无 LLM）。"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from scenario_schema import (
    ANCHOR_COLUMNS,
    load_scenario_schema,
    resolve_knowledge_columns,
    resolve_knowledge_columns_for_request,
    schema_summary,
)
from step1_template import fill_scenario_skeleton


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
DEFAULT_BLOCK_HEIGHT = 4


def create_schema_template_workbook(
    output_path: Path | str,
    schema_path: Path | str,
    *,
    sheet_title: str = "知识萃取",
    knowledge_columns: list[str] | None = None,
) -> dict:
    """
    生成含锚定四列 + schema 知识列的空模板，并返回结构摘要。
    锚定列在数据区纵向合并（默认 4 行一块），与 fill_template_sheet 行为一致。
    """
    schema = load_scenario_schema(schema_path)
    summary = schema_summary(schema, schema_path=schema_path)
    anchor = list(ANCHOR_COLUMNS)
    knowledge = knowledge_columns or resolve_knowledge_columns(schema)
    if not knowledge:
        raise ValueError("请至少定义一列知识字段")
    headers = anchor + knowledge
    summary["knowledge_columns"] = knowledge
    summary["column_count"] = len(headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_title or summary.get("display_name") or "知识萃取")[:31]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(24, len(header) * 2))

    data_start = 2
    block_end = data_start + DEFAULT_BLOCK_HEIGHT - 1
    for col_idx in range(1, len(anchor) + 1):
        ws.merge_cells(
            start_row=data_start,
            start_column=col_idx,
            end_row=block_end,
            end_column=col_idx,
        )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = ws.title
    wb.save(out)
    wb.close()

    summary["sheet"] = sheet_name
    summary["headers"] = headers
    summary["block_height"] = DEFAULT_BLOCK_HEIGHT
    return summary


def generate_skeleton_from_schema(
    schema_path: Path | str,
    output_path: Path | str,
    scenario_name: str,
    scenario_content: str,
    sub_scenarios: list,
    *,
    knowledge_columns: list[str] | None = None,
) -> dict:
    """先按 schema/自定义列生成模板，再填入场景锚定信息，返回 fill_scenario_skeleton 结果 + schema 元数据。"""
    schema_path = Path(schema_path)
    output_path = Path(output_path)
    schema = load_scenario_schema(schema_path)
    cols = resolve_knowledge_columns_for_request(schema, knowledge_columns)
    if not cols:
        raise ValueError("请至少定义一列知识字段")

    temp_tpl = output_path.with_name(f"_schema_tpl_{output_path.stem}.xlsx")

    meta = create_schema_template_workbook(temp_tpl, schema_path, knowledge_columns=cols)
    try:
        fill_result = fill_scenario_skeleton(
            str(temp_tpl),
            str(output_path),
            scenario_name,
            scenario_content,
            sub_scenarios,
        )
    finally:
        if temp_tpl.exists():
            try:
                temp_tpl.unlink()
            except OSError:
                pass

    summary = schema_summary(schema, schema_path=schema_path)
    summary["knowledge_columns"] = cols
    summary["column_count"] = len(ANCHOR_COLUMNS) + len(cols)
    fill_result["schema"] = summary
    fill_result["template_headers"] = meta.get("headers", [])
    fill_result["knowledge_columns"] = cols
    return fill_result
