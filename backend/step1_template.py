#!/usr/bin/env python3
"""Step1 场景锚定：将四项输入填入萃取模板前四列，保留表头与版式，数据区不保留参考列内容。"""

import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

ANCHOR_HEADERS = {
    "scenario": ("场景", "场景名称", "场景信息"),
    "scenario_desc": ("场景说明", "场景内容"),
    "sub_scenario": ("子场景", "子场景名称"),
    "sub_scenario_desc": ("子场景说明", "子场景内容"),
}

# 数据区不保留模板示例值的列（表头行 1–2 仍保留列名）
REFERENCE_HEADER_MARKERS = (
    "环节", "步骤", "访谈方向", "具体方法", "知识类型", "知识引用",
    "规则引用", "专业术语", "关键输出", "名称", "描述",
)


def _norm(val):
    if val is None:
        return ""
    return str(val).strip()


def _match_header(cell_val, aliases):
    v = _norm(cell_val)
    if not v:
        return False
    for alias in aliases:
        if v == alias or alias in v:
            return True
    return False


def normalize_sub_scenarios(sub_scenarios):
    """子场景名称与说明均为可选；仅保留至少有一项非空的条目。"""
    out = []
    for sub in sub_scenarios or []:
        if not isinstance(sub, dict):
            continue
        name = _norm(sub.get("name", ""))
        content = _norm(sub.get("content", ""))
        if name or content:
            out.append({"name": name, "content": content})
    return out


def anchor_has_required_columns(anchor):
    """Step1 至少需「场景」列；场景说明、子场景、子场景说明均可缺。"""
    return bool(anchor) and "scenario" in anchor


def find_anchor_columns(ws):
    """从第 1 行表头解析锚定四列列号。"""
    cols = {}
    for c in range(1, ws.max_column + 1):
        header = ws.cell(1, c).value
        if not header:
            continue
        for key, aliases in ANCHOR_HEADERS.items():
            if key not in cols and _match_header(header, aliases):
                cols[key] = c
    return cols


def detect_header_rows(ws):
    """表头区行数：锚定四列在首行，且存在 A1:A2 类纵向合并时为 2 行。"""
    merged_ranges = getattr(ws, "merged_cells", None)
    if merged_ranges is not None:
        for merged in merged_ranges.ranges:
            if merged.min_row == 1 and merged.max_row == 2 and merged.min_col == 1 and merged.max_col == 1:
                return 2
    # read_only 工作表：根据第 2 行首列是否为空推断
    if ws.cell(2, 1).value is None and ws.cell(1, 1).value and _match_header(ws.cell(1, 1).value, ANCHOR_HEADERS["scenario"]):
        return 2
    return 1


def detect_data_start_row(ws):
    """数据区从表头下一行开始；双行表头模板数据从第 3 行起。"""
    return detect_header_rows(ws) + 1


def detect_anchor_block_height(ws, data_start, anchor_cols):
    """锚定四列在数据区的纵向合并高度，即每个子场景占用的行数（默认 4）。"""
    min_col = min(anchor_cols)
    max_col = max(anchor_cols)
    heights = []
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row >= data_start
            and merged.min_col >= min_col
            and merged.max_col <= max_col
            and merged.min_row == data_start
        ):
            heights.append(merged.max_row - merged.min_row + 1)
    if heights:
        return max(heights)
    return 1


def _writable_cell(ws, row, col):
    """合并单元格仅写入左上角。"""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(merged.min_row, merged.min_col)
    return cell


def _set_cell_value(ws, row, col, value):
    _writable_cell(ws, row, col).value = value


def _copy_cell_style(src, dst):
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def find_reference_columns(ws, anchor_cols):
    """需清空数据区内容的列（锚定四列以外的参考/环节类列）。"""
    anchor_set = set(anchor_cols)
    ref_cols = set()
    header_rows = detect_header_rows(ws)
    for c in range(1, ws.max_column + 1):
        if c in anchor_set:
            continue
        h1 = _norm(ws.cell(1, c).value)
        h2 = _norm(ws.cell(2, c).value) if header_rows >= 2 else ""
        matched = any(
            m in h1 or m in h2 or h1 == m or h2 == m for m in REFERENCE_HEADER_MARKERS
        )
        if matched or c > max(anchor_set):
            ref_cols.add(c)
    return ref_cols


def clear_reference_columns(ws, start_row, end_row, ref_cols):
    """清空数据区参考列单元格（仅清值，保留表头与样式）。"""
    for r in range(start_row, end_row + 1):
        for c in ref_cols:
            _set_cell_value(ws, r, c, None)


def _copy_row_styles(ws, src_row, dst_row, max_col, anchor_cols, copy_values_for_anchor=False):
    anchor_set = set(anchor_cols)
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if c in anchor_set and copy_values_for_anchor:
            dst.value = src.value
        else:
            dst.value = None
        _copy_cell_style(src, dst)


def _duplicate_block(ws, block_start, block_height, anchor_cols, max_col):
    """在块后插入等高的新块，仅复制版式，参考列不复制内容。"""
    insert_at = block_start + block_height
    ws.insert_rows(insert_at, block_height)
    for offset in range(block_height):
        _copy_row_styles(
            ws, block_start + offset, insert_at + offset, max_col, anchor_cols, copy_values_for_anchor=False
        )
    # 为新块的前四列重建纵向合并（与模板首块一致）
    anchor = find_anchor_columns(ws)
    if len(anchor) == 4:
        anchor_cols = sorted(anchor.values())
        block_end = insert_at + block_height - 1
        for col in anchor_cols:
            ws.merge_cells(
                start_row=insert_at,
                start_column=col,
                end_row=block_end,
                end_column=col,
            )


def _fill_anchor_block(ws, block_start, anchor, scenario_name, scenario_content, sub):
    """写入锚定块首行；子场景相关列存在时才写入。"""
    _set_cell_value(ws, block_start, anchor["scenario"], scenario_name)
    if "scenario_desc" in anchor:
        _set_cell_value(ws, block_start, anchor["scenario_desc"], scenario_content)
    if "sub_scenario" in anchor:
        _set_cell_value(ws, block_start, anchor["sub_scenario"], sub.get("name", ""))
    if "sub_scenario_desc" in anchor:
        _set_cell_value(ws, block_start, anchor["sub_scenario_desc"], sub.get("content", ""))


def fill_template_sheet(ws, scenario_name, scenario_content, sub_scenarios):
    """填充单个工作表，返回写入的子场景块数（无子场景内容时为 1 个空块）。"""
    anchor = find_anchor_columns(ws)
    if not anchor_has_required_columns(anchor):
        return 0

    anchor_col_list = list(anchor.values())
    data_start = detect_data_start_row(ws)
    block_height = detect_anchor_block_height(ws, data_start, anchor_col_list)
    max_col = ws.max_column
    ref_cols = find_reference_columns(ws, anchor_col_list)
    subs = normalize_sub_scenarios(sub_scenarios)
    if not subs:
        subs = [{"name": "", "content": ""}]

    # 清空模板数据区中的参考列示例（环节、具体方法、知识引用等）
    clear_reference_columns(ws, data_start, ws.max_row, ref_cols)

    for i in range(1, len(subs)):
        _duplicate_block(ws, data_start, block_height, anchor_col_list, max_col)

    total_end = data_start + len(subs) * block_height - 1
    clear_reference_columns(ws, data_start, total_end, ref_cols)

    for i, sub in enumerate(subs):
        block_start = data_start + i * block_height
        _fill_anchor_block(ws, block_start, anchor, scenario_name, scenario_content, sub)

    return len(subs)


def list_default_step1_templates(samples_dir: Path):
    """列出可用的默认 Step1 模板（按文件名排序）。"""
    if not samples_dir.exists():
        return []
    candidates = []
    for path in sorted(samples_dir.glob("*.xlsx")):
        if path.name.startswith("~$") or path.name.startswith("_"):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            h1 = ws.cell(1, 1).value
            wb.close()
            if h1 and _match_header(h1, ANCHOR_HEADERS["scenario"]):
                candidates.append(path)
        except Exception:
            continue
    return candidates


def find_default_step1_template(samples_dir: Path):
    """优先使用「测试模板」，排除临时/输出文件。"""
    candidates = list_default_step1_templates(samples_dir)
    for path in candidates:
        if "测试" in path.stem or "测试" in path.name:
            return path
    return candidates[0] if candidates else None


def fill_template_into_workbook(wb, scenario_name, scenario_content, sub_scenarios):
    """填充所有含「场景」锚定列的工作表（子场景列可选）。"""
    filled = []
    for ws in wb.worksheets:
        count = fill_template_sheet(ws, scenario_name, scenario_content, sub_scenarios)
        if count > 0:
            filled.append(ws.title)
    if not filled:
        raise ValueError(
            "模板中未找到「场景」锚定列，请使用含场景/场景说明的萃取模板（子场景、子场景说明可选）"
        )
    return filled[0]


def fill_scenario_skeleton(template_path, output_path, scenario_name, scenario_content, sub_scenarios):
    """复制模板并填入场景锚定四项。"""
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    sheet_title = fill_template_into_workbook(wb, scenario_name, scenario_content, sub_scenarios)
    wb.save(output_path)
    wb.close()

    fields_info = []
    wb_read = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    for sn in wb_read.sheetnames:
        sws = wb_read[sn]
        headers = [
            sws.cell(1, c).value
            for c in range(1, sws.max_column + 1)
            if sws.cell(1, c).value
        ]
        if headers:
            fields_info.append({
                "sheet": sn,
                "headers": headers,
                "data_rows": max(0, sws.max_row - detect_header_rows(sws)),
            })
    wb_read.close()

    return {
        "sheet": sheet_title,
        "fields_info": fields_info,
        "sub_scenario_count": len(normalize_sub_scenarios(sub_scenarios)),
    }
