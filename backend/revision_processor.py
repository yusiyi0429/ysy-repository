#!/usr/bin/env python3
"""
知识修订处理器
对比专家会议纪要与Excel知识初稿，生成带标注的修订稿
"""

import argparse
import json
import os
import shutil
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 颜色配置
COLOR_MAP = {
    'modify': 'FFFF00',      # 黄色 - 修改
    'delete': 'FF6B6B',      # 红色 - 删除
    'add': '90EE90',         # 绿色 - 新增
    'supplement': '87CEEB',  # 蓝色 - 补充
    'header': '4472C4',      # 蓝色 - 表头
    'header_text': 'FFFFFF' # 白色 - 表头文字
}

# 修订状态映射
STATUS_MAP = {
    'modify': '修改',
    'delete': '删除',
    'add': '新增',
    'supplement': '补充'
}


REVISION_FONT_COLOR = 'FFFF0000'  # 红色（ARGB）


def load_expert_notes(expert_notes_str, notes_file):
    """加载专家纪要"""
    if expert_notes_str:
        try:
            return json.loads(expert_notes_str)
        except json.JSONDecodeError:
            return []
    elif notes_file:
        with open(notes_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return []


def create_backup(input_path):
    """创建备份文件"""
    backup_path = f"{input_path}.bak"
    if os.path.exists(input_path):
        shutil.copy2(input_path, backup_path)
    return backup_path


def create_revision_columns(ws, start_col, header_rows=1):
    """创建修订信息列（表头与模板表头行对齐，避免覆盖第 1 行场景锚定列名）。"""
    headers = ['修订状态', '原始内容', '修订内容', '修订说明', '修订时间']
    header_rows = max(1, int(header_rows or 1))
    for i, header in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color=COLOR_MAP['header_text'])
        cell.fill = PatternFill(start_color=COLOR_MAP['header'], end_color=COLOR_MAP['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if header_rows > 1:
            ws.merge_cells(
                start_row=1,
                start_column=col,
                end_row=header_rows,
                end_column=col,
            )
    return headers


def apply_cell_style(cell, revision_type):
    """应用修订单元格样式"""
    if revision_type in COLOR_MAP:
        bg_color = COLOR_MAP[revision_type]
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    # 所有修订点统一使用红色字体，便于一眼识别
    font = copy(cell.font) if cell.font else Font()
    font.color = REVISION_FONT_COLOR

    if revision_type == 'delete':
        font.strike = True
    elif revision_type in ['add', 'modify']:
        font.bold = True

    cell.font = font


def _writable_cell(ws, row, col):
    """合并单元格只允许写左上角，返回可写单元格。"""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(merged.min_row, merged.min_col)
    return cell


def process_modify(ws, row, col, old_value, new_value, note, revision_col):
    """处理修改操作"""
    target_cell = _writable_cell(ws, row, col)
    original_value = target_cell.value
    
    # 更新单元格内容
    target_cell.value = new_value
    apply_cell_style(target_cell, 'modify')
    
    # 填写修订信息列
    ws.cell(row=row, column=revision_col, value='修改')
    ws.cell(row=row, column=revision_col + 1, value=str(original_value) if original_value else '')
    ws.cell(row=row, column=revision_col + 2, value=new_value)
    ws.cell(row=row, column=revision_col + 3, value=note)
    ws.cell(row=row, column=revision_col + 4, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
    
    # 应用修订列样式
    for i in range(5):
        apply_cell_style(ws.cell(row=row, column=revision_col + i), 'modify')


def process_delete(ws, row, col, old_value, note, revision_col):
    """处理删除操作"""
    target_cell = _writable_cell(ws, row, col)
    
    apply_cell_style(target_cell, 'delete')
    
    # 填写修订信息列
    ws.cell(row=row, column=revision_col, value='删除')
    ws.cell(row=row, column=revision_col + 1, value=target_cell.value)
    ws.cell(row=row, column=revision_col + 2, value='')
    ws.cell(row=row, column=revision_col + 3, value=note)
    ws.cell(row=row, column=revision_col + 4, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
    
    for i in range(5):
        apply_cell_style(ws.cell(row=row, column=revision_col + i), 'delete')


def process_add(ws, row, col, new_value, note, revision_col, sheet_max_row):
    """处理新增操作"""
    if col == 0:
        # 新增整行
        ws.insert_rows(row)
        values = new_value.split('|') if '|' in new_value else [new_value]
        for i, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=i, value=value.strip())
            apply_cell_style(cell, 'add')
            for j in range(5):
                cell = ws.cell(row=row, column=revision_col + j)
                if j == 0:
                    cell.value = '新增'
                elif j == 3:
                    cell.value = note
                elif j == 4:
                    cell.value = datetime.now().strftime('%Y-%m-%d %H:%M')
                apply_cell_style(cell, 'add')
    else:
        # 新增单元格内容
        target_cell = _writable_cell(ws, row, col)
        target_cell.value = new_value
        apply_cell_style(target_cell, 'add')
        
        ws.cell(row=row, column=revision_col, value='新增')
        ws.cell(row=row, column=revision_col + 1, value='')
        ws.cell(row=row, column=revision_col + 2, value=new_value)
        ws.cell(row=row, column=revision_col + 3, value=note)
        ws.cell(row=row, column=revision_col + 4, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
        
        for i in range(5):
            apply_cell_style(ws.cell(row=row, column=revision_col + i), 'add')


def process_supplement(ws, row, col, new_value, note, revision_col):
    """处理补充操作"""
    target_cell = _writable_cell(ws, row, col)
    original_value = target_cell.value or ''
    
    # 追加补充内容
    if original_value:
        target_cell.value = f"{original_value}\n【补充】{new_value}"
    else:
        target_cell.value = f"【补充】{new_value}"
    
    apply_cell_style(target_cell, 'supplement')
    
    # 填写修订信息列
    ws.cell(row=row, column=revision_col, value='补充')
    ws.cell(row=row, column=revision_col + 1, value='')
    ws.cell(row=row, column=revision_col + 2, value=new_value)
    ws.cell(row=row, column=revision_col + 3, value=note)
    ws.cell(row=row, column=revision_col + 4, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
    
    for i in range(5):
        apply_cell_style(ws.cell(row=row, column=revision_col + i), 'supplement')


def process_delete_row(ws, row, note, revision_col):
    """处理删除整行"""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        apply_cell_style(cell, 'delete')
    
    # 在修订列标记
    ws.cell(row=row, column=revision_col, value='删除整行')
    ws.cell(row=row, column=revision_col + 1, value='整行已删除')
    ws.cell(row=row, column=revision_col + 3, value=note)
    ws.cell(row=row, column=revision_col + 4, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
    
    for i in range(5):
        apply_cell_style(ws.cell(row=row, column=revision_col + i), 'delete')


def apply_revision(ws, action_data, revision_col):
    """应用单条修订"""
    action = action_data.get('action', '').lower()
    row = action_data.get('row', 1)
    col = action_data.get('col', 1)
    note = action_data.get('note', '')
    
    # For add actions, require explicit target row/col to avoid bottom pile-up.
    if action == 'add':
        if row is None or row < 1 or col is None or col < 1:
            return False
    else:
        # 验证行号和列号最小值（openpyxl要求≥1）
        if row is None or row < 1:
            row = ws.max_row + 1
        if col is None or col < 1:
            col = 1
    
    # 验证行号最大值
    if action != 'add' and row > ws.max_row:
        row = ws.max_row
    
    # 处理不同action类型
    if action == 'modify':
        process_modify(
            ws, row, col,
            action_data.get('old_value', ''),
            action_data.get('new_value', ''),
            note, revision_col
        )
    elif action == 'delete':
        process_delete(
            ws, row, col,
            action_data.get('old_value', ''),
            note, revision_col
        )
    elif action == 'add':
        process_add(
            ws, row, col,
            action_data.get('new_value', ''),
            note, revision_col, ws.max_row
        )
    elif action == 'supplement':
        process_supplement(
            ws, row, col,
            action_data.get('new_value', ''),
            note, revision_col
        )
    elif action == 'delete_row':
        process_delete_row(ws, row, note, revision_col)
    else:
        return False
    
    return True


def process_workbook(input_path, expert_notes, output_path, *, layouts=None):
    """处理整个工作簿。layouts: 可选，来自 workbook_layout.build_revision_context。"""
    wb = load_workbook(input_path)

    if layouts is None:
        try:
            from workbook_layout import get_sheet_layout, normalize_revision_notes

            built = {}
            for name in wb.sheetnames:
                built[name] = get_sheet_layout(wb[name])
            expert_notes, _ = normalize_revision_notes(expert_notes, built)
            layouts = built
        except Exception:
            layouts = {}

    # 按sheet分组处理
    notes_by_sheet = {}
    for note in expert_notes:
        sheet_name = note.get('sheet', 'Sheet1')
        if sheet_name not in notes_by_sheet:
            notes_by_sheet[sheet_name] = []
        notes_by_sheet[sheet_name].append(note)

    total_processed = 0

    for sheet_name, notes in notes_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # 确定修订列起始位置（插入到最后一列后）
        revision_col = ws.max_column + 1

        header_rows = 1
        if layouts and sheet_name in layouts:
            header_rows = layouts[sheet_name].get("header_rows", 1)

        # 创建修订信息列
        create_revision_columns(ws, revision_col, header_rows)

        # 按行号排序修订记录，从后往前处理以避免行号偏移问题
        sorted_notes = sorted(notes, key=lambda x: x.get('row', 1), reverse=True)

        # 应用每条修订
        for note in sorted_notes:
            action = note.get('action', '').lower()
            original_row = note.get('row', 1)

            # 从后往前处理时，insert_rows 只影响当前行及之后的行
            # 因为我们是从大到小处理，所以不需要调整行号
            adjusted_row = original_row

            # 更新note的行号为调整后值
            note['row'] = adjusted_row

            if apply_revision(ws, note, revision_col):
                total_processed += 1

    # 保存修订后的文件
    wb.save(output_path)

    return total_processed


def list_sheets(input_path):
    """列出所有工作表名称"""
    wb = load_workbook(input_path, read_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    return sheets


def main():
    parser = argparse.ArgumentParser(description='知识修订处理器 - 对比专家纪要生成标注版修订稿')
    parser.add_argument('--input', '-i', required=True, help='原始Excel文件路径')
    parser.add_argument('--expert-notes', '-e', help='专家纪要JSON字符串')
    parser.add_argument('--expert-notes-file', '-f', help='专家纪要JSON文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径（默认添加_rev后缀）')
    parser.add_argument('--list-sheets', '-l', action='store_true', help='列出工作表名称')
    parser.add_argument('--no-backup', '-n', action='store_true', help='不创建备份文件')

    args = parser.parse_args()

    # 验证输入文件
    if not os.path.exists(args.input):
        result = {"status": "error", "message": f"输入文件不存在: {args.input}"}
        print(json.dumps(result, ensure_ascii=False))
        return 1

    # 列出工作表
    if args.list_sheets:
        sheets = list_sheets(args.input)
        result = {"status": "ok", "sheets": sheets}
        print(json.dumps(result, ensure_ascii=False))
        return 0

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        name, ext = os.path.splitext(args.input)
        output_path = f"{name}_rev{ext}"

    # 加载专家纪要
    expert_notes = load_expert_notes(args.expert_notes, args.expert_notes_file)
    if not expert_notes:
        result = {"status": "error", "message": "未提供有效的专家纪要"}
        print(json.dumps(result, ensure_ascii=False))
        return 1

    # 创建备份
    if not args.no_backup:
        create_backup(args.input)

    # 处理文件
    try:
        count = process_workbook(args.input, expert_notes, output_path)

        # 输出JSON结果
        result = {
            "status": "ok",
            "input_file": args.input,
            "output_file": output_path,
            "revision_count": count
        }
        print(json.dumps(result, ensure_ascii=False))

    except Exception as e:
        import traceback
        result = {"status": "error", "message": str(e), "detail": traceback.format_exc()}
        print(json.dumps(result, ensure_ascii=False))
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
