#!/usr/bin/env python3
"""
Step 5: 智能转化 — 将确认版 Excel 转换为 SKILL.md

读取 Excel 知识条目，按场景 YAML 配置分类，生成结构化 SKILL.md 文件。
转换是确定性程序，不依赖模型再次发挥。
工程化增强：数据校验、版本管理、配置驱动的输出格式、详细质量指标。

支持两种 Excel 格式：
1. 传统格式：包含"知识条目" sheet，每行是一条知识
2. 场景步骤格式：每个场景一个 sheet，按步骤组织知识（步骤、数据规则、知识要点、专家经验等列）
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed"}, ensure_ascii=False))
    sys.exit(1)

try:
    import yaml
except ImportError:
    print(json.dumps({"status": "error", "message": "pyyaml not installed"}, ensure_ascii=False))
    sys.exit(1)

from field_aliases import FIELD_ALIASES as _BASE_FIELD_ALIASES, REVISION_COLUMN_MARKERS


# Known sheet names that are NOT knowledge-item sheets
NON_KNOWLEDGE_SHEETS = {"场景配置", "版本追踪", "配置", "config"}

# 记录未填子场景时的分组占位名，不用于 SKILL 正文分节标题
EMPTY_SUB_FALLBACK = "未分子场景"

CONTENT_HEADER_MARKERS = (
    "具体方法", "知识描述", "知识内容", "数据规则", "知识要点", "专家经验",
    "方法", "描述", "内容", "规则", "要点", "输出", "引用", "逻辑", "术语",
)

# Possible column names for knowledge fields — extends shared base with sheet-structural keys
FIELD_ALIASES = {
    **_BASE_FIELD_ALIASES,
    "场景说明": ["场景说明", "场景内容"],
    "子场景说明": ["子场景说明", "子场景内容"],
    "场景": ["场景"],
    "子场景": ["子场景"],
    "步骤": ["步骤"],
    "知识引用": ["知识引用"],
    "修订状态": ["修订状态", "修改"],
    "修订内容": ["修订内容"],
    "原始内容": ["原始内容"],
}


def _resolve_header(raw_header: str) -> str:
    """Resolve a raw Excel header to a canonical field name using aliases."""
    if not raw_header:
        return ""
    raw = str(raw_header).strip()
    first_line = raw.split("\n")[0].strip()
    if first_line.startswith("步骤"):
        return "步骤"
    # Prefer longer / more specific alias matches (e.g. 子场景 before 场景)
    pairs = []
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            pairs.append((len(alias), alias, canonical))
    for _, alias, canonical in sorted(pairs, reverse=True):
        if first_line == alias or raw == alias:
            return canonical
    return first_line


def _norm_cell(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_revision_column(header: str) -> bool:
    h = _norm_cell(header)
    return any(m in h for m in REVISION_COLUMN_MARKERS)


def _find_knowledge_columns(header_map: dict, field_cols: dict, anchor_col_set: set) -> list:
    """返回 [(表头, 列号), ...]，兼容自定义后段列名。"""
    cols = []
    seen = set()

    def add(header, col):
        if not col or col in anchor_col_set or col in seen:
            return
        if _is_revision_column(header):
            return
        seen.add(col)
        cols.append((header, col))

    if field_cols.get("content"):
        for h, c in header_map.items():
            if c == field_cols["content"]:
                add(h or "知识描述", c)
                break
        else:
            add("知识描述", field_cols["content"])

    for h, c in header_map.items():
        if not h or c in anchor_col_set:
            continue
        if any(m in h for m in CONTENT_HEADER_MARKERS):
            add(h, c)

    return cols


def _infer_knowledge_columns_by_fill(ws, data_start: int, anchor_col_set: set) -> list:
    """无标准列名时：选数据区填充最多的非锚定、非修订列。"""
    try:
        from step2_preextract import build_header_map
    except ImportError:
        return []

    header_map = build_header_map(ws)
    scores = []
    max_row = ws.max_row or data_start
    for h, c in header_map.items():
        if c in anchor_col_set or _is_revision_column(h):
            continue
        count = 0
        for r in range(data_start, max_row + 1):
            if len(_norm_cell(ws.cell(r, c).value)) >= 4:
                count += 1
        if count > 0:
            scores.append((count, h, c))
    scores.sort(reverse=True)
    return [(h, c) for _, h, c in scores[:8]]


def _extract_anchor_template_sheet(ws) -> list:
    """Step1/2/3/4 场景锚定模板：前四列锚定 + 后段任意知识列（含双行表头）。"""
    try:
        from step1_template import (
            detect_anchor_block_height,
            detect_data_start_row,
            find_anchor_columns,
        )
        from step2_preextract import build_header_map, resolve_field_columns
    except ImportError:
        return []

    header_map = build_header_map(ws)
    field_cols = resolve_field_columns(header_map)
    anchor_cols = find_anchor_columns(ws)
    if "scenario" not in anchor_cols:
        return []

    anchor_col_set = set(anchor_cols.values())
    data_start = detect_data_start_row(ws)
    block_height = max(1, detect_anchor_block_height(ws, data_start, sorted(anchor_col_set)))
    knowledge_cols = _find_knowledge_columns(header_map, field_cols, anchor_col_set)
    if not knowledge_cols:
        knowledge_cols = _infer_knowledge_columns_by_fill(ws, data_start, anchor_col_set)
    if not knowledge_cols:
        return []

    records = []
    current_scenario = ""
    current_sub = ""

    for r in range(data_start, (ws.max_row or data_start) + 1):
        if anchor_cols.get("scenario"):
            v = _norm_cell(ws.cell(r, anchor_cols["scenario"]).value)
            if v:
                current_scenario = v
        if anchor_cols.get("sub_scenario"):
            v = _read_cell(ws, r, anchor_cols["sub_scenario"])
            if not v:
                block_start = data_start + ((r - data_start) // block_height) * block_height
                v = _read_cell(ws, block_start, anchor_cols["sub_scenario"])
            if v:
                current_sub = v

        block_start = data_start + ((r - data_start) // block_height) * block_height
        sub_desc = ""
        if anchor_cols.get("sub_scenario_desc"):
            sub_desc = _read_cell(ws, block_start, anchor_cols["sub_scenario_desc"])

        cells = []
        for h, c in knowledge_cols:
            v = _norm_cell(ws.cell(r, c).value)
            if len(v) >= 2:
                cells.append((h, v))
        if not cells:
            continue

        primary = max(cells, key=lambda x: len(x[1]))
        record = {
            "知识分类": current_scenario or ws.title,
            "知识描述": primary[1],
            "适用条件": current_sub,
            "子场景": current_sub,
            "子场景说明": sub_desc,
            "场景": current_scenario,
            "_source_sheet": ws.title,
        }
        if field_cols.get("category"):
            cat = _norm_cell(ws.cell(r, field_cols["category"]).value)
            if cat:
                record["知识分类"] = cat
        for h, v in cells:
            if h != primary[0]:
                record[h] = v
        records.append(record)

    return records


def _is_knowledge_sheet(ws) -> bool:
    """Check if a worksheet looks like it contains knowledge items."""
    try:
        from step1_template import find_anchor_columns
        from step2_preextract import build_header_map

        if "scenario" in find_anchor_columns(ws):
            return True
        header_map = build_header_map(ws)
        hits = 0
        for h in header_map:
            if not h:
                continue
            if _is_revision_column(h):
                continue
            if h in ("场景", "子场景", "场景说明", "子场景说明"):
                hits += 1
            if any(m in h for m in CONTENT_HEADER_MARKERS):
                hits += 1
        if hits >= 2:
            return True
    except ImportError:
        pass

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
    if not any(headers):
        return False
    header_strs = {str(h).strip() for h in headers if h}
    key_fields = {"知识分类", "分类", "知识描述", "描述", "知识内容", "内容", "知识编号", "编号",
                  "知识要点", "数据规则", "专家经验", "步骤", "知识", "场景", "具体方法"}
    return len(header_strs & key_fields) >= 2


def _extract_rows_from_sheet(ws) -> list:
    """Extract knowledge records from a worksheet.
    
    Handles two row patterns:
    1. Traditional: each row is a complete knowledge item
    2. Scenario-step: rows group by scenario/sub-scenario, each row with a step is a knowledge point
    """
    anchor_records = _extract_anchor_template_sheet(ws)
    if anchor_records:
        return anchor_records

    headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
    # Resolve headers to canonical names
    headers = [_resolve_header(h) for h in headers_raw]

    records = []
    
    # Check if this is a scenario-step format (has 场景/步骤 columns)
    has_scenario = "场景" in headers
    has_step = "步骤" in headers
    
    if has_scenario or has_step:
        # Scenario-step format: track current scenario/sub-scenario context
        current_scenario = ""
        current_sub_scenario = ""
        current_scenario_desc = ""
        current_sub_scenario_desc = ""
        
        # Find column indices
        col_indices = {}
        for idx, h in enumerate(headers):
            if h:
                col_indices[h] = idx
        
        scenario_idx = col_indices.get("场景", -1)
        sub_scenario_idx = col_indices.get("子场景", -1)
        scenario_desc_idx = col_indices.get("场景说明", -1)
        sub_scenario_desc_idx = col_indices.get("子场景说明", -1)
        step_idx = col_indices.get("步骤", -1)
        content_idx = col_indices.get("知识描述", -1)
        if content_idx < 0:
            for idx, h in enumerate(headers):
                if h in ("知识描述", "具体方法", "数据规则"):
                    content_idx = idx
                    break

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            
            # Update context from scenario/sub-scenario columns
            if scenario_idx >= 0 and len(row) > scenario_idx and row[scenario_idx]:
                current_scenario = str(row[scenario_idx]).strip()
            if sub_scenario_idx >= 0 and len(row) > sub_scenario_idx and row[sub_scenario_idx]:
                current_sub_scenario = str(row[sub_scenario_idx]).strip()
            if scenario_desc_idx >= 0 and len(row) > scenario_desc_idx and row[scenario_desc_idx]:
                current_scenario_desc = str(row[scenario_desc_idx]).strip()
            if sub_scenario_desc_idx >= 0 and len(row) > sub_scenario_desc_idx and row[sub_scenario_desc_idx]:
                current_sub_scenario_desc = str(row[sub_scenario_desc_idx]).strip()
            
            content = ""
            if content_idx >= 0 and len(row) > content_idx and row[content_idx]:
                content = str(row[content_idx]).strip()

            step_name = ""
            if step_idx >= 0 and len(row) > step_idx and row[step_idx]:
                step_name = str(row[step_idx]).strip()

            if not content and not step_name:
                continue
            if not step_name:
                step_name = content[:60] if content else "知识条目"

            if content or step_idx >= 0:
                # Get knowledge reference
                knowledge_ref = ""
                knowledge_ref_idx = col_indices.get("知识引用", -1)
                if knowledge_ref_idx >= 0 and len(row) > knowledge_ref_idx and row[knowledge_ref_idx]:
                    knowledge_ref = str(row[knowledge_ref_idx]).strip()
                
                # Get rule reference
                rule_ref = ""
                rule_ref_idx = col_indices.get("规则引用", -1)
                if rule_ref_idx >= 0 and len(row) > rule_ref_idx and row[rule_ref_idx]:
                    rule_ref = str(row[rule_ref_idx]).strip()
                
                # Get professional terms
                prof_terms = ""
                prof_terms_idx = col_indices.get("专业术语", -1)
                if prof_terms_idx >= 0 and len(row) > prof_terms_idx and row[prof_terms_idx]:
                    prof_terms = str(row[prof_terms_idx]).strip()
                
                # Build record
                record = {
                    "知识分类": current_scenario or ws.title,
                    "知识描述": content or step_name,
                    "适用条件": current_sub_scenario or current_scenario_desc,
                    "步骤": step_name,
                    "_source_sheet": ws.title,
                }
                
                if knowledge_ref:
                    record["知识引用"] = knowledge_ref
                if rule_ref:
                    record["判断逻辑"] = rule_ref
                if prof_terms:
                    record["专业术语"] = prof_terms
                
                records.append(record)
    else:
        # Traditional format: each row is a complete record
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            record = {}
            for idx, header in enumerate(headers):
                val = row[idx] if idx < len(row) else ""
                if header:
                    record[header] = val
            # Only include rows that have at least a category or description
            if record.get("知识分类") or record.get("知识描述"):
                records.append(record)
    
    return records


def _read_cell(ws, row: int, col: int) -> str:
    try:
        from step2_preextract import _read_effective_cell_value
        return _norm_cell(_read_effective_cell_value(ws, row, col))
    except ImportError:
        return _norm_cell(ws.cell(row, col).value)


def extract_scenario_anchor_from_sheet(ws) -> dict:
    """从 Step1 锚定四列读取场景名称、场景说明（第一步填写）、子场景列表。"""
    try:
        from step1_template import (
            detect_anchor_block_height,
            detect_data_start_row,
            find_anchor_columns,
        )
    except ImportError:
        return {}

    anchor = find_anchor_columns(ws)
    if "scenario" not in anchor:
        return {}

    data_start = detect_data_start_row(ws)
    anchor_cols = sorted(anchor.values())
    block_height = max(1, detect_anchor_block_height(ws, data_start, anchor_cols))
    max_row = ws.max_row or data_start
    block_count = max(1, (max_row - data_start + 1) // block_height)

    first_block = data_start
    scenario_name = _read_cell(ws, first_block, anchor["scenario"]) if anchor.get("scenario") else ""
    scenario_content = ""
    if anchor.get("scenario_desc"):
        scenario_content = _read_cell(ws, first_block, anchor["scenario_desc"])

    sub_scenarios = []
    seen_sub = set()
    for i in range(block_count):
        block_start = data_start + i * block_height
        sub_name = _read_cell(ws, block_start, anchor["sub_scenario"]) if anchor.get("sub_scenario") else ""
        sub_content = _read_cell(ws, block_start, anchor["sub_scenario_desc"]) if anchor.get("sub_scenario_desc") else ""
        if not sub_name and not sub_content:
            continue
        key = sub_name or sub_content[:40]
        if key in seen_sub:
            continue
        seen_sub.add(key)
        sub_scenarios.append({"name": sub_name, "content": sub_content})

    return {
        "scenario_name": scenario_name,
        "scenario_content": scenario_content,
        "sub_scenarios": sub_scenarios,
        "sheet": ws.title,
    }


def extract_scenario_anchor_from_workbook(wb) -> dict:
    for sn in wb.sheetnames:
        if sn in NON_KNOWLEDGE_SHEETS:
            continue
        ctx = extract_scenario_anchor_from_sheet(wb[sn])
        if ctx.get("scenario_name") or ctx.get("scenario_content") or ctx.get("sub_scenarios"):
            return ctx
    return {}


def _meaningful_sub_scenarios(sub_list) -> list:
    """子场景/子场景说明均为可选，过滤全空条目。"""
    try:
        from step1_template import normalize_sub_scenarios
        return normalize_sub_scenarios(sub_list)
    except ImportError:
        out = []
        for sub in sub_list or []:
            if not isinstance(sub, dict):
                continue
            name = str(sub.get("name") or "").strip()
            content = str(sub.get("content") or "").strip()
            if name or content:
                out.append({"name": name, "content": content})
        return out


def merge_scenario_context(excel_ctx: dict, pipeline_ctx: dict | None) -> dict:
    """Excel 锚定列为主；流水线 step1 表单可补全场景说明（第一步原文）。"""
    out = dict(excel_ctx or {})
    pipe = pipeline_ctx or {}
    if pipe.get("scenario_name") and not out.get("scenario_name"):
        out["scenario_name"] = pipe["scenario_name"]
    if pipe.get("scenario_content"):
        out["scenario_content"] = pipe["scenario_content"]
    excel_subs = _meaningful_sub_scenarios(out.get("sub_scenarios"))
    pipe_subs = _meaningful_sub_scenarios(pipe.get("sub_scenarios"))
    if pipe_subs and not excel_subs:
        out["sub_scenarios"] = pipe_subs
    else:
        out["sub_scenarios"] = excel_subs
    out["场景名称"] = out.get("scenario_name") or pipe.get("scenario_name") or ""
    out["场景说明"] = out.get("scenario_content") or pipe.get("scenario_content") or ""
    return out


def read_excel_knowledge(excel_path: str, pipeline_context: dict | None = None) -> tuple:
    """读取 Excel 知识条目和版本信息，返回 (records, version_info)

    Supports two formats:
    1. Legacy: a single sheet named "知识条目"
    2. Multi-sheet: scenario-specific sheets (e.g. "营销", "修订") with knowledge columns
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    records = []
    version_info = {}
    excel_ctx = extract_scenario_anchor_from_workbook(wb)
    version_info["scenario_anchor"] = merge_scenario_context(excel_ctx, pipeline_context)

    # Try legacy format first
    if "知识条目" in wb.sheetnames:
        ws = wb["知识条目"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
        if any(headers):
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or not row[0]:
                    continue
                record = {}
                for idx, header in enumerate(headers):
                    record[header] = row[idx] if idx < len(row) else ""
                records.append(record)
    else:
        # Multi-sheet / 场景锚定模板：逐 sheet 抽取，锚定模板不再依赖单行表头识别
        for sn in wb.sheetnames:
            if sn in NON_KNOWLEDGE_SHEETS:
                continue
            ws = wb[sn]
            sheet_records = _extract_rows_from_sheet(ws)
            if not sheet_records and _is_knowledge_sheet(ws):
                sheet_records = _extract_anchor_template_sheet(ws)
            for rec in sheet_records:
                rec.setdefault("_source_sheet", sn)
            records.extend(sheet_records)

    # 读取版本追踪
    if "版本追踪" in wb.sheetnames:
        ws4 = wb["版本追踪"]
        versions = []
        for row in ws4.iter_rows(min_row=2, values_only=True):
            if row[0]:
                versions.append({
                    "version": str(row[0]),
                    "date": str(row[1] or ""),
                    "operator": str(row[2] or ""),
                    "note": str(row[3] or ""),
                })
        version_info["versions"] = versions

    # 读取场景配置
    if "场景配置" in wb.sheetnames:
        ws2 = wb["场景配置"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                version_info[str(row[0])] = str(row[1])

    wb.close()
    return records, version_info


def load_scenario_config(config_path: str) -> dict:
    """加载场景 YAML 配置"""
    if not config_path or not os.path.exists(config_path):
        return {}
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def validate_records(records: list) -> list:
    """基础数据校验"""
    errors = []
    # For scenario-step format, check 知识分类 and 知识描述
    # For traditional format, check 知识编号, 知识分类, 知识描述
    has_step_format = any("步骤" in rec for rec in records)
    
    for i, rec in enumerate(records):
        if has_step_format:
            # Scenario-step format: only require 知识分类 and 知识描述
            kn_id = rec.get("步骤", f"行{i+2}")
            if not str(rec.get("知识分类", "")).strip():
                errors.append({"id": kn_id, "field": "知识分类", "error": "empty_required"})
            if not str(rec.get("知识描述", "")).strip():
                errors.append({"id": kn_id, "field": "知识描述", "error": "empty_required"})
        else:
            # Traditional format
            kn_id = rec.get("知识编号", f"行{i+5}")
            for field in ["知识编号", "知识分类", "知识描述"]:
                if not str(rec.get(field, "")).strip():
                    errors.append({"id": kn_id, "field": field, "error": "empty_required"})
        
        conf = str(rec.get("置信度", "")).strip()
        if conf and conf not in ("高", "中", "低"):
            errors.append({"id": rec.get("知识编号", rec.get("步骤", f"行{i+2}")), "field": "置信度", "error": "invalid_confidence", "value": conf})
    return errors


def group_by_category(records: list) -> dict:
    """按知识分类分组"""
    groups = {}
    for rec in records:
        category = rec.get("知识分类") or "未分类"
        if category not in groups:
            groups[category] = []
        groups[category].append(rec)
    return groups


def group_by_sub_scenario(records: list) -> dict:
    """按子场景分组（对齐 Step1 模板结构）；未填写子场景时使用占位键。"""
    groups = {}
    for rec in records:
        sub = (
            str(rec.get("子场景") or rec.get("适用条件") or "").strip()
            or EMPTY_SUB_FALLBACK
        )
        groups.setdefault(sub, []).append(rec)
    return groups


def _records_have_explicit_sub(records: list) -> bool:
    return any(
        str(rec.get("子场景") or rec.get("适用条件") or "").strip()
        for rec in records
    )


def _should_use_sub_structure(anchor: dict, records: list, sub_groups: dict) -> bool:
    """仅在有实际子场景信息时按子场景分节，避免空子场景触发错误版式。"""
    if _meaningful_sub_scenarios(anchor.get("sub_scenarios")):
        return True
    if _records_have_explicit_sub(records):
        return True
    explicit_keys = [k for k in sub_groups if k != EMPTY_SUB_FALLBACK]
    return len(explicit_keys) > 1


def format_knowledge_item(rec: dict) -> str:
    """格式化单条知识条目"""
    parts = []
    desc = rec.get("知识描述", "")
    if desc:
        parts.append(f"- **{desc}**")

    # 场景专属字段（非通用字段，排在前面）
    common_field_names = {
        "知识编号", "知识分类", "知识描述", "适用条件", "判断逻辑",
        "反模式/踩坑提示", "来源文档", "来源位置", "原文摘录",
        "置信度", "贡献专家", "确认专家", "备注",
        "场景", "子场景", "子场景说明", "步骤", "_source_sheet",
        "经验判断", "适用边界", "例外情形", "证据数", "突破数",
    }
    for key, val in rec.items():
        if key not in common_field_names and val and str(val).strip():
            parts.append(f"  - {key}：{val}")

    condition = rec.get("适用条件", "")
    if condition:
        parts.append(f"  - 适用条件：{condition}")

    logic = rec.get("判断逻辑", "")
    if logic:
        parts.append(f"  - 判断逻辑：{logic}")

    anti_pattern = rec.get("反模式/踩坑提示", "")
    if anti_pattern:
        parts.append(f"  - 反模式：{anti_pattern}")

    # L2 隐性上下文层
    exp = rec.get("经验判断", "")
    if exp:
        parts.append(f"  - 经验判断（专家）：{exp}")
    boundary = rec.get("适用边界", "")
    if boundary:
        parts.append(f"  - 适用边界：{boundary}")
    exception = rec.get("例外情形", "")
    if exception:
        parts.append(f"  - 例外情形：{exception}")

    source = rec.get("来源文档", "")
    source_loc = rec.get("来源位置", "")
    if source:
        source_str = source
        if source_loc:
            source_str += f" ({source_loc})"
        parts.append(f"  - 来源：{source_str}")

    contributor = rec.get("贡献专家", "")
    confirmer = rec.get("确认专家", "")
    if contributor or confirmer:
        attribution = []
        if contributor:
            attribution.append(f"贡献：{contributor}")
        if confirmer:
            attribution.append(f"确认：{confirmer}")
        parts.append(f"  - 署名：{' | '.join(attribution)}")

    return "\n".join(parts)


def compute_quality_metrics(records: list, groups: dict) -> dict:
    """计算详细质量指标"""
    total = len(records)
    high_conf = sum(1 for r in records if str(r.get("置信度", "")).strip() == "高")
    mid_conf = sum(1 for r in records if str(r.get("置信度", "")).strip() == "中")
    low_conf = sum(1 for r in records if str(r.get("置信度", "")).strip() == "低")
    has_anti = sum(1 for r in records if str(r.get("反模式/踩坑提示", "")).strip())
    has_logic = sum(1 for r in records if str(r.get("判断逻辑", "")).strip())
    has_source = sum(1 for r in records if str(r.get("来源文档", "")).strip())
    has_contributor = sum(1 for r in records if str(r.get("贡献专家", "")).strip())
    has_confirmer = sum(1 for r in records if str(r.get("确认专家", "")).strip())

    return {
        "total_knowledge_items": total,
        "confidence_distribution": {"high": high_conf, "mid": mid_conf, "low": low_conf},
        "high_confidence_ratio": round(high_conf / total, 2) if total > 0 else 0,
        "anti_pattern_count": has_anti,
        "anti_pattern_ratio": round(has_anti / total, 2) if total > 0 else 0,
        "logic_coverage": round(has_logic / total, 2) if total > 0 else 0,
        "source_coverage": round(has_source / total, 2) if total > 0 else 0,
        "contributor_coverage": round(has_contributor / total, 2) if total > 0 else 0,
        "confirmer_coverage": round(has_confirmer / total, 2) if total > 0 else 0,
        "category_count": len(groups),
        "category_distribution": {k: len(v) for k, v in groups.items()},
    }


def _format_overview_scenario(version_info: dict, records: list) -> list:
    """概述：写入 Step1 场景名称与场景说明（第一步填写内容）。"""
    lines = []
    anchor = version_info.get("scenario_anchor") or {}
    scenario_name = anchor.get("场景名称") or anchor.get("scenario_name") or version_info.get("场景名称", "")
    scenario_content = anchor.get("场景说明") or anchor.get("scenario_content") or ""

    lines.append("## 概述")
    lines.append("")
    lines.append("本 Skill 由隐性知识显性化四步法流水线生成，将确认版 Excel 中的场景锚定与知识列转化为结构化 Markdown。")
    lines.append("")

    if scenario_name:
        lines.append("### 场景名称")
        lines.append("")
        lines.append(scenario_name)
        lines.append("")

    if scenario_content:
        lines.append("### 场景说明")
        lines.append("")
        for para in str(scenario_content).split("\n"):
            p = para.strip()
            if p:
                lines.append(p)
            else:
                lines.append("")
        lines.append("")

    sub_list = _meaningful_sub_scenarios(anchor.get("sub_scenarios"))
    if sub_list:
        lines.append("### 子场景一览")
        lines.append("")
        for sub in sub_list:
            name = (sub.get("name") or "").strip()
            content = (sub.get("content") or "").strip()
            if name and content:
                lines.append(f"- **{name}**：{content}")
            elif name:
                lines.append(f"- **{name}**")
            elif content:
                lines.append(f"- {content}")
        lines.append("")

    lines.append("### 知识库统计")
    lines.append("")
    lines.append(f"- 知识条目数：**{len(records)}**")
    ver = version_info.get("模板版本", "v1.0")
    lines.append(f"- 模板版本：{ver}")
    lines.append(f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    return lines


def generate_skill_md(records: list, groups: dict, config: dict, scenario_name: str, version_info: dict) -> str:
    """生成 SKILL.md：概述含 Step1 场景内容，正文按子场景组织知识。"""
    compilation = config.get("compilation", {})
    anchor = version_info.get("scenario_anchor") or {}
    display_name = (
        config.get("display_name")
        or anchor.get("场景名称")
        or version_info.get("场景名称", scenario_name)
    )
    domain = config.get("domain", version_info.get("业务领域", "通用"))
    sub_groups = group_by_sub_scenario(records)
    use_sub_structure = _should_use_sub_structure(anchor, records, sub_groups)

    lines = []

    # Front matter
    lines.append("---")
    slug = (scenario_name or display_name or "skill").replace(" ", "-")
    lines.append(f"name: {slug}")
    desc = f"{display_name}专家经验 Skill：含场景说明（Step1）与 {len(records)} 条确认知识"
    lines.append(f"description: {desc}")
    lines.append("---")
    lines.append("")

    lines.append(f"# {display_name} · 专家经验知识库")
    lines.append("")

    lines.extend(_format_overview_scenario(version_info, records))

    lines.append("## 知识结构（由 Excel 转化）")
    lines.append("")
    if use_sub_structure:
        lines.append("以下按 **子场景 → 知识要点** 组织，字段含义与萃取模板后段列一致。")
    else:
        lines.append("以下按 **知识分类 → 知识要点** 组织，字段含义与萃取模板后段列一致。")
    lines.append("")

    if use_sub_structure:
        sub_meta = {
            (s.get("name") or "").strip(): (s.get("content") or "").strip()
            for s in _meaningful_sub_scenarios(anchor.get("sub_scenarios"))
            if (s.get("name") or "").strip()
        }
        for sub_name, items in sub_groups.items():
            if sub_name == EMPTY_SUB_FALLBACK:
                continue
            lines.append(f"### 子场景：{sub_name}")
            lines.append("")
            if sub_meta.get(sub_name):
                lines.append(f"> {sub_meta[sub_name]}")
                lines.append("")
            for item in items:
                lines.append(format_knowledge_item(item))
                lines.append("")
        fallback_items = sub_groups.get(EMPTY_SUB_FALLBACK)
        if fallback_items:
            lines.append("### 其他知识")
            lines.append("")
            for item in fallback_items:
                lines.append(format_knowledge_item(item))
                lines.append("")
    else:
        for category, items in groups.items():
            lines.append(f"### {category}")
            lines.append("")
            for item in items:
                lines.append(format_knowledge_item(item))
                lines.append("")

    # Anti-patterns summary
    if compilation.get("include_anti_patterns_section", True):
        anti_items = [r for r in records if str(r.get("反模式/踩坑提示", "")).strip()]
        if anti_items:
            lines.append("## 反模式与踩坑总结")
            lines.append("")
            for item in anti_items:
                desc = item.get("知识描述", "未命名")
                anti = item.get("反模式/踩坑提示", "")
                lines.append(f"- **{desc}**：{anti}")
            lines.append("")

    # Usage examples
    if compilation.get("include_usage_examples", True):
        lines.append("## 使用示例")
        lines.append("")
        lines.append("- 示例1：")
        lines.append(f"  - 场景：用户在{domain}决策中遇到不确定情况")
        lines.append(f"  - 用法：加载本 Skill，在知识条目中搜索相关条件，参考判断逻辑和反模式")
        lines.append("")
        lines.append("- 示例2：")
        lines.append(f"  - 场景：新人学习{domain}业务规范")
        lines.append(f"  - 用法：通读本 Skill 的知识条目与反模式总结，快速建立判断框架")
        lines.append("")

    return "\n".join(lines)


def excel_to_skill(excel_path: str, config_path: str, output_dir: str, pipeline_context: dict | None = None) -> dict:
    """主转换流程"""
    # 1. Read Excel
    records, version_info = read_excel_knowledge(excel_path, pipeline_context)
    if not records:
        return {
            "status": "error",
            "message": (
                f"No knowledge items found in {excel_path}。"
                "请确认最终稿/修订稿中后段知识列（具体方法、知识描述等）已有内容，"
                "且非仅含场景四列或修订元数据列。"
            ),
        }

    # 2. Validate (non-fatal: collect warnings but continue)
    errors = validate_records(records)
    warnings = errors if errors else []

    # 3. Load config
    config = load_scenario_config(config_path)

    # 4. Determine scenario name
    scenario_name = config.get("scenario_name", version_info.get("场景名称", os.path.splitext(os.path.basename(excel_path))[0]))

    # 5. Group by category
    groups = group_by_category(records)

    # 6. Generate SKILL.md
    skill_content = generate_skill_md(records, groups, config, scenario_name, version_info)

    # 7. Write output
    os.makedirs(output_dir, exist_ok=True)
    skill_path = os.path.join(output_dir, "SKILL.md")
    with open(skill_path, "w", encoding="utf-8") as f:
        f.write(skill_content)

    # 8. Compute quality metrics
    metrics = compute_quality_metrics(records, groups)

    return {
        "status": "ok",
        "skill_path": skill_path,
        "knowledge_count": len(records),
        "category_count": len(groups),
        "quality_metrics": metrics,
        "version_info": version_info,
        "warnings": warnings,
    }


def main():
    parser = argparse.ArgumentParser(description="将确认版 Excel 知识库转换为 SKILL.md")
    parser.add_argument("--input", required=True, help="输入 Excel v1.0 文件路径")
    parser.add_argument("--config", default="", help="场景 YAML 配置文件路径（可选）")
    parser.add_argument("--output", required=True, help="输出目录路径")
    parser.add_argument("--context-json", default="", help="流水线 Step1 表单 JSON（场景名称/说明/子场景）")
    args = parser.parse_args()

    pipeline_ctx = {}
    if args.context_json:
        try:
            pipeline_ctx = json.loads(args.context_json)
        except json.JSONDecodeError:
            pipeline_ctx = {}

    try:
        result = excel_to_skill(args.input, args.config, args.output, pipeline_ctx)
        print(json.dumps(result, ensure_ascii=False))
    except FileNotFoundError as e:
        print(json.dumps({"status": "error", "message": f"文件不存在: {str(e)}"}, ensure_ascii=False))
        sys.exit(1)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(json.dumps({"status": "error", "message": str(e), "detail": error_detail}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
