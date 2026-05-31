#!/usr/bin/env python3
"""校验 Step2 萃取仅产出 .xlsx（不依赖 LLM / 网络）。"""

import sys
import tempfile
from pathlib import Path

import openpyxl

BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND))

from step2_preextract import write_preextract_excel  # noqa: E402


def _make_step1_template(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "萃取模板"
    headers = ["场景", "场景说明", "子场景", "子场景说明", "环节", "具体方法", "知识类型"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, value=h)
    ws.cell(3, 1, "演示场景")
    ws.cell(3, 2, "场景说明文本")
    ws.cell(3, 3, "子场景A")
    ws.cell(3, 4, "子场景说明")
    wb.save(path)
    wb.close()


def main():
    errors = []
    tmp = Path(tempfile.mkdtemp(prefix="step2_validate_"))
    try:
        step1 = tmp / "step1_skeleton.xlsx"
        out = tmp / "preextract_test.xlsx"
        _make_step1_template(step1)

        items = [
            {
                "category": "判断规则",
                "content": "校验条目：授信额度需结合客户评级",
                "trigger_condition": "新客户准入",
                "judgment_logic": "评级A且收入稳定则可通过",
                "anti_pattern": "仅看抵押物不看现金流",
                "confidence": "高",
            },
            {
                "知识分类": "操作流程",
                "知识描述": "第二条中文键名条目",
                "适用条件": "续贷",
                "置信度": "中",
            },
        ]

        meta = write_preextract_excel(step1_path=step1, output_path=out, items=items)

        if not out.exists():
            errors.append("输出文件未生成")
        elif out.suffix.lower() != ".xlsx":
            errors.append(f"扩展名错误: {out.suffix}")

        if meta.get("filled_rows", 0) < 1:
            errors.append(f"未写入数据行 filled_rows={meta.get('filled_rows')}")

        wb = openpyxl.load_workbook(out, data_only=True)
        try:
            ws = wb.active
            v1 = ws.cell(3, 6).value
            v2 = ws.cell(4, 6).value if ws.max_row >= 4 else None
        finally:
            wb.close()

        if not v1 or "校验条目" not in str(v1):
            errors.append(f"第3行具体方法未填入: {v1!r}")
        if not v2 or "第二条" not in str(v2):
            errors.append(f"第4行具体方法未填入: {v2!r}")

        # 无 Step1 时应仍能生成标准表
        out2 = tmp / "preextract_fallback.xlsx"
        write_preextract_excel(step1_path=None, output_path=out2, items=items[:1])
        if not out2.exists():
            errors.append("无模板回退文件未生成")
    finally:
        import shutil
        try:
            shutil.rmtree(tmp, ignore_errors=True)
        except Exception:
            pass

    if errors:
        print("FAIL")
        for e in errors:
            print(" -", e)
        sys.exit(1)

    print("OK: Step2 萃取 Excel 校验通过（模板回填 + 回退表）")
    print(f"  filled_rows={meta['filled_rows']} used_step1_template={meta['used_step1_template']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
