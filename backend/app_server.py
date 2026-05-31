#!/usr/bin/env python3
"""Flask backend server for the Tacit Knowledge Extraction web application."""

import argparse
import datetime
import json
import logging
import os
import re
import subprocess
import sys
import tempfile
import uuid
import shutil
import threading
from contextlib import contextmanager
from pathlib import Path

import yaml
from flask import Flask, request, jsonify, send_file, send_from_directory, Response, stream_with_context
from llm_client import (
    API_TYPE_CCB,
    API_TYPE_OPENAI,
    LlmApiError,
    call_llm,
    call_llm_with_retry,
    extract_assistant_content,
    iter_llm_stream,
    normalize_llm_url,
)
import openpyxl

from pipeline_artifacts import (
    basename_only,
    downstream_output_keys,
    keys_to_clear_from_step,
    is_download_allowed,
    is_step1_filename,
    is_step2_preextract_filename,
    is_step3_revision_filename,
    is_step4_final_filename,
    resolve_cache_file_path,
    resolve_client_excel_path,
    safe_workspace_path,
    validate_step_data_patch,
    resolve_knowledge_workbook_path,
    PROTECTED_WORKSPACE_FILES,
)
from release_info import STEP2_EXCEL_BUILD, get_release_info

# 旧部署曾误用私有函数名，保留别名避免 NameError
_is_step2_preextract_filename = is_step2_preextract_filename

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB upload limit

# Add scripts directory to Python path for skill imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Resolve paths relative to this script
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
SAMPLES_DIR = PROJECT_DIR / "data" / "samples"
CONFIG_DIR = PROJECT_DIR / "config"
FRONTEND_DIR = PROJECT_DIR / "frontend"
SCHEMA_PATH = CONFIG_DIR / "scenario-schema.yaml"
LLM_CONFIG_PATH = CONFIG_DIR / "llm-config.yaml"
LLM_CONFIG_LOCAL_PATH = CONFIG_DIR / "llm-config.local.yaml"

# Optional API token auth (set APP_AUTH_TOKEN in production)
APP_AUTH_TOKEN = os.environ.get("APP_AUTH_TOKEN", "").strip()

# Workspace: where uploaded/generated files are stored (Docker: WORKSPACE_DIR=/app/workspace)
WORKSPACE = Path(os.environ.get("WORKSPACE_DIR", Path(tempfile.gettempdir()) / "tacit_knowledge_app"))
WORKSPACE.mkdir(parents=True, exist_ok=True)

# Custom models persistence
CUSTOM_MODELS_PATH = WORKSPACE / "custom_models.json"
PRESET_OVERRIDES_PATH = WORKSPACE / "preset_overrides.json"

# Pipelines persistence
PIPELINES_PATH = WORKSPACE / "pipelines.json"

# Thread-safe model state
_models_lock = threading.Lock()
_pipelines_lock = threading.Lock()

AUTH_EXEMPT_PATHS = frozenset({"/api/health", "/api/version", "/api/auth/config"})


def _mask_api_key(key: str) -> str:
    key = key or ""
    if len(key) > 8:
        return key[:4] + "****" + key[-4:]
    return "****" if key else ""


def _sanitize_model_for_client(model_cfg: dict) -> dict:
    d = {k: v for k, v in model_cfg.items() if k not in ("api_key", "fst_attr_rmrk")}
    d["api_key_masked"] = _mask_api_key(model_cfg.get("api_key", ""))
    d["has_api_key"] = bool((model_cfg.get("api_key") or "").strip())
    return d


@app.before_request
def _require_api_auth():
    if not APP_AUTH_TOKEN:
        return None
    if not request.path.startswith("/api/"):
        return None
    if request.path in AUTH_EXEMPT_PATHS:
        return None
    auth = (request.headers.get("Authorization") or "").strip()
    if auth == f"Bearer {APP_AUTH_TOKEN}":
        return None
    return jsonify({"status": "error", "error": "未授权访问"}), 401


EXTRACT_STYLE_RULES = {
    "标准萃取": {
        "temperature": 0.25,
        "max_tokens": 4096,
        "min_items": 8,
        "max_items": 22,
        "prompt_hint": "平衡覆盖核心规则、流程与经验，优先可执行条目。",
    },
    "深度萃取": {
        "temperature": 0.35,
        "max_tokens": 4096,
        "min_items": 12,
        "max_items": 40,
        "prompt_hint": "优先完整覆盖，尽量补全触发条件、判断逻辑、反模式与证据字段。",
    },
    "精简萃取": {
        "temperature": 0.15,
        "max_tokens": 2048,
        "min_items": 5,
        "max_items": 10,
        "prompt_hint": "只保留高价值高置信条目，减少冗余与重复。",
    },
}

REVISION_STYLE_RULES = {
    "标准修订": {
        "temperature": 0.25,
        "max_actions": 40,
        "allowed_actions": {"modify", "supplement", "add", "delete"},
        "prompt_hint": "平衡修订：采纳明确建议，同时保持原有合理内容。",
    },
    "严格修订": {
        "temperature": 0.1,
        "max_actions": 20,
        "allowed_actions": {"modify", "supplement"},
        "prompt_hint": "保守修订：仅处理证据充分、定位明确的修改/补充，禁止新增和删除。",
    },
    "宽松修订": {
        "temperature": 0.35,
        "max_actions": 80,
        "allowed_actions": {"modify", "supplement", "add", "delete"},
        "prompt_hint": "积极修订：尽可能采纳专家建议，允许新增与删除。",
    },
}


# Configure structured logging
_logger = logging.getLogger("tacit_knowledge")
_logger.setLevel(logging.INFO)
if not _logger.handlers:
    _ch = logging.StreamHandler()
    _ch.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    ))
    _logger.addHandler(_ch)


@contextmanager
def _safe_workbook(path, read_only=True, data_only=True):
    """Context manager for openpyxl workbook — ensures proper close on exception."""
    wb = None
    try:
        wb = openpyxl.load_workbook(str(path), read_only=read_only, data_only=data_only)
        yield wb
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


# ─── Logging stubs (restored after _strip_debug_logs.py ran) ────────

def _debug_log(level, source, message, data=None):
    """Structured logging — writes to python logging module."""
    extra = json.dumps(data or {}, ensure_ascii=False, default=str)
    _logger.info("[%s] %s | %s", level, source, f"{message} {extra}")


def _agent_debug_log(run_id, level, source, message, data=None):
    """Agent debug logging."""
    extra = json.dumps(data or {}, ensure_ascii=False, default=str)
    _logger.info("[agent:%s][%s] %s | %s", run_id, level, source, f"{message} {extra}")


def _decode_text_by_filename(filename: str, raw: bytes) -> str:
    """Decode bytes by file extension, supporting TXT/MD/DOCX/PDF."""
    filename = (filename or "").lower()

    if filename.endswith(".pdf"):
        try:
            from PyPDF2 import PdfReader
            import io
            reader = PdfReader(io.BytesIO(raw))
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n".join(pages)
        except Exception:
            return raw.decode("utf-8", errors="replace")

    if filename.endswith(".docx"):
        try:
            from docx import Document
            import io
            doc = Document(io.BytesIO(raw))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)
        except Exception:
            return raw.decode("utf-8", errors="replace")

    for enc in ("utf-8", "gbk", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def extract_text_from_file(file_obj) -> str:
    """Read text from uploaded file, supporting TXT, MD, DOCX, PDF."""
    filename = (file_obj.filename or "").lower()
    raw = file_obj.read()
    return _decode_text_by_filename(filename, raw)


def extract_text_from_path(file_path: str) -> str:
    """Read text from a cached workspace file path."""
    p = Path(file_path)
    raw = p.read_bytes()
    return _decode_text_by_filename(p.name, raw)


def _pipeline_prefers_markdown(pipeline_id: str) -> bool:
    """Whether current pipeline requested markdown-first artifacts."""
    if not pipeline_id:
        return False
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p.get("id") != pipeline_id:
                continue
            sd = p.get("step_data", {}) or {}
            if str(sd.get("step1_output_format", "")).strip().lower() == "markdown":
                return True
            form = sd.get("step1_form_data", {}) or {}
            return str(form.get("output_format", "")).strip().lower() == "markdown"
    return False


def _excel_to_markdown_file(excel_path: str | Path, md_path: str | Path, *, title: str = "") -> None:
    """Render an Excel workbook to markdown tables for user preview/download."""
    with _safe_workbook(str(excel_path)) as wb:
        lines: list[str] = []
        if title:
            lines.append(f"# {title}")
            lines.append("")
        for ws in wb.worksheets:
            lines.append(f"## 工作表：{ws.title}")
            lines.append("")
            rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
            if not rows:
                lines.append("（空表）")
                lines.append("")
                continue
            max_cols = max((len(r or ()) for r in rows), default=0)
            if max_cols <= 0:
                lines.append("（无可用列）")
                lines.append("")
                continue
            header_row = rows[0] or ()
            headers = []
            for i in range(max_cols):
                val = header_row[i] if i < len(header_row) else ""
                headers.append(str(val or "").replace("|", "\\|"))
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * max_cols) + " |")
            for row in rows[1:]:
                vals = []
                row = row or ()
                for i in range(max_cols):
                    val = row[i] if i < len(row) else ""
                    s = str(val or "").replace("\n", " ").replace("|", "\\|")
                    vals.append(s)
                lines.append("| " + " | ".join(vals) + " |")
            lines.append("")
    Path(md_path).write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def _maybe_generate_markdown_artifact(pipeline_id: str, excel_name: str, *, md_prefix: str, title: str) -> tuple[str, str]:
    """Generate markdown artifact from excel when pipeline prefers markdown."""
    if not _pipeline_prefers_markdown(pipeline_id):
        return "", ""
    excel_path = safe_workspace_path(WORKSPACE, excel_name, must_exist=True)
    if not excel_path:
        return "", ""
    stem_id = uuid.uuid4().hex[:8]
    md_name = f"{md_prefix}_{stem_id}.md"
    md_path = WORKSPACE / md_name
    try:
        _excel_to_markdown_file(excel_path, md_path, title=title)
        return md_name, f"/downloads/{md_name}"
    except Exception:
        return "", ""


# ─── LLM Config Helpers ──────────────────────────────────────────


def _extract_balanced_json_slice(text: str, open_ch: str, close_ch: str) -> str:
    """Extract first balanced [...] or {...} slice, respecting quoted strings."""
    start = text.find(open_ch)
    if start < 0:
        return ""
    depth = 0
    in_str = False
    esc = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
            continue
        if ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return ""


def _repair_json_text(text: str) -> str:
    """Best-effort fixes for common LLM JSON mistakes."""
    if not text:
        return text
    t = text.strip()
    t = t.replace("\u201c", '"').replace("\u201d", '"').replace("\u2018", "'").replace("\u2019", "'")
    # Remove trailing commas before } or ]
    t = re.sub(r",\s*([}\]])", r"\1", t)
    return t


def _extract_json_from_text(text: str) -> str:
    """Extract JSON string from LLM response that may contain markdown code blocks."""
    text = (text or "").strip()
    if not text:
        return ""
    # Try markdown code blocks first
    for marker in ["```json", "```"]:
        if marker in text:
            parts = text.split(marker, 1)
            if len(parts) > 1:
                inner = parts[1].split("```", 1)[0].strip()
                if inner:
                    text = inner
                    break
    text = text.strip()
    if text.startswith("["):
        slice_text = _extract_balanced_json_slice(text, "[", "]")
        return slice_text or text
    if text.startswith("{"):
        slice_text = _extract_balanced_json_slice(text, "{", "}")
        return slice_text or text
    # Preamble before JSON array/object (common with custom templates)
    for open_ch, close_ch in (("[", "]"), ("{", "}")):
        slice_text = _extract_balanced_json_slice(text, open_ch, close_ch)
        if slice_text:
            return slice_text
    return text


def _items_from_parsed_root(parsed) -> list | None:
    """Normalize parsed JSON root to a list of item dicts."""
    if isinstance(parsed, list):
        return parsed
    if not isinstance(parsed, dict):
        return None
    for key in (
        "items", "data", "results", "result", "entries", "knowledge",
        "knowledge_items", "extracted", "extracted_items", "知识", "知识条目", "条目",
    ):
        val = parsed.get(key)
        if isinstance(val, list):
            return val
    return None


def _extract_top_level_json_objects(text: str) -> list[str]:
    """Extract complete top-level JSON objects from possibly truncated text."""
    objects = []
    depth = 0
    in_str = False
    esc = False
    start = -1
    for i, ch in enumerate(text):
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue

        if ch == '"':
            in_str = True
            continue
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
            continue
        if ch == "}":
            if depth > 0:
                depth -= 1
                if depth == 0 and start >= 0:
                    objects.append(text[start:i + 1])
                    start = -1
    return objects


def _parse_extracted_items(raw_text: str) -> tuple[list, str]:
    """Parse LLM extraction output with robust fallbacks."""
    cleaned = _extract_json_from_text(raw_text or "")
    if not cleaned.strip():
        return [], "empty"

    for candidate in (cleaned, _repair_json_text(cleaned)):
        if not candidate.strip():
            continue
        try:
            parsed = json.loads(candidate)
            items = _items_from_parsed_root(parsed)
            if items is not None:
                return items, "json_list" if isinstance(parsed, list) else "json_items"
        except Exception:
            pass

    # JSONL: one object per line
    line_items = []
    for line in cleaned.splitlines():
        line = line.strip().rstrip(",")
        if not line.startswith("{"):
            continue
        try:
            obj = json.loads(_repair_json_text(line))
            if isinstance(obj, dict):
                line_items.append(obj)
        except Exception:
            continue
    if line_items:
        return line_items, "jsonl"

    # Fallback: recover complete objects from truncated/dirty output.
    recovered = []
    for obj_text in _extract_top_level_json_objects(_repair_json_text(cleaned)):
        try:
            obj = json.loads(obj_text)
            if isinstance(obj, dict):
                recovered.append(obj)
        except Exception:
            continue
    if recovered:
        return recovered, "object_recovery"

    # Last resort: scan full raw text for embedded objects
    for obj_text in _extract_top_level_json_objects(raw_text or ""):
        try:
            obj = json.loads(_repair_json_text(obj_text))
            if isinstance(obj, dict):
                recovered.append(obj)
        except Exception:
            continue
    if recovered:
        return recovered, "object_recovery"

    return [], "parse_failed"


def _align_item_keys_to_template(item: dict, target_columns: list[str]) -> dict:
    """Map LLM keys (e.g. 具体方法) to template composite keys (环节-具体方法)."""
    if not isinstance(item, dict) or not target_columns:
        return item
    out = dict(item)
    col_norm = {c: c.replace(" ", "") for c in target_columns}
    for col in target_columns:
        if col in out and str(out.get(col) or "").strip():
            continue
        parts = [p.strip() for p in col.replace("：", ":").split("-") if p.strip()]
        suffix = parts[-1] if parts else col
        for k, v in item.items():
            if v is None or isinstance(v, (dict, list)):
                continue
            kn = str(k).replace(" ", "")
            if kn == col.replace(" ", "") or kn == suffix.replace(" ", "") or suffix in str(k):
                out[col] = v
                break
    return out


def _normalize_extracted_items(items: list, target_columns: list[str]) -> list:
    normalized = []
    for raw in items or []:
        if isinstance(raw, dict):
            normalized.append(_align_item_keys_to_template(raw, target_columns))
        elif isinstance(raw, (str, int, float, bool)):
            normalized.append({"content": str(raw)})
    return normalized


def load_preset_overrides():
    """Load runtime overrides for preset models (by name)."""
    if not PRESET_OVERRIDES_PATH.exists():
        return {}
    try:
        with open(str(PRESET_OVERRIDES_PATH), "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, IOError):
        return {}


def save_preset_overrides(overrides):
    """Persist preset model overrides."""
    with open(str(PRESET_OVERRIDES_PATH), "w", encoding="utf-8") as f:
        json.dump(overrides, f, ensure_ascii=False, indent=2)


def load_base_presets():
    """Load preset definitions from YAML + optional local secrets file."""
    if not LLM_CONFIG_PATH.exists():
        return []
    with open(str(LLM_CONFIG_PATH), "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    presets = [dict(m) for m in cfg.get("presets", [])]
    if LLM_CONFIG_LOCAL_PATH.exists():
        with open(str(LLM_CONFIG_LOCAL_PATH), "r", encoding="utf-8") as f:
            local_cfg = yaml.safe_load(f) or {}
        local_by_name = local_cfg.get("presets") or {}
        if isinstance(local_by_name, dict):
            for entry in presets:
                name = entry.get("name")
                if name and name in local_by_name:
                    override = local_by_name[name]
                    if isinstance(override, dict):
                        entry.update(override)
    return presets


def load_llm_config():
    """Load presets from YAML + overrides + custom models from JSON."""
    models = []
    overrides = load_preset_overrides()
    for m in load_base_presets():
        entry = dict(m)
        if entry.get("name") in overrides:
            entry.update(overrides[entry["name"]])
        if entry.get("url"):
            api_type = (entry.get("api_type") or API_TYPE_OPENAI).strip().lower()
            entry["url"] = normalize_llm_url(entry["url"], api_type)
        entry["is_preset"] = True
        models.append(entry)
    custom = []
    if CUSTOM_MODELS_PATH.exists():
        try:
            with open(str(CUSTOM_MODELS_PATH), "r", encoding="utf-8") as f:
                custom = json.load(f)
        except (json.JSONDecodeError, IOError):
            custom = []
    for m in custom:
        entry = dict(m)
        if entry.get("url"):
            api_type = (entry.get("api_type") or API_TYPE_OPENAI).strip().lower()
            entry["url"] = normalize_llm_url(entry["url"], api_type)
        entry["is_preset"] = False
        models.append(entry)
    return models


def save_custom_models(custom_list):
    """Persist custom models list to JSON."""
    # Strip runtime keys before saving
    clean = []
    for m in custom_list:
        d = {k: v for k, v in m.items() if k != "is_preset"}
        clean.append(d)
    with open(str(CUSTOM_MODELS_PATH), "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)


def get_model_by_name(name):
    """Find a model config by name."""
    for m in load_llm_config():
        if m["name"] == name:
            return m
    return None


# ─── Static File Serving ──────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(str(FRONTEND_DIR), "index.html")

@app.route("/css/<path:filename>")
def css(filename):
    return send_from_directory(str(FRONTEND_DIR / "css"), filename)

@app.route("/js/<path:filename>")
def js(filename):
    return send_from_directory(str(FRONTEND_DIR / "js"), filename)

@app.route("/vendor/<path:filename>")
def vendor_static(filename):
    """Luckysheet / jQuery 等离线静态资源（内网部署）"""
    return send_from_directory(str(FRONTEND_DIR / "vendor"), filename)

@app.route("/downloads/<path:filename>")
def downloads(filename):
    base = basename_only(filename)
    if not is_download_allowed(base):
        return jsonify({"status": "error", "error": "不允许下载该文件"}), 403
    path = safe_workspace_path(WORKSPACE, base, must_exist=True)
    if not path:
        return jsonify({"status": "error", "error": "文件不存在"}), 404
    return send_from_directory(str(WORKSPACE), base, as_attachment=True)


@app.route("/api/files/read", methods=["GET"])
def api_file_read():
    """Read a file from workspace and return its content as text."""
    file_name = request.args.get("file_name", "")
    if not file_name:
        return jsonify({"status": "error", "error": "缺少 file_name 参数"})
    # Security: prevent path traversal
    file_name = basename_only(file_name)
    if file_name in PROTECTED_WORKSPACE_FILES or not is_download_allowed(file_name):
        return jsonify({"status": "error", "error": "不允许读取该文件"})
    file_path = safe_workspace_path(WORKSPACE, file_name, must_exist=True)
    if not file_path:
        return jsonify({"status": "error", "error": f"文件不存在: {file_name}"})
    try:
        with open(str(file_path), "r", encoding="utf-8") as f:
            content = f.read()
        return jsonify({"status": "ok", "content": content, "file_name": file_name})
    except Exception as e:
        return jsonify({"status": "error", "error": f"读取失败: {str(e)}"})


@app.route("/api/files/save", methods=["POST"])
def api_file_save():
    """Save text content to a file in workspace."""
    data = request.get_json(force=True)
    file_name = (data.get("file_name") or "").strip()
    content = data.get("content", "")
    if not file_name:
        return jsonify({"status": "error", "error": "缺少 file_name 参数"})
    # Security: prevent path traversal
    file_name = basename_only(file_name)
    if file_name in PROTECTED_WORKSPACE_FILES or not is_download_allowed(file_name):
        return jsonify({"status": "error", "error": "不允许写入该文件"})
    file_path = safe_workspace_path(WORKSPACE, file_name, must_exist=False)
    if not file_path:
        return jsonify({"status": "error", "error": "非法文件路径"})
    try:
        with open(str(file_path), "w", encoding="utf-8") as f:
            f.write(content)
        return jsonify({"status": "ok", "file_name": file_name, "size": len(content)})
    except Exception as e:
        return jsonify({"status": "error", "error": f"保存失败: {str(e)}"})


@app.route("/api/files/cache_upload", methods=["POST"])
def api_file_cache_upload():
    """Cache uploaded source file in workspace, return cached filename."""
    file_obj = request.files.get("file")
    pipeline_id = request.form.get("pipeline_id", "").strip()
    step = request.form.get("step", "").strip()
    if not file_obj or not file_obj.filename:
        return jsonify({"status": "error", "error": "缺少上传文件"})

    try:
        saved_path = save_upload(file_obj, prefix=f"cache_s{step or 'x'}")
        base = os.path.basename(saved_path)
        resp = {"status": "ok", "file_name": base}

        if pipeline_id and step in {"2", "3", "4"}:
            with _pipelines_lock:
                pipelines = load_pipelines()
                for p in pipelines:
                    if p["id"] == pipeline_id:
                        sd = p.setdefault("step_data", {})
                        sd[f"step{step}_cached_file"] = base
                        sd[f"step{step}_cached_name"] = file_obj.filename
                        save_pipelines(pipelines)
                        break
        return jsonify(resp)
    except Exception as e:
        return jsonify({"status": "error", "error": f"缓存上传失败: {str(e)}"})


# ─── Health Check ─────────────────────────────────────────────────

@app.route("/api/frontend/vendor-check")
def api_frontend_vendor_check():
    """检查 Excel 在线编辑所需静态资源是否存在（内网部署自检）"""
    checks = {
        "vendor_route": True,
        "files": {},
    }
    paths = {
        "plugin_js": FRONTEND_DIR / "vendor" / "luckysheet" / "plugins" / "js" / "plugin.js",
        "luckysheet_umd": FRONTEND_DIR / "vendor" / "luckysheet" / "luckysheet.umd.js",
    }
    all_ok = True
    for key, p in paths.items():
        exists = p.is_file()
        checks["files"][key] = {"path": str(p), "exists": exists}
        if not exists:
            all_ok = False
    checks["ok"] = all_ok
    checks["hint"] = (
        "就绪"
        if all_ok
        else "缺少 frontend/vendor，请执行: node scripts/copy-frontend-vendor.js"
    )
    return jsonify(checks)



@app.route("/api/auth/config", methods=["GET"])
def api_auth_config():
    return jsonify({"status": "ok", "auth_required": bool(APP_AUTH_TOKEN)})


@app.route("/api/build_info", methods=["GET"])
def api_build_info():
    info = get_release_info()
    return jsonify({
        "status": "ok",
        "step2_excel": True,
        **info,
    })


@app.route("/api/version", methods=["GET"])
def api_version():
    """标准版本查询（内网部署识别 / 升级比对）。"""
    return jsonify({"status": "ok", **get_release_info()})


@app.route("/api/health")
def health():
    return jsonify({"status": "ok", **get_release_info()})


# ─── Helper: run script and capture JSON output ──────────────────

def run_script(script_name, args_list):
    """Run a script in scripts/ directory and return parsed JSON from stdout."""
    script_path = SCRIPT_DIR / script_name
    cmd = [sys.executable, str(script_path)] + args_list
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=120,
            cwd=str(PROJECT_DIR)
        )
        # 优先尝试从 stdout 解析 JSON（即使 returncode != 0）
        stdout = result.stdout.strip()
        if stdout:
            lines = [l for l in stdout.split("\n") if l.strip()]
            for line in reversed(lines):
                try:
                    parsed = json.loads(line)
                    if isinstance(parsed, dict):
                        return parsed
                except json.JSONDecodeError:
                    continue
        # 如果 stdout 无可解析 JSON，返回 stderr 或通用错误
        stderr = result.stderr.strip()
        if result.returncode != 0:
            return {"status": "error", "error": stderr or f"Exit code {result.returncode}"}
        if not stdout:
            return {"status": "error", "error": "脚本无输出"}
        return {"status": "error", "error": f"无法解析脚本输出: {stdout[:200]}"}
    except subprocess.TimeoutExpired:
        return {"status": "error", "error": "脚本执行超时(120s)"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


def save_upload(file_obj, prefix="upload"):
    """Save an uploaded file to workspace, return the path."""
    ext = Path(file_obj.filename).suffix if file_obj.filename else ".xlsx"
    fname = f"{prefix}_{uuid.uuid4().hex[:8]}{ext}"
    fpath = WORKSPACE / fname
    file_obj.save(str(fpath))
    return str(fpath)


# ─── Pipeline Management ──────────────────────────────────────────


def load_pipelines():
    """Load pipelines from JSON file, return list."""
    if PIPELINES_PATH.exists():
        try:
            with open(str(PIPELINES_PATH), "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []
    return []


def save_pipelines(pipelines):
    """Persist pipelines list to JSON."""
    with open(str(PIPELINES_PATH), "w", encoding="utf-8") as f:
        json.dump(pipelines, f, ensure_ascii=False, indent=2)


@app.route("/api/pipelines", methods=["GET"])
def api_list_pipelines():
    """List all pipelines, newest first."""
    with _pipelines_lock:
        pipelines = load_pipelines()
    pipelines.sort(key=lambda p: p.get("updated_at", ""), reverse=True)
    return jsonify({"status": "ok", "pipelines": pipelines})


@app.route("/api/pipelines", methods=["POST"])
def api_create_pipeline():
    """Create a new pipeline."""
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    scenario = (data.get("scenario") or "").strip()
    domain = (data.get("domain") or "").strip()

    if not name:
        return jsonify({"status": "error", "error": "流水线名称不能为空"})

    now = datetime.datetime.now().isoformat()
    pipeline = {
        "id": uuid.uuid4().hex[:12],
        "name": name,
        "scenario": scenario,
        "domain": domain or scenario,
        "current_step": 1,
        "step_status": {
            "1": "pending",
            "2": "pending",
            "3": "pending",
            "4": "pending",
        },
        "step_data": {},
        "created_at": now,
        "updated_at": now,
    }

    with _pipelines_lock:
        pipelines = load_pipelines()
        pipelines.append(pipeline)
        save_pipelines(pipelines)

    return jsonify({"status": "ok", "pipeline": pipeline})


@app.route("/api/pipelines/<pipeline_id>", methods=["GET"])
def api_get_pipeline(pipeline_id):
    """Get a single pipeline by ID."""
    with _pipelines_lock:
        pipelines = load_pipelines()
    for p in pipelines:
        if p["id"] == pipeline_id:
            return jsonify({"status": "ok", "pipeline": p})
    return jsonify({"status": "error", "error": "流水线不存在"})


@app.route("/api/pipelines/<pipeline_id>", methods=["PUT"])
def api_update_pipeline(pipeline_id):
    """Update a pipeline's step status/data."""
    data = request.get_json(force=True)

    with _pipelines_lock:
        pipelines = load_pipelines()
        target = None
        for p in pipelines:
            if p["id"] == pipeline_id:
                target = p
                break
        if not target:
            return jsonify({"status": "error", "error": "流水线不存在"})

        if "current_step" in data:
            target["current_step"] = data["current_step"]
        if "step_status" in data:
            target["step_status"].update(data["step_status"])
        if "step_data" in data:
            patch = data["step_data"]
            if not isinstance(patch, dict):
                return jsonify({"status": "error", "error": "step_data 格式错误"})
            err = validate_step_data_patch(patch)
            if err:
                return jsonify({"status": "error", "error": err})
            target["step_data"].update(patch)
        if "name" in data:
            target["name"] = data["name"]

        target["updated_at"] = datetime.datetime.now().isoformat()

        save_pipelines(pipelines)

    return jsonify({"status": "ok", "pipeline": target})


@app.route("/api/pipelines/<pipeline_id>", methods=["DELETE"])
def api_delete_pipeline(pipeline_id):
    """Delete a pipeline."""
    with _pipelines_lock:
        pipelines = load_pipelines()
        before = len(pipelines)
        pipelines = [p for p in pipelines if p["id"] != pipeline_id]
        if len(pipelines) == before:
            return jsonify({"status": "error", "error": "流水线不存在"})
        save_pipelines(pipelines)

    return jsonify({"status": "ok"})


@app.route("/api/pipelines/<pipeline_id>/clear", methods=["POST"])
def api_clear_pipeline(pipeline_id):
    """Clear a pipeline's step data and reset all steps to pending."""
    with _pipelines_lock:
        pipelines = load_pipelines()
        pipeline = next((p for p in pipelines if p["id"] == pipeline_id), None)
        if not pipeline:
            return jsonify({"status": "error", "error": "流水线不存在"})
        pipeline["step_data"] = {}
        pipeline["step_status"] = {str(i): "pending" for i in range(1, 5)}
        pipeline["current_step"] = 1
        pipeline["updated_at"] = datetime.datetime.now().isoformat()
        save_pipelines(pipelines)
    return jsonify({"status": "ok"})


@app.route("/api/pipelines/<pipeline_id>/rollback/<int:step>", methods=["POST"])
def api_rollback_pipeline(pipeline_id, step):
    """Roll back a pipeline to a previous step; reset downstream step status and outputs."""
    if step < 1 or step > 4:
        return jsonify({"status": "error", "error": "步骤号必须在 1-4 之间"})

    with _pipelines_lock:
        pipelines = load_pipelines()
        pipeline = next((p for p in pipelines if p["id"] == pipeline_id), None)
        if not pipeline:
            return jsonify({"status": "error", "error": "流水线不存在"})

        for s in range(step, 5):
            pipeline["step_status"][str(s)] = "pending"

        sd = pipeline.setdefault("step_data", {})
        for key in keys_to_clear_from_step(step):
            sd.pop(key, None)

        pipeline["current_step"] = step

        pipeline["updated_at"] = datetime.datetime.now().isoformat()
        save_pipelines(pipelines)

    return jsonify({"status": "ok", "pipeline": pipeline})


# ─── Step 1: Generate Template ────────────────────────────────────

@app.route("/api/step1/schema", methods=["GET"])
def api_step1_schema():
    """返回当前知识结构方案（scenario-schema.yaml）摘要，供 Step1 界面展示。"""
    from scenario_schema import load_scenario_schema, schema_summary

    if not SCHEMA_PATH.exists():
        return jsonify({"status": "error", "error": "未找到 scenario-schema.yaml"})
    schema = load_scenario_schema(SCHEMA_PATH)
    return jsonify({
        "status": "ok",
        "schema": schema_summary(schema, schema_path=SCHEMA_PATH),
    })


@app.route("/api/step1/templates", methods=["GET"])
def api_step1_templates():
    """列出 Step1 模板来源：默认 schema + 可选 legacy Excel 模板。"""
    from scenario_schema import load_scenario_schema, schema_summary
    from step1_template import list_default_step1_templates, find_default_step1_template

    schema_info = {}
    if SCHEMA_PATH.exists():
        schema_info = schema_summary(load_scenario_schema(SCHEMA_PATH), schema_path=SCHEMA_PATH)

    all_templates = list_default_step1_templates(SAMPLES_DIR)
    templates = [p for p in all_templates if "测试" in p.stem or "测试" in p.name]
    default_tpl = find_default_step1_template(SAMPLES_DIR)
    legacy_default = default_tpl.name if default_tpl else ""
    if default_tpl and not ("测试" in default_tpl.stem or "测试" in default_tpl.name):
        legacy_default = templates[0].name if templates else ""

    return jsonify({
        "status": "ok",
        "schema": schema_info,
        "default_mode": "schema",
        "templates": [
            {"name": p.name, "label": p.stem, "kind": "legacy"}
            for p in templates
        ],
        "default_template": "__schema__",
        "legacy_default_template": legacy_default,
    })


@app.route("/api/step1/generate", methods=["POST"])
def api_step1_generate():
    """Step1: 将场景四项填入萃取模板前四列，保留模板表结构（无 LLM）。"""
    from step1_template import (
        fill_scenario_skeleton,
        find_default_step1_template,
        list_default_step1_templates,
    )
    from scenario_schema import (
        enrich_knowledge_columns_for_markdown,
        load_scenario_schema,
        normalize_knowledge_columns,
        resolve_knowledge_columns_for_request,
    )
    from step1_markdown_builder import generate_markdown_skeleton
    from step1_schema_builder import generate_skeleton_from_schema

    scenario_name = request.form.get("scenario_name", "").strip()
    scenario_content = request.form.get("scenario_content", "").strip()
    sub_scenarios_json = request.form.get("sub_scenarios", "[]")
    pipeline_id = request.form.get("pipeline_id", "") or request.args.get("pipeline_id", "")
    template_mode = (request.form.get("template_mode", "") or "").strip().lower()
    output_format = (request.form.get("output_format", "excel") or "excel").strip().lower()
    if output_format not in {"excel", "markdown"}:
        output_format = "excel"
    knowledge_columns_json = request.form.get("knowledge_columns", "[]")

    if not scenario_name:
        return jsonify({"status": "error", "error": "场景名称不能为空"})
    _debug_log(
        "H1",
        "app_server.py:api_step1_generate",
        "step1 request received",
        {
            "has_pipeline_id": bool(pipeline_id),
            "sub_scenarios_json_len": len(sub_scenarios_json or ""),
            "template_mode": template_mode or "schema",
            "output_format": output_format,
        },
    )

    try:
        sub_scenarios = json.loads(sub_scenarios_json) if sub_scenarios_json else []
    except json.JSONDecodeError:
        sub_scenarios = []

    try:
        user_columns_raw = json.loads(knowledge_columns_json) if knowledge_columns_json else []
        if not isinstance(user_columns_raw, list):
            user_columns_raw = []
    except json.JSONDecodeError:
        user_columns_raw = []

    schema = load_scenario_schema(SCHEMA_PATH) if SCHEMA_PATH.exists() else {}
    knowledge_columns = resolve_knowledge_columns_for_request(schema, user_columns_raw)
    has_custom_columns = bool(normalize_knowledge_columns(user_columns_raw))
    columns_enriched = False
    user_knowledge_columns = list(knowledge_columns)
    if output_format == "markdown":
        knowledge_columns, columns_enriched, _substantive = enrich_knowledge_columns_for_markdown(
            user_columns_raw, schema
        )
    if not knowledge_columns and output_format != "upload":
        return jsonify({"status": "error", "error": "请至少定义一列知识字段"})

    template_path = None
    selected_default_template = request.form.get("default_template", "").strip()
    template_source = "schema"
    template_name = ""
    schema_meta = None
    md_name = ""
    md_download_url = ""
    upload = request.files.get("template")
    if upload and upload.filename:
        ext = os.path.splitext(upload.filename)[1].lower()
        if ext not in (".xlsx", ".xls"):
            return jsonify({"status": "error", "error": "模板仅支持 .xlsx / .xls 格式"})
        temp_name = f"upload_tpl_{uuid.uuid4().hex[:8]}{ext}"
        template_path = str(WORKSPACE / temp_name)
        upload.save(template_path)
        template_source = "upload"
        template_name = upload.filename
    elif template_mode == "legacy" and not has_custom_columns:
        default_tpl = None
        if selected_default_template and selected_default_template != "__schema__":
            candidates = {p.name: p for p in list_default_step1_templates(SAMPLES_DIR)}
            default_tpl = candidates.get(selected_default_template)
            if not default_tpl:
                return jsonify({"status": "error", "error": "所选 Excel 模板不存在，请刷新后重试"})
        else:
            default_tpl = find_default_step1_template(SAMPLES_DIR)

        if default_tpl:
            template_path = str(default_tpl)
            template_source = "legacy"
            template_name = default_tpl.name

    uid = uuid.uuid4().hex[:8]
    output_name = f"template_{uid}.xlsx"
    output_path = str(WORKSPACE / output_name)
    primary_download_name = output_name
    primary_download_url = "/downloads/" + output_name

    try:
        if template_path and os.path.exists(template_path):
            fill_result = fill_scenario_skeleton(
                template_path, output_path, scenario_name, scenario_content, sub_scenarios
            )
            knowledge_columns = knowledge_columns or []
            if output_format == "markdown":
                md_name = f"template_{uid}.md"
                md_path = WORKSPACE / md_name
                _excel_to_markdown_file(output_path, md_path, title=f"场景锚定骨架 · {scenario_name}")
                primary_download_name = md_name
                primary_download_url = "/downloads/" + md_name
                md_download_url = primary_download_url
                if template_source == "upload":
                    template_source = "upload_markdown"
                elif template_source == "legacy":
                    template_source = "legacy_markdown"
        elif output_format == "markdown":
            if not SCHEMA_PATH.exists():
                return jsonify({
                    "status": "error",
                    "error": "未找到 config/scenario-schema.yaml，无法生成骨架",
                })
            md_name = f"template_{uid}.md"
            md_path = str(WORKSPACE / md_name)
            generate_markdown_skeleton(
                md_path,
                scenario_name,
                scenario_content,
                sub_scenarios,
                knowledge_columns,
            )
            fill_result = generate_skeleton_from_schema(
                SCHEMA_PATH,
                output_path,
                scenario_name,
                scenario_content,
                sub_scenarios,
                knowledge_columns=knowledge_columns,
            )
            schema_meta = fill_result.get("schema") or {}
            template_source = "schema_markdown"
            template_name = "自定义列 · Markdown + Excel"
            primary_download_name = md_name
            primary_download_url = "/downloads/" + md_name
            md_download_url = primary_download_url
        else:
            if not SCHEMA_PATH.exists():
                return jsonify({
                    "status": "error",
                    "error": "未找到 config/scenario-schema.yaml，无法按结构方案生成模板",
                })
            fill_result = generate_skeleton_from_schema(
                SCHEMA_PATH,
                output_path,
                scenario_name,
                scenario_content,
                sub_scenarios,
                knowledge_columns=knowledge_columns,
            )
            schema_meta = fill_result.get("schema") or {}
            template_source = "schema"
            template_name = (
                f"{schema_meta.get('display_name', '自定义结构')} "
                f"({schema_meta.get('version', 'v1.0')})"
            ).strip()
    except ValueError as e:
        return jsonify({"status": "error", "error": str(e)})
    except Exception as e:
        return jsonify({"status": "error", "error": f"生成场景骨架失败: {str(e)}"})

    if schema_meta is None and SCHEMA_PATH.exists():
        from scenario_schema import schema_summary
        schema_meta = schema_summary(schema, schema_path=SCHEMA_PATH)
        schema_meta["knowledge_columns"] = knowledge_columns

    result = {
        "status": "ok",
        "scenario": scenario_name,
        "download_url": primary_download_url,
        "file_name": primary_download_name,
        "excel_file": output_name,
        "excel_download_url": "/downloads/" + output_name,
        "output_format": output_format,
        "knowledge_columns": knowledge_columns,
        "user_knowledge_columns": user_knowledge_columns,
        "columns_enriched": columns_enriched,
        "version": (schema_meta or {}).get("version", "v0.1"),
        "fields_info": fill_result.get("fields_info", []),
        "sub_scenario_count": fill_result.get("sub_scenario_count", len(sub_scenarios)),
        "template_source": template_source,
        "template_name": template_name,
        "schema": schema_meta,
    }
    if md_name:
        result["markdown_file"] = md_name
        result["markdown_download_url"] = md_download_url

    # 保存到 pipeline step_data
    saved_pipeline = None
    if pipeline_id:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    for key in keys_to_clear_from_step(1):
                        sd.pop(key, None)
                    sd["step1_output_file"] = output_name
                    sd["step1_download_url"] = "/downloads/" + output_name
                    sd["step1_template_source"] = template_source
                    sd["step1_template_name"] = template_name
                    sd["step1_output_format"] = output_format
                    sd["step1_knowledge_columns"] = knowledge_columns
                    sd["step1_user_knowledge_columns"] = user_knowledge_columns
                    sd["step1_columns_enriched"] = columns_enriched
                    if md_name:
                        sd["step1_md_file"] = md_name
                        sd["step1_md_download_url"] = md_download_url
                    else:
                        sd.pop("step1_md_file", None)
                        sd.pop("step1_md_download_url", None)
                    p["scenario"] = scenario_name
                    p["domain"] = scenario_name
                    save_pipelines(pipelines)
                    saved_pipeline = dict(p)
                    saved_pipeline["step_data"] = dict(p.get("step_data", {}))
                    break

    if saved_pipeline:
        result["pipeline"] = saved_pipeline
    return jsonify(result)


# ─── Step 2: Get Step 1 Output ─────────────────────────────────────

@app.route("/api/step2/prev_output", methods=["GET"])
def api_step2_prev_output():
    """获取当前流水线 Step1 的输出件信息，供 Step2 引用"""
    pipeline_id = request.args.get("pipeline_id", "")
    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    with _pipelines_lock:
        pipelines = load_pipelines()
        pipeline = None
        for p in pipelines:
            if p["id"] == pipeline_id:
                # Deep copy to avoid race conditions after lock release
                pipeline = dict(p)
                pipeline["step_data"] = dict(p.get("step_data", {}))
                break

    if not pipeline:
        return jsonify({"status": "error", "error": "流水线不存在"})

    step_data = pipeline.get("step_data", {})
    step1_file = step_data.get("step1_output_file", "")
    step1_url = step_data.get("step1_download_url", "")
    step1_md_file = step_data.get("step1_md_file", "")
    step1_md_url = step_data.get("step1_md_download_url", "")

    if not step1_file:
        return jsonify({
            "status": "ok",
            "has_output": False,
            "hint": "请先在「场景锚定」生成场景骨架（需已创建并进入流水线）",
        })

    file_path = str(WORKSPACE / step1_file)
    if not os.path.exists(file_path):
        return jsonify({
            "status": "ok",
            "has_output": False,
            "hint": "场景骨架文件已丢失，请回到场景锚定重新生成",
        })

    # Read Excel template structure for context
    fields_info = []
    if os.path.exists(file_path):
        try:
            with _safe_workbook(file_path) as wb:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = []
                    for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
                        if cell.value:
                            headers.append(str(cell.value))
                    if headers:
                        fields_info.append({"sheet": sheet_name, "headers": headers})
        except Exception:
            pass

    return jsonify({
        "status": "ok",
        "has_output": True,
        "file_name": step1_file,
        "download_url": step1_url,
        "markdown_file": step1_md_file,
        "markdown_download_url": step1_md_url,
        "output_format": step_data.get("step1_output_format", "excel"),
        "scenario": pipeline.get("scenario", ""),
        "domain": pipeline.get("domain", ""),
        "fields_info": fields_info,
    })


# ─── Validation (Steps 2, 3, 4) ──────────────────────────────────

# ─── Step 5: Compile ──────────────────────────────────────────────

@app.route("/api/step5/prev_output", methods=["GET"])
def api_step5_prev_output():
    """Get alignment output info for compile step."""
    pipeline_id = request.args.get("pipeline_id", "")
    with _pipelines_lock:
        pipelines = load_pipelines()
        pipeline = None
        for p in pipelines:
            if p["id"] == pipeline_id:
                # Deep copy to avoid race conditions after lock release
                pipeline = dict(p)
                pipeline["step_data"] = dict(p.get("step_data", {}))
                break
    if not pipeline:
        return jsonify({"status": "error", "error": "Pipeline not found"})

    step_data = pipeline.get("step_data", {})
    resolved_path, source_key = resolve_knowledge_workbook_path(WORKSPACE, step_data, purpose="compile")
    if not resolved_path:
        return jsonify({"status": "ok", "has_output": False})
    final_file = resolved_path.name

    result = {"status": "ok", "has_output": True, "file_name": final_file}
    if source_key == "step4_final_file":
        result["download_url"] = step_data.get("step4_download_url", f"/downloads/{final_file}")
        result["markdown_file"] = step_data.get("step4_md_file", "")
        result["markdown_download_url"] = step_data.get("step4_md_download_url", "")
    elif source_key == "step3_revision_file":
        result["download_url"] = step_data.get("step3_download_url", f"/downloads/{final_file}")
        result["markdown_file"] = step_data.get("step3_md_file", "")
        result["markdown_download_url"] = step_data.get("step3_md_download_url", "")
    else:
        result["download_url"] = step_data.get("step2_download_url", f"/downloads/{final_file}")
        result["markdown_file"] = step_data.get("step2_md_file", "")
        result["markdown_download_url"] = step_data.get("step2_md_download_url", "")

    # Read Excel fields info
    try:
        with _safe_workbook(str(resolved_path)) as wb:
            fields_info = []
            total_rows = 0
            for ws in wb.worksheets:
                headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
                row_count = max(0, ws.max_row - 1) if ws.max_row else 0
                total_rows += row_count
                fields_info.append({"sheet": ws.title, "headers": headers, "data_rows": row_count})
        result["fields_info"] = fields_info
        result["total_rows"] = total_rows
    except Exception:
        pass

    return jsonify(result)


@app.route("/api/step5/compile", methods=["POST"])
def api_step5_compile():
    """智能转化：生成思维链 / QA 对 / OpenClaw Skill 三类交付物。"""
    pipeline_id = request.form.get("pipeline_id", "")
    excel_file = request.files.get("excel")
    # region agent log
    _agent_debug_log(
        "run-2",
        "H6",
        "app_server.py:api_step5_compile:entry",
        "step5 compile request received",
        {
            "has_pipeline_id": bool(pipeline_id),
            "has_upload": bool(excel_file),
            "formats": request.form.get("formats", ""),
        },
    )
    # endregion

    input_path = None
    pipeline = None
    if pipeline_id:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    pipeline = dict(p)
                    break
    if pipeline_id and not excel_file and pipeline:
            step_data = pipeline.get("step_data", {})
            resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, step_data, purpose="compile")
            if resolved:
                input_path = str(resolved)

    if not input_path:
        if excel_file:
            input_path = save_upload(excel_file, prefix="compile")
        else:
            return jsonify({
                "status": "error",
                "error": "未找到可转化的知识稿。请先完成「知识萃取」，并在「知识对齐」节点生成 final_*.xlsx（或上传确认版 Excel）",
            })
    # region agent log
    _agent_debug_log(
        "run-2",
        "H6",
        "app_server.py:api_step5_compile:input",
        "step5 compile input resolved",
        {"input_basename": os.path.basename(input_path), "input_exists": os.path.exists(input_path)},
    )
    # endregion
    _debug_log(
        "H5",
        "app_server.py:api_step5_compile",
        "step5 compile start",
        {"has_pipeline_id": bool(pipeline_id), "input_basename": os.path.basename(input_path)},
    )

    from knowledge_delivery import excel_to_delivery_bundle

    output_dir = str(WORKSPACE / f"delivery_{uuid.uuid4().hex[:8]}")
    config_path = str(SCHEMA_PATH) if SCHEMA_PATH.exists() else ""

    pipeline_ctx = {}
    if pipeline_id and pipeline:
        sd = pipeline.get("step_data") or {}
        step1_form = sd.get("step1_form_data") or {}
        pipeline_ctx = {
            "scenario_name": pipeline.get("scenario") or step1_form.get("scenario_name", ""),
            "scenario_content": step1_form.get("scenario_content", ""),
            "sub_scenarios": step1_form.get("sub_scenarios") or [],
        }

    formats_raw = request.form.get("formats", "").strip()
    formats = [f.strip() for f in formats_raw.split(",") if f.strip()] if formats_raw else None

    try:
        result = excel_to_delivery_bundle(
            input_path, config_path, output_dir, pipeline_ctx or None, formats=formats
        )
    except Exception as e:
        # region agent log
        _agent_debug_log(
            "run-2",
            "H7",
            "app_server.py:api_step5_compile:exception",
            "step5 delivery generation raised exception",
            {"error": str(e)[:500], "input_basename": os.path.basename(input_path)},
        )
        # endregion
        _debug_log(
            "H5",
            "app_server.py:api_step5_compile",
            "step5 compile exception",
            {"error": str(e)[:300]},
        )
        return jsonify({"status": "error", "error": f"智能转化失败: {str(e)}"})

    if result.get("status") != "ok":
        err = result.get("error") or result.get("message") or "智能转化失败"
        # region agent log
        _agent_debug_log(
            "run-2",
            "H7",
            "app_server.py:api_step5_compile:result_error",
            "step5 delivery generation returned error",
            {"error": str(err)[:500], "input_basename": os.path.basename(input_path)},
        )
        # endregion
        _debug_log(
            "H5",
            "app_server.py:api_step5_compile",
            "step5 compile failed",
            {"error": str(err)[:300]},
        )
        return jsonify({"status": "error", "error": err})

    artifacts = result.get("artifacts") or {}
    downloads = {}

    def _publish_artifact(key: str, src_path: str, prefix: str, ext: str):
        if not src_path or not os.path.isfile(src_path):
            return None
        name = f"{prefix}_{uuid.uuid4().hex[:8]}{ext}"
        dest = WORKSPACE / name
        shutil.copy2(src_path, str(dest))
        info = {
            "file_name": name,
            "download_url": "/downloads/" + name,
        }
        meta = artifacts.get(key) or {}
        if meta.get("count") is not None:
            info["count"] = meta["count"]
        if meta.get("label"):
            info["label"] = meta["label"]
        downloads[key] = info
        return name, "/downloads/" + name

    cot_pub = _publish_artifact(
        "cot", (artifacts.get("cot") or {}).get("path"), "COT", ".md"
    )
    qa_pub = _publish_artifact(
        "qa", (artifacts.get("qa") or {}).get("path"), "QA", ".json"
    )
    qa_md_pub = _publish_artifact(
        "qa_md", (artifacts.get("qa") or {}).get("markdown_path"), "QA", ".md"
    )
    skill_pub = _publish_artifact(
        "skill", (artifacts.get("skill") or {}).get("path"), "SKILL", ".md"
    )
    manifest_pub = _publish_artifact(
        "openclaw_manifest",
        (artifacts.get("skill") or {}).get("manifest_path"),
        "openclaw",
        ".json",
    )

    result["artifacts_download"] = downloads
    if skill_pub:
        result["download_url"] = skill_pub[1]
        result["download_name"] = skill_pub[0]
    if cot_pub:
        result["cot_download_url"] = cot_pub[1]
        result["cot_download_name"] = cot_pub[0]
    if qa_pub:
        result["qa_download_url"] = qa_pub[1]
        result["qa_download_name"] = qa_pub[0]
    if qa_md_pub:
        result["qa_md_download_url"] = qa_md_pub[1]
        result["qa_md_download_name"] = qa_md_pub[0]
    if manifest_pub:
        result["openclaw_manifest_url"] = manifest_pub[1]
        result["openclaw_manifest_name"] = manifest_pub[0]

    if pipeline_id:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    if skill_pub:
                        sd["step5_skill_file"] = skill_pub[0]
                        sd["step5_download_url"] = skill_pub[1]
                    if cot_pub:
                        sd["step5_cot_file"] = cot_pub[0]
                        sd["step5_cot_download_url"] = cot_pub[1]
                    if qa_pub:
                        sd["step5_qa_file"] = qa_pub[0]
                        sd["step5_qa_download_url"] = qa_pub[1]
                    if qa_md_pub:
                        sd["step5_qa_md_file"] = qa_md_pub[0]
                        sd["step5_qa_md_download_url"] = qa_md_pub[1]
                    if manifest_pub:
                        sd["step5_openclaw_manifest_file"] = manifest_pub[0]
                        sd["step5_openclaw_manifest_url"] = manifest_pub[1]
                    save_pipelines(pipelines)
                    break
    return jsonify(result)


@app.route("/api/step5/quality", methods=["POST"])
def api_step5_quality():
    """Generate quality report — auto-reads Step4 output if pipeline_id provided."""
    pipeline_id = request.form.get("pipeline_id", "")
    excel_file = request.files.get("excel")

    input_path = None
    if pipeline_id and not excel_file:
        with _pipelines_lock:
            pipelines = load_pipelines()
            pipeline = None
            for p in pipelines:
                if p["id"] == pipeline_id:
                    pipeline = dict(p)
                    break
        if pipeline:
            step_data = pipeline.get("step_data", {})
            resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, step_data, purpose="compile")
            if resolved:
                input_path = str(resolved)

    if not input_path:
        if excel_file:
            input_path = save_upload(excel_file, prefix="quality")
        else:
            return jsonify({
                "status": "error",
                "error": "未找到可分析的知识稿。请先完成「知识对齐」生成 final_*.xlsx（或上传确认版 Excel）",
            })

    report_name = f"quality_report_{uuid.uuid4().hex[:8]}.md"
    report_path = str(WORKSPACE / report_name)

    args = ["--input", input_path, "--output", report_path]
    if SCHEMA_PATH.exists():
        args.extend(["--config", str(SCHEMA_PATH)])

    result = run_script("quality_report.py", args)

    if os.path.exists(report_path):
        result["download_url"] = "/downloads/" + report_name
        # Save to pipeline step_data
        if pipeline_id:
            with _pipelines_lock:
                pipelines = load_pipelines()
                for p in pipelines:
                    if p["id"] == pipeline_id:
                        p.setdefault("step_data", {})["step5_quality_file"] = report_name
                        p.setdefault("step_data", {})["step5_quality_url"] = "/downloads/" + report_name
                        save_pipelines(pipelines)
                        break

    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════
# LLM API Endpoints
# ═══════════════════════════════════════════════════════════════════

@app.route("/api/llm/models", methods=["GET"])
def api_llm_list_models():
    """List all available models (presets + custom); API keys are masked."""
    models = load_llm_config()
    result = [_sanitize_model_for_client(m) for m in models]
    return jsonify({"status": "ok", "models": result})


def _parse_model_payload(data, existing=None, require_api_key=True):
    """Parse and validate model fields from request JSON."""
    name = (data.get("name") or (existing or {}).get("name") or "").strip()
    raw_url = (data.get("url") or (existing or {}).get("url") or "").strip()
    model = (data.get("model") or (existing or {}).get("model") or "").strip()
    api_key = (data.get("api_key") or "").strip()
    if not api_key and existing:
        api_key = (existing.get("api_key") or "").strip()
    api_type = (data.get("api_type") or (existing or {}).get("api_type") or API_TYPE_OPENAI).strip().lower()
    if api_type not in (API_TYPE_OPENAI, API_TYPE_CCB):
        return None, f"api_type 无效，应为 {API_TYPE_OPENAI} 或 {API_TYPE_CCB}"
    url = normalize_llm_url(raw_url, api_type)
    if not all([name, url, model]):
        return None, "name/url/model 均为必填"
    if require_api_key and not api_key:
        return None, "api_key 为必填"
    tx_code = (data.get("tx_code") or (existing or {}).get("tx_code") or "").strip()
    sec_node_no = (data.get("sec_node_no") or (existing or {}).get("sec_node_no") or "").strip()
    if api_type == API_TYPE_CCB:
        if not tx_code or not sec_node_no:
            return None, "建行接口需填写 Tx-Code 与 Sec-Node-No"
    parsed = {
        "name": name,
        "url": url,
        "model": model,
        "api_key": api_key,
        "api_type": api_type,
        "max_tokens": data.get("max_tokens", (existing or {}).get("max_tokens", 4096)),
        "temperature": data.get("temperature", (existing or {}).get("temperature", 0.7)),
        "description": data.get("description", (existing or {}).get("description", "")),
    }
    if api_type == API_TYPE_CCB:
        parsed["tx_code"] = tx_code
        parsed["sec_node_no"] = sec_node_no
        fst = (data.get("fst_attr_rmrk") or (existing or {}).get("fst_attr_rmrk") or "").strip()
        if fst:
            parsed["fst_attr_rmrk"] = fst
    return parsed, None


@app.route("/api/llm/models", methods=["POST"])
def api_llm_add_model():
    """Add a custom model."""
    data = request.get_json(force=True)
    parsed, err = _parse_model_payload(data, require_api_key=True)
    if err:
        return jsonify({"status": "error", "error": err})
    name = parsed["name"]

    with _models_lock:
        custom = []
        if CUSTOM_MODELS_PATH.exists():
            try:
                with open(str(CUSTOM_MODELS_PATH), "r", encoding="utf-8") as f:
                    custom = json.load(f)
            except (json.JSONDecodeError, IOError):
                custom = []

        existing_names = [m["name"] for m in load_llm_config()]
        if name in existing_names:
            return jsonify({"status": "error", "error": f"模型名称 '{name}' 已存在"})

        custom.append(parsed)
        save_custom_models(custom)

    return jsonify({"status": "ok", "message": f"模型 '{name}' 已添加"})


@app.route("/api/llm/models/<path:model_name>", methods=["GET"])
def api_llm_get_model(model_name):
    """Get model config for editing (API key not returned)."""
    model_cfg = get_model_by_name(model_name)
    if not model_cfg:
        return jsonify({"status": "error", "error": f"模型 '{model_name}' 不存在"})
    d = _sanitize_model_for_client(model_cfg)
    d["is_preset"] = model_cfg.get("is_preset", False)
    return jsonify({"status": "ok", "model": d})


@app.route("/api/llm/models/<path:model_name>", methods=["PUT"])
def api_llm_update_model(model_name):
    """Update a preset (saved as override) or custom model."""
    existing = get_model_by_name(model_name)
    if not existing:
        return jsonify({"status": "error", "error": f"模型 '{model_name}' 不存在"})

    data = request.get_json(force=True)
    parsed, err = _parse_model_payload(data, existing=existing, require_api_key=False)
    if err:
        return jsonify({"status": "error", "error": err})

    new_name = parsed["name"]
    if new_name != model_name:
        return jsonify({"status": "error", "error": "暂不支持修改模型名称，请删除后重新添加"})

    save_fields = {k: v for k, v in parsed.items() if k != "name"}

    with _models_lock:
        if existing.get("is_preset"):
            overrides = load_preset_overrides()
            overrides[model_name] = save_fields
            save_preset_overrides(overrides)
        else:
            custom = []
            if CUSTOM_MODELS_PATH.exists():
                try:
                    with open(str(CUSTOM_MODELS_PATH), "r", encoding="utf-8") as f:
                        custom = json.load(f)
                except (json.JSONDecodeError, IOError):
                    custom = []
            updated = False
            for i, m in enumerate(custom):
                if m["name"] == model_name:
                    custom[i] = parsed
                    updated = True
                    break
            if not updated:
                return jsonify({"status": "error", "error": f"自定义模型 '{model_name}' 不存在"})
            save_custom_models(custom)

    return jsonify({"status": "ok", "message": f"模型 '{model_name}' 已更新"})


@app.route("/api/llm/models/<path:model_name>", methods=["DELETE"])
def api_llm_delete_model(model_name):
    """Delete a custom model (presets cannot be deleted)."""
    with _models_lock:
        custom = []
        if CUSTOM_MODELS_PATH.exists():
            try:
                with open(str(CUSTOM_MODELS_PATH), "r", encoding="utf-8") as f:
                    custom = json.load(f)
            except (json.JSONDecodeError, IOError):
                custom = []

        before = len(custom)
        custom = [m for m in custom if m["name"] != model_name]
        if len(custom) == before:
            return jsonify({"status": "error", "error": f"自定义模型 '{model_name}' 不存在或为预设模型不可删除"})

        save_custom_models(custom)

    return jsonify({"status": "ok", "message": f"模型 '{model_name}' 已删除"})


@app.route("/api/llm/test", methods=["POST"])
def api_llm_test():
    """Test connection to a model."""
    data = request.get_json(force=True)
    model_name = data.get("name", "")
    model_cfg = get_model_by_name(model_name)
    if not model_cfg:
        return jsonify({"status": "error", "error": f"模型 '{model_name}' 不存在"})

    try:
        test_cfg = dict(model_cfg)
        test_cfg["timeout"] = min(int(test_cfg.get("timeout", 300)), 20)
        result = call_llm(
            test_cfg,
            [{"role": "user", "content": "Hi"}],
            stream=False,
            max_tokens=10,
        )
        content = extract_assistant_content(result) if isinstance(result, dict) else ""
        return jsonify({"status": "ok", "message": f"连接成功，模型回复: {content[:50]}"})
    except LlmApiError as e:
        return jsonify({"status": "error", "error": str(e)})
    except Exception as e:
        return jsonify({"status": "error", "error": f"连接失败: {str(e)}"})


@app.route("/api/llm/stream-test", methods=["POST"])
def api_llm_stream_test():
    """Stream-test LLM connection; forwards deltas as SSE to the browser."""
    data = request.get_json(force=True) or {}
    model_name = data.get("name", "")
    prompt = (data.get("prompt") or "你好，请用一句话介绍你自己。").strip()
    model_cfg = get_model_by_name(model_name)
    if not model_cfg:
        return jsonify({"status": "error", "error": f"模型 '{model_name}' 不存在"})

    messages = [{"role": "user", "content": prompt}]

    def generate():
        try:
            test_cfg = dict(model_cfg)
            test_cfg["timeout"] = min(int(test_cfg.get("timeout", 300)), 120)
            resp = call_llm(test_cfg, messages, stream=True, max_tokens=256)
            for delta in iter_llm_stream(test_cfg, resp):
                payload = json.dumps({"delta": delta}, ensure_ascii=False)
                yield f"data: {payload}\n\n"
            yield f"data: {json.dumps({'done': True}, ensure_ascii=False)}\n\n"
        except LlmApiError as e:
            err = json.dumps({"error": str(e)}, ensure_ascii=False)
            yield f"data: {err}\n\n"
        except Exception as e:
            err = json.dumps({"error": f"流式连接失败: {str(e)}"}, ensure_ascii=False)
            yield f"data: {err}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─── Excel Online Editor API ──────────────────────────────────────

def _trim_excel_rows(rows, max_rows=200, max_cols=40):
    """裁剪尾部空行空列，避免 Luckysheet 渲染超大稀疏表卡顿。"""
    if not rows:
        return [[]]
    trimmed = []
    for row in rows[:max_rows]:
        trimmed.append([cell if cell is not None else "" for cell in (row or [])])
    last_r = len(trimmed) - 1
    while last_r > 0:
        if any(str(c or "").strip() for c in trimmed[last_r]):
            break
        last_r -= 1
    trimmed = trimmed[: last_r + 1]
    if not trimmed:
        return [[]]
    last_c = 0
    for row in trimmed:
        for i, c in enumerate(row):
            if str(c or "").strip():
                last_c = max(last_c, i)
    last_c = min(last_c, max_cols - 1)
    out = []
    for row in trimmed:
        r = list(row[: last_c + 1])
        while len(r) <= last_c:
            r.append("")
        out.append(r)
    return out or [[]]


@app.route("/api/excel/read", methods=["POST"])
def api_excel_read():
    """Read Excel file and return structured data for online editing."""
    input_path = None

    # Support both file upload and file_name from JSON/form
    excel_file = request.files.get("excel")
    if excel_file:
        input_path = save_upload(excel_file, prefix="edit_read")
    else:
        # Try JSON body or form data with file_name
        data = request.get_json(silent=True) or request.form.to_dict()
        file_name = data.get("file_name", "")
        if file_name:
            resolved = safe_workspace_path(WORKSPACE, file_name, must_exist=True)
            if not resolved:
                return jsonify({"status": "error", "error": f"文件不存在或路径非法: {file_name}"})
            input_path = str(resolved)

    if not input_path:
        return jsonify({"status": "error", "error": "请上传 Excel 文件或提供文件名"})

    wb = None
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = []
            for row in ws.iter_rows(values_only=True):
                raw_rows.append([cell if cell is not None else "" for cell in row])
            rows = _trim_excel_rows(raw_rows)
            merges = []
            for merge_range in list(ws.merged_cells.ranges):
                merges.append({
                    "r": merge_range.min_row - 1,
                    "c": merge_range.min_col - 1,
                    "rs": merge_range.max_row - merge_range.min_row + 1,
                    "cs": merge_range.max_col - merge_range.min_col + 1,
                })
            sheets[sheet_name] = {"rows": rows, "merges": merges}

        # Store the file path in pipeline data for later save
        pipeline_id = request.form.get("pipeline_id", "") or (request.get_json(silent=True) or {}).get("pipeline_id", "")
        step = request.form.get("step", "3") or (request.get_json(silent=True) or {}).get("step", "3")
        if pipeline_id:
            with _pipelines_lock:
                pipelines = load_pipelines()
                for p in pipelines:
                    if p["id"] == pipeline_id:
                        sd = p.setdefault("step_data", {})
                        base = basename_only(input_path)
                        sd[f"step{step}_excel_path"] = base
                        if str(step) == "1" and is_step1_filename(base):
                            sd["step1_output_file"] = base
                            sd["step1_download_url"] = "/downloads/" + base
                        save_pipelines(pipelines)
                        break

        return jsonify({"status": "ok", "sheets": sheets, "file_path": basename_only(input_path)})
    except Exception as e:
        return jsonify({"status": "error", "error": f"读取 Excel 失败: {str(e)}"})
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


@app.route("/api/excel/save", methods=["POST"])
def api_excel_save():
    """Save edited Excel data back to file and return download URL."""
    data = request.get_json(force=True)
    file_path = data.get("file_path", "")
    file_name = data.get("file_name", "")
    sheets = data.get("sheets", {})
    pipeline_id = data.get("pipeline_id", "")
    step = data.get("step", "3")

    resolved = resolve_client_excel_path(WORKSPACE, file_path, file_name)
    if not resolved:
        return jsonify({"status": "error", "error": "源文件不存在或路径非法，请重新打开编辑"})
    file_path = str(resolved)

    if not sheets:
        return jsonify({"status": "error", "error": "无数据可保存"})

    wb = None
    try:
        wb = openpyxl.load_workbook(file_path)

        def _sheet_payload(raw):
            if isinstance(raw, list):
                return raw, []
            if isinstance(raw, dict):
                return raw.get("rows", []), raw.get("merges", [])
            return [], []

        for sheet_name, raw in sheets.items():
            rows, merges = _sheet_payload(raw)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]

            for merge_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merge_range))

            if ws.max_row:
                ws.delete_rows(1, ws.max_row)

            for r_idx, row_data in enumerate(rows, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value if value != "" else None)

            for m in merges:
                rs = int(m.get("rs", 1))
                cs = int(m.get("cs", 1))
                if rs > 1 or cs > 1:
                    r0 = int(m.get("r", 0)) + 1
                    c0 = int(m.get("c", 0)) + 1
                    ws.merge_cells(
                        start_row=r0,
                        start_column=c0,
                        end_row=r0 + rs - 1,
                        end_column=c0 + cs - 1,
                    )

        # Save to a new file for download and overwrite the original
        save_name = f"edited_step{step}_{uuid.uuid4().hex[:8]}.xlsx"
        save_path = str(WORKSPACE / save_name)
        wb.save(save_path)
        wb.save(file_path)

        # Update pipeline step_data
        if pipeline_id:
            with _pipelines_lock:
                pipelines = load_pipelines()
                for p in pipelines:
                    if p["id"] == pipeline_id:
                        sd = p.setdefault("step_data", {})
                        base = basename_only(file_path)
                        sd[f"step{step}_excel_path"] = base
                        if str(step) == "1" and is_step1_filename(base):
                            sd["step1_output_file"] = base
                            sd["step1_download_url"] = "/downloads/" + base
                        elif str(step) == "2" and is_step2_preextract_filename(base):
                            sd["step2_output_file"] = base
                            sd["step2_download_url"] = "/downloads/" + base
                            for stale in ("step2_preview_name", "step2_preview_url", "step2_download"):
                                sd.pop(stale, None)
                        elif str(step) == "3" and is_step3_revision_filename(base):
                            sd["step3_revision_file"] = base
                            sd["step3_download_url"] = "/downloads/" + save_name
                        elif str(step) == "4" and is_step4_final_filename(base):
                            sd["step4_final_file"] = base
                            sd["step4_download_url"] = "/downloads/" + save_name
                        elif str(step) == "4" and is_step3_revision_filename(base):
                            sd["step3_revision_file"] = base
                            sd["step3_download_url"] = "/downloads/" + save_name
                        save_pipelines(pipelines)
                        break

        return jsonify({
            "status": "ok",
            "download_url": "/downloads/" + save_name,
            "file_path": basename_only(file_path),
            "file_name": basename_only(file_path),
            "message": "保存成功"
        })
    except Exception as e:
        return jsonify({"status": "error", "error": f"保存 Excel 失败: {str(e)}"})
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


# ─── Skill Registry ───────────────────────────────────────────────

SKILL_REGISTRY = {
    "knowledge-extraction": {
        "id": "knowledge-extraction",
        "name": "知识萃取",
        "version": "1.0.0",
        "description": "从非结构化文档中自动提取结构化知识条目，生成知识萃取稿。支持文档模式和案例复盘模式两种输入方式。",
        "detailed_description": (
            "这是隐性知识显性化的核心引擎。系统会将您上传的制度文件、操作手册、案例复盘等非结构化文本，"
            "通过大语言模型自动抽取为结构化的知识条目，逐条填入场景锚定模板的对应列中。\n\n"
            "**两种工作模式**：\n"
            "- 📄 文档萃取：上传制度文件/操作手册（TXT/MD/DOCX/PDF），系统自动分段、识别关键知识并填入模板\n"
            "- 📋 案例复盘：填写结构化复盘表单（背景→判断→结果→重来→习惯），系统不仅提取可执行知识，还专门检测反模式、破例逻辑和直觉信号\n\n"
            "**三种萃取风格**：\n"
            "标准（8-22条，平衡覆盖）、深度（12-40条，完整覆盖所有字段）、精简（5-10条，仅保留高价值高置信条目）"
        ),
        "usage_guide": (
            "1. 先在「场景锚定」完成知识结构定义（确定Excel模板的列名）\n"
            "2. 选择输入模式：上传文档或切换到案例复盘表单\n"
            "3. 选择萃取风格（不确定时用「标准萃取」）\n"
            "4. 选择模型和Skill，点击「执行」\n"
            "5. 等待10-60秒（取决于文档长度），下载萃取Excel或Markdown"
        ),
        "input_example": "上传一份《科技企业普惠贷款业务操作指引》.txt 或填写一个芯片设计企业流水异常的案例复盘",
        "output_example": "生成 preextract_*.xlsx，包含按模板列填充的知识条目，每条标注了置信度。案例复盘模式额外输出「隐性信号」和「信号类型」字段",
        "business_value": (
            "将散落在文档和个人经验中的知识，转化为团队可共享、可检索、可执行的结构化知识资产。"
            "案例复盘模式特别适合把老信贷员的「踩坑故事」转化为团队的「避坑指南」。"
        ),
        "applicable_scenarios": [
            "新业务制度发布后的知识入库",
            "老员工离职前的经验抢救",
            "案例复盘会后的结构化沉淀",
            "部门操作手册的知识化改造",
        ],
        "limitations": [
            "无法自动判断知识的正确性——萃取的可信度取决于输入文档的质量",
            "对极度专业的领域术语可能需要人工校验",
            "案例复盘模式需要专家愿意花时间填写结构化表单",
        ],
        "triggers": ["上传文档", "上传制度文件", "知识萃取", "知识预萃", "提取知识条目", "案例复盘"],
        "supported_formats": ["TXT", "MD", "DOCX", "PDF"],
        "max_file_size_mb": 50,
        "output_styles": ["标准萃取（8-22条）", "深度萃取（12-40条）", "精简萃取（5-10条）"],
        "capabilities": [
            "文档智能解析与分段",
            "知识条目结构化提取",
            "字段自动填充（知识分类/判断逻辑/适用条件）",
            "反模式与踩坑提示识别",
            "置信度自动评估",
            "案例复盘中的隐性信号检测（反模式/破例/直觉/关系依赖）",
        ],
        "related_step": 2,
        "enabled": True,
    },
    "knowledge-revision": {
        "id": "knowledge-revision",
        "name": "知识对齐",
        "version": "1.1.0",
        "description": "融合专家修订意见与AI分析，逐条审核修订建议，支持采纳/驳回/编辑，一键生成可发布的对齐稿。",
        "detailed_description": (
            "知识萃取后，专家面对的不是空白Excel，而是一份「待审稿」。本Skill将专家的修订意见解析为结构化的修订指令，"
            "逐条标注修改/删除/新增/补充，用不同颜色高亮显示变更，并在Excel右侧自动生成修订信息列（修订状态/原始内容/修订内容/修订说明/修订时间）。\n\n"
            "**关键设计理念**：专家不是来「审稿」的，是来「对话」的。因此系统在专家每次采纳修订建议时，自动弹出「隐性注释追问卡片」，"
            "引导专家分享修订背后的经验判断——这些隐性注释将作为知识的「上下文层」永久保存。\n\n"
            "**无意见直通**：如果专家对萃取稿满意、无需修订，系统会自动将萃取稿确认为对齐稿，无需额外操作。"
        ),
        "usage_guide": (
            "1. 在输入区查看Step2生成的萃取稿预览\n"
            "2. 在「专家输入」中填写修订意见（或上传专家纪要文件），也可以留空直通\n"
            "3. 点击「发送并智能修订」，系统生成修订建议列表\n"
            "4. 逐条审核：采纳/驳回/编辑每条建议\n"
            "5. 采纳时弹出的追问卡片可选填——这是沉淀隐性知识的最佳时机\n"
            "6. 点击「生成对齐稿」，生成 final_*.xlsx"
        ),
        "input_example": "专家意见：「第5条知识中，成立满2年的条件对芯片行业过于严格，建议改为成立满1年但创始人需来自中科院或985高校相关专业」",
        "output_example": "生成 final_*.xlsx，修订单元格用颜色标注（黄=修改/红=删除/绿=新增/蓝=补充），右侧增加修订信息列。专家填写的隐性注释保存在 step_data 中。",
        "business_value": (
            "传统方式下，专家的修订意见散落在微信、邮件、会议纪要中，修订理由（为什么这么改）往往丢失。"
            "本Skill把修订过程结构化，并将修订背后的隐性判断以「隐性注释」形式保存——这是知识库中最珍贵的一层信息。"
        ),
        "applicable_scenarios": [
            "专家审核萃取稿并提出修改意见",
            "多人会议纪要需要合并为统一修订",
            "合规/风控部门对知识条目进行合规审查",
            "老员工对新人萃取的知识进行校准",
        ],
        "limitations": [
            "修订建议的质量依赖于专家意见的明确程度——模糊的意见会产生模糊的修订",
            "不适用于大量增删的场景（超过80条修订建议审核体验变差）",
            "隐性注释追问卡片依赖专家配合——如果专家跳过，这部分价值就丢失了",
        ],
        "triggers": ["知识对齐", "专家对齐", "最终确认", "生成对齐稿"],
        "supported_formats": ["XLSX"],
        "max_file_size_mb": 50,
        "output_styles": ["标准修订", "严格修订（保守）", "宽松修订（积极）"],
        "capabilities": [
            "Excel结构解析与比对",
            "修订点智能识别（修改/删除/新增/补充）",
            "多类型标注（颜色+状态列+说明）",
            "版本追踪与备份",
            "无意见直通（不强制提交专家意见）",
            "隐性注释追问卡片（自动捕获修订背后的经验判断）",
        ],
        "related_step": 3,
        "enabled": True,
    },
    "knowledge-pattern-mining": {
        "id": "knowledge-pattern-mining",
        "name": "跨案例模式发现",
        "version": "1.0.0",
        "description": "对多个案例复盘进行交叉分析，发现反复出现的隐性信号、高频反模式和系统性风险盲区。",
        "detailed_description": (
            "单个案例复盘能沉淀一个故事，但看不出规律。当你积累了3个、5个、10个案例复盘后，"
            "本Skill会像一位资深风控专家一样，横跨所有案例寻找「反复出现的信号」。\n\n"
            "**五步分析流程**：\n"
            "1. 识别重复出现的预警信号 → 哪些信号在多个案例中反复出现？\n"
            "2. 发现系统性风险盲区 → 为什么按标准流程操作仍然没能提前发现风险？\n"
            "3. 提炼跨案例隐性知识 → 从案例之间的共同模式中提炼可操作的新知识\n"
            "4. 信号优先级排序 → 按「出现频率 × 损失严重度」给所有信号排序\n"
            "5. 生成行动建议 → 三条具体的、可落地的改进行动\n\n"
            "**核心价值**：一个信贷员踩过的坑是故事，三个信贷员在不同场景下踩过类似的坑就是系统性风险盲区。"
        ),
        "usage_guide": (
            "1. 准备至少2个案例复盘文本（可以从之前的Step2案例复盘输出中获取）\n"
            "2. 将所有案例文本粘贴到输入框，或用多文件上传同时选择多个案例文件\n"
            "3. 选择模型，点击执行\n"
            "4. 系统自动进行交叉分析，生成跨案例模式发现报告\n"
            "5. 下载Markdown报告，在团队会议中讨论发现的模式"
        ),
        "input_example": "输入2-5个案例复盘文本，每个案例至少包含背景、判断、结果、反思四个部分。案例之间最好来自同一业务领域（如都是科技企业贷款案例），这样发现的模式更有针对性。",
        "output_example": (
            "生成 pattern_mining_*.md 报告，包含五部分：\n"
            "一、重复出现的预警信号（含出现频次和最早信号标记）\n"
            "二、系统性风险盲区\n"
            "三、跨案例隐性知识（每条知识的执行来源和具体做法）\n"
            "四、信号优先级排序\n"
            "五、行动建议"
        ),
        "business_value": (
            "这是从「经验」到「规律」的关键一步。单独一个案例容易被当作「特殊情况」忽略，"
            "但5个案例同时指向同一个信号时，它就是必须被修正的流程缺陷。"
        ),
        "applicable_scenarios": [
            "季度/年度案例复盘总结会议的前置分析",
            "新员工培训材料的案例化改造",
            "风控流程优化的数据支撑",
            "跨部门/跨支行的经验共享和教训总结",
        ],
        "limitations": [
            "案例数量太少（<2个）时分析价值有限，建议至少积累3个以上相关案例再使用",
            "案例之间的相关性影响分析质量——完全无关的案例放在一起会产生噪音",
            "分析结果需要人工验证——LLM可能发现「统计学相关」但「业务逻辑无关」的模式",
        ],
        "triggers": ["案例复盘", "跨案例分析", "模式发现", "信号挖掘", "交叉验证"],
        "supported_formats": ["TXT", "MD", "JSON"],
        "max_file_size_mb": 100,
        "output_styles": ["标准模式发现", "深度挖掘"],
        "capabilities": [
            "多案例交叉对比分析",
            "高频预警信号自动识别与排名",
            "跨案例隐性知识的涌现模式发现",
            "系统性风险盲区检测",
            "生成跨案例洞察报告（含优先级排序和行动建议）",
        ],
        "related_step": 2,
        "enabled": True,
    },
    "knowledge-gap-analysis": {
        "id": "knowledge-gap-analysis",
        "name": "知识盲区检测",
        "version": "1.0.0",
        "description": "基于场景Schema与已萃取的知识条目，检测知识覆盖的空白区域，识别「应该知道但还不知道」的内容。",
        "detailed_description": (
            "风控最大的敌人不是信息不足，而是「不知道自己有盲区」。本Skill对比你在场景锚定中定义的知识结构（Schema）"
            "和实际萃取的知识条目，逐列计算填充率，自动标记那些「定义了但没填充」的盲区列。\n\n"
            "**双重分析**：\n"
            "- 程序化统计：逐列计算填充率、空值率、样本内容，不需要LLM参与\n"
            "- LLM解读：基于盲区统计结果，生成3-5条定向补全建议，包括「应该补充什么」「建议找谁」「为什么高风险」\n\n"
            "**典型发现**：Schema定义了「反模式/踩坑提示」列，但实际85%的条目此列为空——说明知识库缺乏风险防范信息，是高危盲区。"
        ),
        "usage_guide": (
            "1. 确保已完成Step2知识萃取（有关联的知识Excel文件）\n"
            "2. 在Step2的Skill下拉中选择「知识盲区检测」\n"
            "3. 选择模型（用于生成解读建议），点击执行\n"
            "4. 查看盲区报告，标记高优先级盲区\n"
            "5. 根据报告中的「补全建议」，定向邀请相关领域专家补充知识"
        ),
        "input_example": "关联当前流水线（自动读取Step1 Schema和Step2萃取结果），或直接上传一份知识Excel文件",
        "output_example": (
            "生成 gap_analysis_*.md 报告，包含三部分：\n"
            "一、各列填充率统计表\n"
            "二、盲区清单（按严重度排序，高危=填充率<20%，中危=填充率<50%）\n"
            "三、补全建议（LLM生成的3-5条具体可操作建议）"
        ),
        "business_value": (
            "明确了「缺什么」等于完成了知识补充工作量的一半。这份报告可以作为团队知识建设的行动计划——"
            "本季度重点补全反模式列、下季度补全判断逻辑列。"
        ),
        "applicable_scenarios": [
            "知识库建设初期——快速评估哪些领域是重灾区",
            "新业务领域扩展——确认新领域知识覆盖是否完整",
            "季度知识质量巡检——定期检测知识库健康状况",
            "专家访谈前的准备——明确哪些问题需要向专家定向请教",
        ],
        "limitations": [
            "只能检测Schema中已定义的列——如果Schema本身就漏掉了重要维度，盲区检测也发现不了",
            "填充率不等于质量——即使字段非空，也不代表内容是高质量的",
            "建议在知识库达到一定规模（50+条目）后再使用，小样本统计意义有限",
        ],
        "triggers": ["知识盲区", "盲区检测", "覆盖分析", "知识缺口", "gap分析"],
        "supported_formats": ["XLSX"],
        "max_file_size_mb": 50,
        "output_styles": ["标准检测", "深度检测"],
        "capabilities": [
            "Schema字段覆盖率统计",
            "知识分类分布不均检测",
            "空值/低质量字段自动标记",
            "定向专家采访建议生成",
            "知行差距识别（制度 vs. 实际做法）",
        ],
        "related_step": 2,
        "enabled": True,
    },
    "knowledge-freshness-audit": {
        "id": "knowledge-freshness-audit",
        "name": "知识保鲜度审计",
        "version": "1.0.0",
        "description": "审计知识库中条目的时效性，检测可能过时的规则、被案例突破的知识点，生成保鲜优先级建议。",
        "detailed_description": (
            "知识不是永恒不变的。监管政策会更新、市场环境会变化、案例复盘会揭示旧规则的失效。"
            "本Skill对知识库做全面的「保鲜体检」：\n\n"
            "**四维审计**：\n"
            "1. 置信度分布 — 高/中/低置信度条目的比例（低置信度过高=知识库缺乏验证）\n"
            "2. 溯源完整性 — 有多少条目标注了来源文档和贡献专家（无来源=无法追溯=高风险）\n"
            "3. 隐性注释交叉引用 — Step3中收集的专家隐性注释是否暗示某些知识已被实际突破\n"
            "4. LLM深度审计 — 识别依赖特定时间窗口的规则（如'近三年'）、依赖特定政策的规则、可能已过时的行业惯例\n\n"
            "**特殊能力**：如果在Step3中专家填写了隐性注释（如「这条规则在芯片行业经常被突破」），"
            "本Skill会自动检测这些注释对应的知识条目，标记为「已被质疑」并建议重新评估。"
        ),
        "usage_guide": (
            "1. 进入Step5（智能转化）后的保鲜审计，或直接访问Skill面板\n"
            "2. 选择关联的流水线ID（自动读取知识Excel + 隐性注释），或直接上传知识Excel文件\n"
            "3. 选择模型（用于深度审计），点击执行\n"
            "4. 查看保鲜风险指标和深度审计建议\n"
            "5. 根据报告的优先级列表，逐条更新过时知识"
        ),
        "input_example": "关联当前流水线（自动读取 final_*.xlsx + Step3中积累的隐性注释），或上传一份知识Excel文件",
        "output_example": (
            "生成 freshness_audit_*.md 报告，包含：\n"
            "一、保鲜风险指标（置信度分布、来源覆盖率、专家署名率、隐性注释数量）\n"
            "二、深度审计与保鲜建议（LLM识别出的可能过时规则及更新优先级）"
        ),
        "business_value": (
            "隐性知识衰减速度极快——老张退休后他的经验就没了，但更可怕的是老张的经验已经过时了但没人知道。"
            "这个Skill确保知识库的「保鲜期」可追踪、可审计。"
        ),
        "applicable_scenarios": [
            "季度知识库健康检查",
            "监管政策重大变化后的知识库集中更新",
            "老员工离职/调岗前的知识交接审计",
            "新案例大量涌现后，对旧知识的重新验证",
        ],
        "limitations": [
            "保鲜度判断依赖元数据质量——如果知识条目从不标注「来源」和「贡献专家」，审计能做的事有限",
            "LLM无法实时感知最新的监管政策变化——它只能基于训练数据中的知识来识别可能过时的模式",
            "建议每个季度执行一次，而非高频使用——审计的价值在于周期性对比",
        ],
        "triggers": ["保鲜审计", "时效审计", "知识老化", "保鲜度", "知识更新"],
        "supported_formats": ["XLSX"],
        "max_file_size_mb": 50,
        "output_styles": ["标准审计", "深度审计"],
        "capabilities": [
            "知识最后验证时间追踪",
            "过期/被突破规则自动标记",
            "置信度衰减模式检测",
            "案例复盘与知识条目的矛盾检测",
            "保鲜优先级排序与更新建议",
        ],
        "related_step": 5,
        "enabled": True,
    },
}
@app.route("/api/skills", methods=["GET"])
def api_skills_list():
    """返回所有已注册 Skill 的简要信息"""
    skills = []
    for sid, info in SKILL_REGISTRY.items():
        skills.append({
            "id": sid,
            "name": info["name"],
            "version": info["version"],
            "description": info["description"],
            "enabled": info["enabled"],
        })
    return jsonify({"status": "ok", "skills": skills})


@app.route("/api/skills/<path:skill_id>", methods=["GET"])
def api_skill_detail(skill_id):
    """返回指定 Skill 的详细信息"""
    info = SKILL_REGISTRY.get(skill_id)
    if not info:
        return jsonify({"status": "error", "error": "Skill 不存在"}), 404
    return jsonify({"status": "ok", "skill": info})


@app.route("/api/skills/<path:skill_id>", methods=["PUT"])
def api_skill_update(skill_id):
    """更新 Skill 配置（如启用/禁用）"""
    info = SKILL_REGISTRY.get(skill_id)
    if not info:
        return jsonify({"status": "error", "error": "Skill 不存在"}), 404
    data = request.get_json(force=True)
    if "enabled" in data:
        info["enabled"] = bool(data["enabled"])
    return jsonify({"status": "ok", "skill": info})


@app.route("/api/skills/execute", methods=["POST"])
def api_skill_execute():
    """通用 Skill 执行入口，根据 skill_id 路由到对应处理器"""
    skill_id = request.form.get("skill_id", "").strip()
    if not skill_id:
        return jsonify({"status": "error", "error": "缺少 skill_id 参数"})

    info = SKILL_REGISTRY.get(skill_id)
    if not info or not info.get("enabled"):
        return jsonify({"status": "error", "error": f"Skill '{skill_id}' 不存在或未启用"})

    # Route based on skill_id
    if skill_id == "knowledge-extraction":
        return _execute_knowledge_extraction()
    elif skill_id == "knowledge-revision":
        return _execute_knowledge_revision()
    elif skill_id == "knowledge-pattern-mining":
        return _execute_pattern_mining()
    elif skill_id == "knowledge-gap-analysis":
        return _execute_gap_analysis()
    elif skill_id == "knowledge-freshness-audit":
        return _execute_freshness_audit()
    else:
        return jsonify({"status": "error", "error": f"Skill '{skill_id}' 暂无执行处理器"})


def _resolve_step1_workbook_path(pipeline_id: str):
    """从流水线 step_data 解析 Step1 场景骨架 Excel 路径。"""
    if not pipeline_id:
        return None
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                step1_file = p.get("step_data", {}).get("step1_output_file", "")
                if step1_file:
                    candidate = WORKSPACE / step1_file
                    if candidate.exists():
                        return candidate
                break
    return None


def _write_step2_preextract_excel(pipeline_id: str, extracted_items: list):
    from step2_preextract import write_preextract_excel

    step1_path = _resolve_step1_workbook_path(pipeline_id)
    output_name = f"preextract_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = WORKSPACE / output_name
    meta = write_preextract_excel(
        step1_path=step1_path,
        output_path=output_path,
        items=extracted_items or [],
    )
    return output_name, meta


def _persist_step2_excel_pipeline(
    pipeline_id,
    output_name,
    extracted_text,
    style,
    count,
    *,
    md_name: str = "",
    md_url: str = "",
):
    if not pipeline_id:
        return
    if not is_step2_preextract_filename(output_name):
        raise ValueError(f"Step2 输出文件名非法（不得使用场景骨架 template_ 文件）: {output_name}")
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.setdefault("step_data", {})
                sd["skill_extract_result"] = extracted_text
                sd["skill_extract_style"] = style
                sd["step2_output_file"] = output_name
                sd["step2_download_url"] = f"/downloads/{output_name}"
                if md_name:
                    sd["step2_md_file"] = md_name
                    sd["step2_md_download_url"] = md_url or f"/downloads/{md_name}"
                else:
                    sd.pop("step2_md_file", None)
                    sd.pop("step2_md_download_url", None)
                sd["step2_extracted_count"] = count
                for stale in ("step2_preview_name", "step2_preview_url"):
                    sd.pop(stale, None)
                save_pipelines(pipelines)
                break


def _step1_knowledge_columns_from_pipeline(pipeline_id: str) -> list[str]:
    if not pipeline_id:
        return []
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                cols = p.get("step_data", {}).get("step1_knowledge_columns") or []
                if isinstance(cols, list):
                    return [str(c).strip() for c in cols if str(c).strip()]
                return []
    return []


def _extract_step2_target_columns(pipeline_id: str) -> list[str]:
    """Read Step1 template headers and produce Step2 target keys."""
    step1_path = _resolve_step1_workbook_path(pipeline_id)
    if not step1_path or not Path(step1_path).exists():
        return _step1_knowledge_columns_from_pipeline(pipeline_id)

    try:
        from step1_template import detect_header_rows, find_anchor_columns

        with _safe_workbook(step1_path) as wb:
            ws = wb[wb.sheetnames[0]]
            anchor_cols = set(find_anchor_columns(ws).values())
            header_rows = detect_header_rows(ws)

            cols = []
            seen = set()
            for c in range(1, (ws.max_column or 1) + 1):
                if c in anchor_cols:
                    continue
                h1 = str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else ""
                h2 = str(ws.cell(2, c).value).strip() if header_rows >= 2 and ws.cell(2, c).value else ""
                if h1 and h2 and h1 != h2:
                    key = f"{h1}-{h2}"
                else:
                    key = h2 or h1
                if not key:
                    continue
                if key not in seen:
                    seen.add(key)
                    cols.append(key)

        if len(cols) < 3:
            fallback = _step1_knowledge_columns_from_pipeline(pipeline_id)
            if len(fallback) > len(cols):
                return fallback
        return cols
    except Exception:
        return _step1_knowledge_columns_from_pipeline(pipeline_id)


def _normalize_extract_style(style: str) -> str:
    style = (style or "").strip()
    return style if style in EXTRACT_STYLE_RULES else "标准萃取"


def _pick_text(item: dict, keys: tuple[str, ...]) -> str:
    for k in keys:
        v = item.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def _extract_item_content(item: dict, target_columns: list | None = None) -> str:
    # 自定义模板：优先从 Step1 解析出的后段列名取值（前四列锚定列不在此列表中）
    if target_columns:
        for col in target_columns:
            v = item.get(col)
            if v is not None and str(v).strip():
                return str(v).strip()
        for col in target_columns:
            suffix = col.split("-")[-1].strip() if "-" in col else ""
            if not suffix:
                continue
            for k, v in item.items():
                if v is not None and str(v).strip() and (str(k).endswith(suffix) or suffix in str(k)):
                    return str(v).strip()

    text = _pick_text(
        item,
        (
            "content", "知识描述", "知识内容", "具体方法",
            "category", "知识分类", "步骤", "名称", "描述",
            "trigger_condition", "适用条件", "触发条件",
            "excerpt", "原文摘录", "知识引用",
        ),
    )
    if text:
        return text

    # Fallback for template-specific keys (e.g. "关键输出-名称"/"xxx-描述").
    for k, v in item.items():
        if v is None:
            continue
        key = str(k or "")
        if not key:
            continue
        s = str(v).strip()
        if not s:
            continue
        if any(mark in key for mark in ("名称", "描述", "内容", "步骤", "方法", "输出", "逻辑", "条件", "引用")):
            return s

    # Last-resort: first non-empty scalar string value.
    for v in item.values():
        if isinstance(v, (dict, list)):
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def _extract_item_confidence_rank(item: dict) -> int:
    conf = _pick_text(item, ("confidence", "置信度")).lower()
    if "高" in conf or "high" in conf:
        return 3
    if "中" in conf or "medium" in conf or "med" in conf:
        return 2
    if "低" in conf or "low" in conf:
        return 1
    return 2


def _extract_item_richness(item: dict) -> int:
    keys = (
        "category", "知识分类", "步骤",
        "content", "知识描述", "知识内容", "具体方法",
        "trigger_condition", "适用条件", "触发条件", "访谈方向",
        "judgment_logic", "判断逻辑", "规则引用",
        "anti_pattern", "反模式", "反模式/踩坑提示", "描述",
        "source", "来源", "来源文档",
        "confidence", "置信度",
        "excerpt", "原文摘录", "知识引用",
    )
    seen_values = set()
    richness = 0
    for k in keys:
        v = item.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if s in seen_values:
            continue
        seen_values.add(s)
        richness += 1
    return richness


def _apply_extract_style_rules(items: list, style: str, target_columns: list | None = None) -> tuple[list, dict]:
    """Deterministic post-processing so style differences are stable."""
    style = _normalize_extract_style(style)
    rule = EXTRACT_STYLE_RULES[style]
    template_mode = bool(target_columns)

    candidates = []
    seen = set()
    for raw in items or []:
        if not isinstance(raw, dict):
            if isinstance(raw, (str, int, float, bool)):
                raw = {"content": str(raw)}
            else:
                continue
        text = _extract_item_content(raw, target_columns)
        if not text:
            continue
        key = "".join(text.lower().split())
        if key in seen:
            continue
        seen.add(key)

        conf_rank = _extract_item_confidence_rank(raw)
        richness = _extract_item_richness(raw)

        # Hard filter for concise mode: keep high-value entries only.
        # 自定义模板（仅后段列名变化）时放宽，避免列名不含「方法/描述」等导致被滤光。
        if style == "精简萃取" and not template_mode:
            if conf_rank < 2:
                continue
            if richness < 3 and len(text) < 24:
                continue
        elif style == "精简萃取" and template_mode and len(text) < 8:
            continue

        # Different deterministic ranking per style.
        if style == "深度萃取":
            score = (richness, conf_rank, len(text))
        elif style == "精简萃取":
            score = (conf_rank, richness, -len(text))
        else:
            score = (conf_rank, richness, len(text))
        candidates.append((score, raw))

    candidates.sort(key=lambda x: x[0], reverse=True)
    processed = [x[1] for x in candidates[: rule["max_items"]]]

    # If concise-mode hard filters accidentally drop everything, keep top candidates.
    fallback_applied = False
    if not processed and candidates:
        processed = [x[1] for x in candidates[: rule["max_items"]]]
        fallback_applied = True

    stats = {
        "raw_count": len(items or []),
        "candidate_count": len(candidates),
        "processed_count": len(processed),
        "fallback_applied": fallback_applied,
        "min_items": rule["min_items"],
        "max_items": rule["max_items"],
    }
    return processed, stats


def _execute_knowledge_extraction():
    """知识萃取 Skill 执行"""
    skill_id = request.form.get("skill_id", "knowledge-extraction")
    info = SKILL_REGISTRY.get(skill_id, {})
    model_name = request.form.get("model", "")
    style = _normalize_extract_style(request.form.get("style", "标准萃取"))
    style_rule = EXTRACT_STYLE_RULES[style]
    pipeline_id = request.form.get("pipeline_id", "")
    content = request.form.get("content", "")
    source_file = request.files.get("file")
    cached_file = os.path.basename(request.form.get("cached_file", "").strip())

    if not content and not source_file and not cached_file:
        return jsonify({"status": "error", "error": "请提供文档内容或上传文件"})
    _debug_log(
        "H2",
        "app_server.py:_execute_knowledge_extraction",
        "step2 extraction start",
        {"has_content": bool(content), "has_upload": bool(source_file), "has_cached_file": bool(cached_file)},
    )

    if not model_name:
        models_list = load_llm_config()
        if models_list:
            model_name = models_list[0]["name"]
    model_cfg = get_model_by_name(model_name)
    if not model_cfg:
        return jsonify({"status": "error", "error": f"模型 '{model_name}' 不存在或无可用模型"})

    # Read file content if uploaded
    doc_text = content
    if source_file and not doc_text:
        try:
            doc_text = extract_text_from_file(source_file)
        except Exception:
            return jsonify({"status": "error", "error": "文件读取失败"})
    elif cached_file and not doc_text:
        cached_path = resolve_cache_file_path(WORKSPACE, cached_file)
        if cached_path:
            try:
                doc_text = extract_text_from_path(str(cached_path))
            except Exception:
                pass

    target_columns = _extract_step2_target_columns(pipeline_id)
    skill_caps = info.get("capabilities", []) if isinstance(info, dict) else []
    cap_text = "；".join(skill_caps) if skill_caps else "结构化知识提取"
    max_tokens = style_rule["max_tokens"]
    if target_columns:
        max_tokens = max(max_tokens, 6144 if len(target_columns) > 10 else 5120)

    # ── 案例复盘模式：提取隐性信号 + 可执行知识 ──
    content_type = (request.form.get("content_type", "") or "").strip()
    if content_type == "case_review":
        system_prompt = (
            f"你是一位资深银行知识工程专家，正在从「案例复盘」中同时提取两类内容：\n"
            f"萃取风格：{style}\n\n"
            f"## 任务一：识别隐性信号（重点）\n"
            f"案例复盘中的隐性知识往往不是直接说出来的。请你特别注意以下四类信号：\n"
            f"1. **规则覆盖不到的地方**：专家提到了哪些标准流程中没有的检查步骤？哪些「多余的动作」？\n"
            f"2. **情感/直觉表达**：专家用了哪些不安/不对劲/怪怪的情感词汇？这些情感背后对应了什么可观测信号？\n"
            f"3. **破例逻辑**：专家在哪次决策中突破了标准规则？他用来合理化的理由是什么？是否值得固化为例外条件？\n"
            f"4. **关系依赖**：专家提到「问了某某人」吗？那个人知道什么别人不知道的东西？\n\n"
            f"## 任务二：抽取可执行知识条目\n"
            f"同时从案例中提取以下格式的结构化知识条目。\n\n"
            f"请按以下JSON格式输出（一个数组，不要Markdown代码块，不要任何前后说明文字）：\n"
            f'[{{\"隐性信号\": \"描述一个规则覆盖不到的场景或直觉信号（一句话）\", '
            f'\"信号类型\": \"反模式|破例|直觉|关系依赖\", '
            f'\"可执行知识\": \"从这个信号中可以提炼出什么可操作的知识？\", '
            f'\"触发条件\": \"什么情况下应该特别关注这个信号？\", '
            f'\"来源\": \"来自本案例复盘的哪个部分（标题/背景/判断/结果/重来/习惯）\", '
            f'\"置信度\": \"高|中|低\"}}]\n\n'
            f"要求：\n"
            f"1. 每条隐性信号必须是完整、自包含的陈述\n"
            f"2. 优先提取反模式和破例逻辑——这些是隐性知识的关键入口\n"
            f"3. 输出条数尽量 {style_rule['min_items']}~{style_rule['max_items']} 条\n"
            f"4. 可执行知识要具体——不能只写「注意风险」，要写「注意什么风险、怎么看、看哪里」"
        )
        # Override target columns for case review output
        target_columns = ["隐性信号", "信号类型", "可执行知识", "触发条件", "来源", "置信度"]
    elif target_columns:
        target_cols_json = json.dumps(target_columns, ensure_ascii=False)
        example_obj = {k: "" for k in target_columns}
        content_key = next(
            (k for k in target_columns if any(m in k for m in ("方法", "描述", "内容", "引用"))),
            target_columns[0],
        )
        example_obj[content_key] = "（示例：从文档抽取的一条可执行知识）"
        example_json = json.dumps([example_obj], ensure_ascii=False)
        system_prompt = (
            f"你是一位知识工程专家，正在执行隐性知识显性化的第二步——知识萃取。\n"
            f"萃取风格：{style}\n"
            f"Skill能力参考：{cap_text}\n\n"
            f"风格硬规则：{style_rule['prompt_hint']}\n"
            f"请按用户上传的萃取模板抽取知识。前四列（场景/场景说明/子场景/子场景说明）已由系统填写，"
            f"JSON 只需包含下列第5列及之后的字段（键名与表头完全一致）：\n"
            f"【输出格式 — 必须严格遵守】\n"
            f"1. 只输出一个 JSON 数组，不要用 Markdown 代码块，不要写任何前后说明文字。\n"
            f"2. 数组元素为对象；每个对象的键名必须与下列列表完全一致（含连字符）：{target_cols_json}\n"
            f"3. 键名与值均使用英文双引号；无信息的字段填空字符串 \"\"。\n"
            f"4. 输出条数尽量 {style_rule['min_items']}~{style_rule['max_items']} 条。\n"
            f"5. 输出示例（结构参考，请替换为真实抽取内容）：\n{example_json}"
        )
        if _pipeline_prefers_markdown(pipeline_id) or len(target_columns) >= 8:
            system_prompt += (
                "\n\n【深度萃取 — 多语义列】\n"
                "适用条件、判断逻辑、反模式/踩坑提示、知识描述、知识引用等长文本字段须写完整"
                "（每条通常不少于一两句），勿只填占位词；尽量让每条记录在多数语义列上都有实质内容。"
            )
    else:
        system_prompt = (
            f"你是一位知识工程专家，正在执行隐性知识显性化的第二步——知识萃取。\n"
            f"萃取风格：{style}\n"
            f"Skill能力参考：{cap_text}\n\n"
            f"风格硬规则：{style_rule['prompt_hint']}\n"
            f"请从以下文档中提取所有已显性化的知识条目，按以下JSON数组格式输出：\n"
            f'[{{"category": "判断规则|操作流程|反模式|审批标准|经验法则", '
            f'"content": "知识内容（一句话完整陈述）", '
            f'"trigger_condition": "触发条件", '
            f'"judgment_logic": "判断逻辑", '
            f'"anti_pattern": "常见反模式/踩坑提醒", '
            f'"source": "来源文档名", '
            f'"confidence": "高|中|低"}}]\n\n'
            f"要求：\n"
            f"1. 每条知识必须是完整的、自包含的陈述\n"
            f"2. category 只能是：判断规则、操作流程、反模式、审批标准、经验法则\n"
            f"3. 输出条数尽量满足 {style_rule['min_items']}~{style_rule['max_items']} 条\n"
            f"4. 尽量提取判断逻辑和反模式，这是隐性知识的关键入口"
        )

    extracted = ""
    parse_mode = "none"
    extracted_items = []
    extract_stats = {"raw_count": 0, "processed_count": 0, "min_items": style_rule["min_items"], "max_items": style_rule["max_items"]}
    output_name = None
    excel_meta = {}
    try:
        result = call_llm_with_retry(model_cfg, [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"请从以下文档中提取知识条目：\n\n{doc_text}"}
        ], stream=False, temperature=style_rule["temperature"], max_tokens=max_tokens)

        if isinstance(result, dict):
            extracted = extract_assistant_content(result)
        extracted_items, parse_mode = _parse_extracted_items(extracted)
        extracted_items = _normalize_extracted_items(extracted_items, target_columns)

        extracted_items, extract_stats = _apply_extract_style_rules(extracted_items, style, target_columns)

        # Never generate a misleading empty pre-extract workbook when model did return text.
        if not extracted_items and (extracted or "").strip():
            if parse_mode in {"json_list", "json_items", "object_recovery", "jsonl"}:
                err_text = (
                    "知识提取结果在风格规则过滤后为空，已阻止生成空萃取稿。"
                    "请切换到「标准萃取」重试，或检查自定义模板表头是否含可填写的知识列（非仅场景四列）。"
                )
            else:
                err_text = (
                    "知识提取结果解析失败（模型未返回可解析的 JSON 数组）。"
                    "自定义模板列较多时更易出现；请重试并优先使用「标准萃取」，或简化模板表头。"
                )
            return jsonify({
                "status": "error",
                "error": err_text,
                "style": style,
                "parse_mode": parse_mode,
                "target_column_count": len(target_columns),
                "used_template_columns": bool(target_columns),
                "extracted_preview": (extracted or "")[:2000],
                "style_rule": {
                    "raw_count": extract_stats.get("raw_count", 0),
                    "candidate_count": extract_stats.get("candidate_count", 0),
                    "processed_count": extract_stats.get("processed_count", 0),
                    "fallback_applied": extract_stats.get("fallback_applied", False),
                },
                "build": STEP2_EXCEL_BUILD,
            })

        output_name, excel_meta = _write_step2_preextract_excel(pipeline_id, extracted_items)
        step2_md_name, step2_md_url = _maybe_generate_markdown_artifact(
            pipeline_id,
            output_name,
            md_prefix="preextract",
            title=f"Step2 知识萃取 · {pipeline_id[:8]}",
        )
        _persist_step2_excel_pipeline(
            pipeline_id,
            output_name,
            extracted,
            style,
            len(extracted_items),
            md_name=step2_md_name,
            md_url=step2_md_url,
        )
        step1_source = ""
        if pipeline_id:
            with _pipelines_lock:
                for p in load_pipelines():
                    if p["id"] == pipeline_id:
                        step1_source = p.get("step_data", {}).get("step1_output_file", "")
                        break

        return jsonify({
            "status": "ok",
            "skill_name": info.get("name", skill_id),
            "skill_id": skill_id,
            "model": model_name,
            "style": style,
            "extracted": extracted,
            "extracted_count": len(extracted_items),
            "style_rule": {
                "mode": style,
                "min_items": extract_stats.get("min_items", style_rule["min_items"]),
                "max_items": extract_stats.get("max_items", style_rule["max_items"]),
                "raw_count": extract_stats.get("raw_count", 0),
                "processed_count": extract_stats.get("processed_count", len(extracted_items)),
            },
            "parse_mode": parse_mode,
            "download_name": output_name,
            "download_url": f"/downloads/{output_name}",
            "markdown_file": step2_md_name,
            "markdown_download_url": step2_md_url,
            "output_kind": "preextract",
            "step1_source_file": step1_source,
            "filled_rows": excel_meta.get("filled_rows", 0),
            "used_step1_template": excel_meta.get("used_step1_template", False),
            "build": STEP2_EXCEL_BUILD,
        })
    except LlmApiError as e:
        payload = {"status": "error", "error": str(e), "build": STEP2_EXCEL_BUILD}
        if output_name:
            payload.update({
                "download_name": output_name,
                "download_url": f"/downloads/{output_name}",
            })
        return jsonify(payload)
    except Exception as e:
        _debug_log(
            "H2",
            "app_server.py:_execute_knowledge_extraction",
            "step2 generic error",
            {"error": str(e)[:300]},
        )
        payload = {"status": "error", "error": f"萃取失败: {str(e)}", "build": STEP2_EXCEL_BUILD}
        return jsonify(payload)



def _resolve_step2_excel_path(pipeline_id: str):
    """Step3 修订底稿：仅使用 Step2 萃取 Excel，禁止回退到 Step1 场景骨架。"""
    if not pipeline_id:
        return None, ""
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] != pipeline_id:
                continue
            sd = p.get("step_data", {})
            step2_file = sd.get("step2_output_file", "")
            if not is_step2_preextract_filename(step2_file):
                return None, ""
            path = safe_workspace_path(WORKSPACE, step2_file, must_exist=True)
            if path:
                return str(path), sd.get("skill_extract_result", "")[:3000]
            break
    return None, ""


def _normalize_revision_style(style: str) -> str:
    style = (style or "").strip()
    return style if style in REVISION_STYLE_RULES else "标准修订"


_ALIGN_CHAT_ROUND_RE = re.compile(r"^第\d+轮专家意见[：:]\s*", re.MULTILINE)
_ALIGN_REVISION_CUE_RE = re.compile(
    r"(第\s*\d+\s*行|第\s*\d+\s*列|改为|修改|删除|新增|补充(?!意见)|调整|更正|修订|替换|"
    r"单元格|sheet|列[abcde]|行\s*\d+)",
    re.IGNORECASE,
)
_EXPLICIT_NO_OPINION_PHRASES = (
    "没有意见", "无意见", "无异议", "无需修订", "无需修改", "无修改", "没有修改",
    "无修订", "没有修订", "无变更", "保持不变", "专家无意见", "暂无意见",
    "无专家意见", "确认通过", "确认无误", "可以发布", "同意通过", "不需修改",
    "不需要修改", "没有异议", "无会议纪要修订", "无修订意见", "无修改意见",
    "无补充意见", "无补充",
)


def _normalize_expert_text_for_align(expert_text: str) -> str:
    """去掉对话轮次前缀与占位话术，便于判断是否有修订意图。"""
    t = (expert_text or "").strip()
    if not t:
        return ""
    t = _ALIGN_CHAT_ROUND_RE.sub("", t)
    t = re.sub(r"[（(]无补充意见[）)]", "", t)
    return t.strip()


def _text_is_explicit_no_opinion(expert_text: str) -> bool:
    t = _normalize_expert_text_for_align(expert_text)
    if not t:
        return True
    if _ALIGN_REVISION_CUE_RE.search(t):
        return False
    compact = re.sub(r"[\s,.，。、；;：:!！?？\-—_（）()]+", "", t.lower())
    if any(p in compact for p in _EXPLICIT_NO_OPINION_PHRASES):
        return True
    if compact in ("无", "没有", "同意", "通过", "ok", "none", "na", "n/a", "暂无"):
        return True
    return False


def _uploaded_expert_material_is_substantive(uploaded_text: str) -> bool:
    """会议纪要/访谈记录等上传内容达到可触发智能修订的阈值。"""
    t = (uploaded_text or "").strip()
    if len(t) < 30:
        return False
    return not _text_is_explicit_no_opinion(t)


def _should_pass_through_preextract(expert_text: str, uploaded_material_text: str = "") -> bool:
    """
    知识对齐直通预萃稿（不调用 LLM）：
    - 无专家文本且无实质上传材料；
    - 或专家明确表示无意见（如「暂无意见」）。
    """
    if _uploaded_expert_material_is_substantive(uploaded_material_text):
        return False
    return _text_is_explicit_no_opinion(expert_text)


def _alignment_llm_guard_rules() -> str:
    return """
## 重要约束
1. 仅根据「专家意见」中**明确写出**的修订要求生成条目；禁止仅依据知识稿内容自行推断、优化或补充修订。
2. 若专家明确表示无意见、无需修改、确认通过等，必须输出空数组 []。
3. 不得将知识稿中的待完善项自动转为修订建议，除非专家意见中点名要求修改。
"""


def _resolve_align_source_for_pipeline(pipeline_id: str) -> tuple[str | None, str]:
    """返回 (source_file_path, basename)。"""
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {})
                resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="align")
                if resolved:
                    return str(resolved), resolved.name
                break
    return None, ""


def _publish_final_from_source(
    pipeline_id: str,
    source_file_path: str,
    expert_text: str = "",
    style: str = "标准修订",
):
    """将当前对齐输入稿复制为 final_*.xlsx（无修订）。"""
    from datetime import datetime
    import shutil

    output_name = f"final_{pipeline_id[:8]}_{datetime.now().strftime('%H%M%S')}.xlsx"
    output_path = WORKSPACE / output_name
    shutil.copy2(source_file_path, str(output_path))
    md_name, md_url = _maybe_generate_markdown_artifact(
        pipeline_id,
        output_name,
        md_prefix="final",
        title=f"Step3 知识对齐 · {pipeline_id[:8]}",
    )
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.setdefault("step_data", {})
                sd["step4_final_file"] = output_name
                sd["step4_download_url"] = "/downloads/" + output_name
                if md_name:
                    sd["step4_md_file"] = md_name
                    sd["step4_md_download_url"] = md_url
                else:
                    sd.pop("step4_md_file", None)
                    sd.pop("step4_md_download_url", None)
                sd["step4_final_notes"] = (expert_text or "")[:500]
                sd["step4_final_style"] = style
                sd["step4_final_count"] = 0
                sd.pop("_align_preview_notes", None)
                sd.pop("_align_preview_style", None)
                sd.pop("_align_source_file", None)
                sd.pop("_align_chat_history", None)
                save_pipelines(pipelines)
                break
    return output_name, md_name, md_url


def _align_no_opinion_success_payload(
    pipeline_id: str,
    source_file_path: str,
    expert_text: str,
    style: str,
    style_rule: dict,
    *,
    auto_publish: bool = True,
) -> dict:
    """无修订意见时返回成功；默认自动发布 final 稿，避免用户额外点击确认。"""
    payload = {
        "status": "ok",
        "notes": [],
        "total": 0,
        "no_opinion": True,
        "auto_finalized": False,
        "align_mode": "pass_through",
        "message": "专家意见为无需修订，已自动将当前稿确认为对齐稿。",
        "style": style,
        "style_rule": {
            "mode": style,
            "max_actions": style_rule.get("max_actions", 0),
            "raw_count": 0,
            "processed_count": 0,
        },
    }
    if auto_publish and source_file_path:
        output_name, md_name, md_url = _publish_final_from_source(
            pipeline_id,
            source_file_path,
            expert_text=expert_text,
            style=style,
        )
        payload["auto_finalized"] = True
        payload["revision_count"] = 0
        payload["accepted_count"] = 0
        payload["output_file"] = output_name
        payload["download_name"] = output_name
        payload["download_url"] = "/downloads/" + output_name
        payload["markdown_file"] = md_name
        payload["markdown_download_url"] = md_url or ""
        payload["message"] = "已自动将当前萃取稿确认为对齐稿（无修订），可直接下载。"
    else:
        payload["message"] = "专家意见为无需修订。可直接确认当前稿为对齐稿。"
    return payload


def _normalize_revision_action(action: str) -> str:
    a = (action or "").strip().lower()
    alias = {
        "修改": "modify",
        "delete": "delete",
        "删除": "delete",
        "新增": "add",
        "add": "add",
        "补充": "supplement",
        "supplement": "supplement",
    }
    return alias.get(a, a)


def _to_int(val, default=0) -> int:
    if isinstance(val, int):
        return val
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default

    # Excel column letters: A->1, B->2, AA->27
    if s.isalpha():
        n = 0
        for ch in s.upper():
            if "A" <= ch <= "Z":
                n = n * 26 + (ord(ch) - ord("A") + 1)
            else:
                return default
        return n or default

    # Strings like "第6列"/"col=8" -> extract first integer.
    digits = []
    sign = 1
    found = False
    for i, ch in enumerate(s):
        if ch == "-" and not found:
            sign = -1
            continue
        if ch.isdigit():
            digits.append(ch)
            found = True
        elif found:
            break
    if digits:
        try:
            return sign * int("".join(digits))
        except Exception:
            return default

    return default


def _apply_revision_style_rules(expert_notes: list, style: str) -> tuple[list, dict]:
    """Deterministic filtering to make revision styles behaviorally distinct."""
    style = _normalize_revision_style(style)
    rule = REVISION_STYLE_RULES[style]
    allowed_actions = rule["allowed_actions"]

    dedup = {}
    raw_count = 0
    for raw in expert_notes or []:
        if not isinstance(raw, dict):
            continue
        raw_count += 1

        action = _normalize_revision_action(raw.get("action", ""))
        if action not in allowed_actions:
            continue

        sheet = str(raw.get("sheet") or "").strip() or "Sheet1"
        row = _to_int(raw.get("row"), 0)
        col = _to_int(raw.get("col"), 0)
        old_value = str(raw.get("old_value") or "").strip()
        new_value = str(raw.get("new_value") or "").strip()
        note = str(raw.get("note") or "").strip()

        # Hard validity checks
        if action in {"modify", "supplement", "add"} and not new_value:
            continue
        if action == "modify" and not old_value:
            continue
        if action != "add" and row <= 0:
            continue
        # New rows must be explicitly positioned; otherwise they tend to pile up.
        if action == "add" and (row <= 0 or col <= 0):
            continue
        if style == "严格修订" and (action in {"add", "delete"}):
            continue

        normalized = {
            "sheet": sheet,
            "row": row,
            "col": col if col > 0 else 1,
            "action": action,
            "old_value": old_value,
            "new_value": new_value,
            "note": note,
        }

        if action == "add":
            dedup_key = (sheet, action, normalized["new_value"][:120])
        else:
            dedup_key = (sheet, row, normalized["col"], action)

        current = dedup.get(dedup_key)
        if current is None or len(normalized["note"]) >= len(current.get("note", "")):
            dedup[dedup_key] = normalized

    filtered = list(dedup.values())

    # Deterministic ranking by style
    if style == "严格修订":
        priority = {"modify": 3, "supplement": 2, "add": 0, "delete": 0}
    elif style == "宽松修订":
        priority = {"add": 4, "supplement": 3, "modify": 2, "delete": 1}
    else:
        priority = {"modify": 4, "supplement": 3, "add": 2, "delete": 1}

    filtered.sort(
        key=lambda n: (
            priority.get(n.get("action", ""), 0),
            len((n.get("note") or "")),
            len((n.get("new_value") or "")),
        ),
        reverse=True,
    )
    filtered = filtered[: rule["max_actions"]]

    stats = {
        "raw_count": raw_count,
        "processed_count": len(filtered),
        "max_actions": rule["max_actions"],
    }
    return filtered, stats


def _run_knowledge_revision(pipeline_id: str, expert_text: str, style: str, model_name: str = ""):
    """从 Step2/Step1 Excel + 专家意见生成带标注的修订稿。"""
    style = _normalize_revision_style(style)
    style_rule = REVISION_STYLE_RULES[style]
    excel_file, step2_content = _resolve_step2_excel_path(pipeline_id)
    if not excel_file or not os.path.exists(excel_file):
        return {"status": "error", "error": "未找到知识萃取文件，请先完成知识萃取（Step2）"}
    _debug_log(
        "H3",
        "app_server.py:_run_knowledge_revision",
        "step3 revision start",
        {"has_pipeline_id": bool(pipeline_id), "style": style, "expert_text_len": len(expert_text or "")},
    )

    try:
        from workbook_layout import build_revision_context, layout_prompt_rules, normalize_revision_notes

        excel_context_str, layout_map = build_revision_context(excel_file)
    except Exception as e:
        return {"status": "error", "error": f"读取Excel文件失败: {str(e)}"}

    style_desc = style_rule["prompt_hint"]

    step2_section = ""
    if step2_content:
        step2_section = f"\n## Step2知识萃取稿内容（参考）\n{step2_content}\n"

    prompt = f"""你是一个知识修订助手。你的任务是将专家修订意见解析为结构化的修订JSON。

## 当前Excel知识文件（含 excel_row 物理行号）
{excel_context_str}
{step2_section}
## 专家修订意见
{expert_text}

## 修订风格：{style_desc}

{layout_prompt_rules()}

## 输出要求
请生成一个JSON数组，每个元素是一条修订操作，格式如下：
```json
[
  {{
    "sheet": "工作表名称",
    "row": excel_row,
    "col": 列号(1-based，A列=1),
    "action": "modify|delete|add|supplement",
    "old_value": "原始内容（modify/delete时必填）",
    "new_value": "修订后内容（modify/add/supplement时必填）",
    "note": "修订说明"
  }}
]
```

注意：
1. row 必须使用预览中的 excel_row，不要修改表头行
2. col 使用 1-based 数字列号（A=1,B=2）
3. add 操作必须给出明确的 row 与 col（禁止 row=0/col=0）
4. 仅输出JSON数组，不要输出其他内容"""

    models_list = load_llm_config()
    if not model_name and models_list:
        model_name = models_list[0]["name"]
    model_cfg = get_model_by_name(model_name)
    if not model_cfg:
        return {"status": "error", "error": f"模型 '{model_name}' 不存在或无可用模型"}

    llm_text = ""
    try:
        llm_result = call_llm_with_retry(
            model_cfg,
            messages=[
                {"role": "system", "content": "你是一个知识修订助手，负责将专家修订意见解析为结构化的修订JSON。仅输出JSON数组，不要输出其他内容。"},
                {"role": "user", "content": prompt},
            ],
            stream=False,
            temperature=style_rule["temperature"],
        )
        llm_text = extract_assistant_content(llm_result) if isinstance(llm_result, dict) else ""
        llm_text = _extract_json_from_text(llm_text.strip())
        expert_notes = json.loads(llm_text)
        if not isinstance(expert_notes, list):
            expert_notes = [expert_notes]
        expert_notes, note_stats = _apply_revision_style_rules(expert_notes, style)
        expert_notes, row_stats = normalize_revision_notes(expert_notes, layout_map)
        note_stats["row_adjusted"] = row_stats.get("adjusted", 0)
        note_stats["row_skipped_header"] = row_stats.get("skipped_header", 0)
    except LlmApiError as e:
        return {"status": "error", "error": str(e)}
    except json.JSONDecodeError as e:
        return {"status": "error", "error": f"LLM输出解析失败: {str(e)}", "raw_output": (llm_text or "")[:500]}
    except Exception as e:
        return {"status": "error", "error": f"LLM调用失败: {str(e)}"}

    try:
        from revision_processor import process_workbook
        from datetime import datetime

        output_name = f"revision_{pipeline_id[:8]}_{datetime.now().strftime('%H%M%S')}.xlsx"
        output_path = os.path.join(WORKSPACE, output_name)
        revision_count = process_workbook(excel_file, expert_notes, output_path, layouts=layout_map)

        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    sd["step3_revision_file"] = output_name
                    sd["step3_download_url"] = f"/downloads/{output_name}"
                    sd["step3_revision_notes"] = expert_text
                    sd["step3_revision_style"] = style
                    sd["step3_revision_count"] = revision_count
                    sd["step3_excel_path"] = output_name
                    save_pipelines(pipelines)
                    break

        return {
            "status": "ok",
            "revision_count": revision_count,
            "output_file": output_name,
            "download_name": output_name,
            "download_url": f"/downloads/{output_name}",
            "expert_notes_json": expert_notes,
            "style_rule": {
                "mode": style,
                "max_actions": note_stats["max_actions"],
                "raw_count": note_stats["raw_count"],
                "processed_count": note_stats["processed_count"],
            },
        }
        
    except Exception as e:
        _debug_log(
            "H3",
            "app_server.py:_run_knowledge_revision",
            "step3 revision processing failure",
            {"error": str(e)[:300]},
        )
        return {"status": "error", "error": f"修订处理失败: {str(e)}"}


def _execute_knowledge_revision():
    """知识修订 Skill 执行：基于 Step2 萃取稿 + 专家意见生成修订稿。

    专家无意见时，直接透传 Step2 萃取稿作为对齐输出（不调用 LLM）。
    """
    skill_id = request.form.get("skill_id", "knowledge-revision")
    info = SKILL_REGISTRY.get(skill_id, {})
    pipeline_id = request.form.get("pipeline_id", "")
    expert_text = request.form.get("expert_text", "")
    expert_cached_file = os.path.basename(request.form.get("expert_cached_file", "").strip())
    style = request.form.get("style", "标准修订")
    model_name = request.form.get("model", "")

    if not info or not info.get("enabled"):
        return jsonify({"status": "error", "error": "知识修订技能未启用"})
    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    # 加载上传材料（会议纪要/访谈记录等）
    uploaded_material_text = _load_align_expert_upload_text(
        uploaded_file=request.files.get("expert_file"),
        cached_file_name=expert_cached_file,
    )
    if not expert_text and expert_cached_file:
        try:
            cached_path = WORKSPACE / expert_cached_file
            if cached_path.exists():
                expert_text = extract_text_from_path(str(cached_path))
        except Exception:
            pass
    if not expert_text and uploaded_material_text:
        expert_text = uploaded_material_text

    # ── 无意见直通：不调用 LLM，直接透传 Step2 萃取稿 ──
    style = _normalize_revision_style(style)
    style_rule = REVISION_STYLE_RULES[style]
    if _should_pass_through_preextract(expert_text, uploaded_material_text):
        source_file_path, _ = _resolve_step2_excel_path(pipeline_id)
        if not source_file_path:
            return jsonify({"status": "error", "error": "未找到知识萃取文件，请先完成知识萃取（Step2）"})
        payload = _align_no_opinion_success_payload(
            pipeline_id, source_file_path, expert_text or "", style, style_rule
        )
        payload["skill_name"] = info.get("name", skill_id)
        payload["skill_id"] = skill_id
        return jsonify(payload)

    # ── 有实质意见：调用 LLM 生成修订稿 ──
    result = _run_knowledge_revision(pipeline_id, expert_text, style, model_name)
    if result.get("status") == "ok":
        result["skill_name"] = info.get("name", skill_id)
        result["skill_id"] = skill_id
    return jsonify(result)


# ─── 新 Skill: 跨案例模式发现 ─────────────────────────────────

def _execute_pattern_mining():
    """跨案例模式发现：对多个案例复盘进行交叉分析，发现反复出现的隐性信号。"""
    skill_id = request.form.get("skill_id", "knowledge-pattern-mining")
    info = SKILL_REGISTRY.get(skill_id, {})
    model_name = request.form.get("model", "")
    pipeline_id = request.form.get("pipeline_id", "")
    content = request.form.get("content", "")
    style = request.form.get("style", "标准模式发现")
    # 支持多文件上传
    uploaded_files = request.files.getlist("files")

    if not content and not uploaded_files:
        return jsonify({"status": "error", "error": "请提供至少两个案例复盘的文本，或上传案例文件"})

    if not model_name:
        models_list = load_llm_config()
        if models_list:
            model_name = models_list[0]["name"]
    model_cfg = get_model_by_name(model_name)
    if not model_cfg:
        return jsonify({"status": "error", "error": f"模型 '{model_name}' 不存在"})

    # 汇总所有案例文本
    all_cases = []
    if content:
        all_cases.append(content)
    for uf in uploaded_files:
        try:
            case_text = extract_text_from_file(uf)
            if case_text.strip():
                all_cases.append(f"=== 案例文件: {uf.filename} ===\n{case_text.strip()}")
        except Exception:
            pass

    if len(all_cases) < 2:
        return jsonify({"status": "error", "error": "跨案例分析需要至少 2 个案例，请补充更多案例复盘内容"})

    combined = "\n\n---分隔线---\n\n".join(all_cases)

    system_prompt = (
        f"你是一位资深银行风控专家，正在对多个案例复盘进行交叉分析，寻找**跨案例涌现的隐性知识模式**。\n\n"
        f"## 分析任务\n"
        f"请仔细阅读以下 {len(all_cases)} 个案例复盘，执行以下五步分析：\n\n"
        f"### 第一步：识别重复出现的预警信号\n"
        f"哪些具体的信号在多个案例中反复出现？请逐一列出，标注每个信号出现在哪几个案例中。\n"
        f"优先关注：财务指标之外的信号（水电费变化、人员变动、工商变更、关联交易、非正式信息源等）。\n\n"
        f"### 第二步：发现系统性风险盲区\n"
        f"这些案例共同揭示了一个什么样的**规则/流程层面**的盲区？\n"
        f"即：为什么多个案例中，按标准流程操作仍然没能提前发现风险？\n\n"
        f"### 第三步：提炼跨案例隐性知识\n"
        f"从这些案例中能提炼出哪些**可操作的新知识**？\n"
        f"这些知识不是来自单个案例，而是来自案例之间的共同模式。\n\n"
        f"### 第四步：信号优先级排序\n"
        f"按\"出现频率 × 损失严重度\"给所有信号排序，标注哪个信号是最早出现的（即最有预警价值的）。\n\n"
        f"### 第五步：生成行动建议\n"
        f"基于以上分析，给出三条具体的、可落地的行动建议。\n\n"
        f"## 输出格式\n"
        f"请严格按以下 JSON 输出（不要 Markdown 代码块，不要任何前后说明）：\n"
        f'{{\n'
        f'  "recurring_signals": [\n'
        f'    {{"signal": "信号描述", "cases": ["案例1标题", "案例2标题"], "frequency": 2, "earliest_indicator": true/false}}\n'
        f'  ],\n'
        f'  "systemic_blind_spots": ["盲区描述1", "盲区描述2"],\n'
        f'  "cross_case_knowledge": [\n'
        f'    {{"knowledge": "可执行知识", "source_signals": ["信号A", "信号B"], "actionable": "具体怎么做"}}\n'
        f'  ],\n'
        f'  "priority_ranking": [\n'
        f'    {{"rank": 1, "signal": "信号", "rationale": "为什么排第一"}}\n'
        f'  ],\n'
        f'  "action_recommendations": ["建议1", "建议2", "建议3"]\n'
        f'}}'
    )

    try:
        result = call_llm_with_retry(model_cfg, [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"以下是要交叉分析的 {len(all_cases)} 个案例复盘：\n\n{combined[:14000]}"}
        ], stream=False, temperature=0.3, max_tokens=4096)

        llm_text = extract_assistant_content(result) if isinstance(result, dict) else ""
        llm_text = _extract_json_from_text(llm_text)
        analysis = json.loads(_repair_json_text(llm_text))

        # 生成可下载的 Markdown 报告
        report_name = f"pattern_mining_{uuid.uuid4().hex[:8]}.md"
        report_path = WORKSPACE / report_name
        _write_pattern_mining_report(report_path, analysis, len(all_cases))

        # 持久化到 pipeline，供 Step3 修订上下文使用
        if pipeline_id:
            with _pipelines_lock:
                pipelines = load_pipelines()
                for p in pipelines:
                    if p["id"] == pipeline_id:
                        sd = p.setdefault("step_data", {})
                        sd["step2_pattern_mining_report"] = report_name
                        sd["step2_pattern_mining_url"] = f"/downloads/{report_name}"
                        sd["step2_pattern_mining_summary"] = {
                            "case_count": len(all_cases),
                            "top_signals": [
                                s.get("signal", "") for s in (analysis.get("recurring_signals") or [])[:5]
                            ],
                            "blind_spots": (analysis.get("systemic_blind_spots") or [])[:3],
                        }
                        save_pipelines(pipelines)
                        break

        return jsonify({
            "status": "ok",
            "skill_name": info.get("name", skill_id),
            "skill_id": skill_id,
            "model": model_name,
            "case_count": len(all_cases),
            "analysis": analysis,
            "report_name": report_name,
            "download_url": f"/downloads/{report_name}",
        })
    except json.JSONDecodeError as e:
        return jsonify({"status": "error", "error": f"LLM输出解析失败: {str(e)}", "raw": (llm_text or "")[:500]})
    except LlmApiError as e:
        return jsonify({"status": "error", "error": str(e)})
    except Exception as e:
        return jsonify({"status": "error", "error": f"模式发现失败: {str(e)}"})


def _write_pattern_mining_report(path: Path, analysis: dict, case_count: int) -> None:
    lines = [
        f"# 跨案例模式发现报告",
        f"",
        f"- 分析案例数：**{case_count}**",
        f"- 生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"",
        f"## 一、重复出现的预警信号",
        f"",
    ]
    for s in (analysis.get("recurring_signals") or []):
        cases_str = "、".join(s.get("cases", []))
        early = " ⚡最早信号" if s.get("earliest_indicator") else ""
        lines.append(f"- **{s.get('signal', '')}**（出现 {s.get('frequency', 0)} 次）{early}")
        lines.append(f"  - 涉及案例：{cases_str}")
        lines.append("")

    lines.append("## 二、系统性风险盲区")
    lines.append("")
    for b in (analysis.get("systemic_blind_spots") or []):
        lines.append(f"- {b}")
    lines.append("")

    lines.append("## 三、跨案例隐性知识")
    lines.append("")
    for k in (analysis.get("cross_case_knowledge") or []):
        lines.append(f"### {k.get('knowledge', '')}")
        lines.append(f"- 来源信号：{'、'.join(k.get('source_signals', []))}")
        lines.append(f"- 具体做法：{k.get('actionable', '')}")
        lines.append("")

    lines.append("## 四、信号优先级排序")
    lines.append("")
    for r in (analysis.get("priority_ranking") or []):
        lines.append(f"{r.get('rank', '?')}. **{r.get('signal', '')}** — {r.get('rationale', '')}")
    lines.append("")

    lines.append("## 五、行动建议")
    lines.append("")
    for i, a in enumerate((analysis.get("action_recommendations") or []), 1):
        lines.append(f"{i}. {a}")

    path.write_text("\n".join(lines), encoding="utf-8")


# ─── 新 Skill: 知识盲区检测 ─────────────────────────────────

def _execute_gap_analysis():
    """知识盲区检测：对比 Schema 定义的知识列与实际填充率，识别空白区域。"""
    skill_id = request.form.get("skill_id", "knowledge-gap-analysis")
    info = SKILL_REGISTRY.get(skill_id, {})
    pipeline_id = request.form.get("pipeline_id", "")
    model_name = request.form.get("model", "")
    excel_file = request.files.get("excel")

    if not pipeline_id and not excel_file:
        return jsonify({"status": "error", "error": "请提供 pipeline_id 或上传知识 Excel 文件"})

    # 解析输入文件
    input_path = None
    schema_cols = []
    if pipeline_id:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.get("step_data", {})
                    schema_cols = sd.get("step1_knowledge_columns") or []
                    resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="compile")
                    if resolved:
                        input_path = str(resolved)
                    break
    if excel_file:
        input_path = save_upload(excel_file, prefix="gap_analysis")

    if not input_path or not os.path.exists(input_path):
        return jsonify({"status": "error", "error": "未找到知识 Excel 文件"})

    if not schema_cols and SCHEMA_PATH.exists():
        from scenario_schema import load_scenario_schema, resolve_knowledge_columns
        schema_cols = resolve_knowledge_columns(load_scenario_schema(SCHEMA_PATH))

    # 程序化统计每列填充率（不需要 LLM）
    col_stats = {}
    total_rows = 0
    try:
        with _safe_workbook(input_path) as wb:
            for ws in wb.worksheets:
                headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
                col_indices = {h: i for i, h in enumerate(headers) if h}
                for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not any(row):
                        continue
                    total_rows += 1
                    for col_name, ci in col_indices.items():
                        val = str(row[ci] or "").strip() if ci < len(row or ()) else ""
                        stat = col_stats.setdefault(col_name, {"filled": 0, "total": 0, "samples": []})
                        stat["total"] += 1
                        if len(val) >= 4:
                            stat["filled"] += 1
                            if len(stat["samples"]) < 3:
                                stat["samples"].append(val[:80])
    except Exception as e:
        return jsonify({"status": "error", "error": f"读取 Excel 失败: {str(e)}"})

    # 生成盲区清单
    gaps = []
    for col, st in col_stats.items():
        rate = st["filled"] / max(st["total"], 1)
        if rate < 0.5:
            gaps.append({
                "column": col,
                "fill_rate": round(rate * 100, 1),
                "filled": st["filled"],
                "total": st["total"],
                "severity": "high" if rate < 0.2 else "medium",
            })

    gaps.sort(key=lambda g: g["fill_rate"])

    # 用 LLM 生成盲区解读和建议
    narrative = ""
    if gaps and model_name:
        model_cfg = get_model_by_name(model_name)
        if not model_cfg:
            models_list = load_llm_config()
            model_cfg = models_list[0] if models_list else None
        if model_cfg:
            gap_summary = "\n".join(
                f"- {g['column']}: 填充率 {g['fill_rate']}%（{g['filled']}/{g['total']}）" for g in gaps[:10]
            )
            schema_summary = "、".join(schema_cols) if schema_cols else "未获取到Schema列定义"
            try:
                llm_result = call_llm_with_retry(model_cfg, [
                    {"role": "system", "content": "你是一位知识工程专家，正在分析知识库的盲区。请用简洁的语言给出3-5条可操作的补全建议。"},
                    {"role": "user", "content": (
                        f"知识库schema定义了以下列：{schema_summary}\n\n"
                        f"以下是填充率低于50%的列（即知识盲区）：\n{gap_summary}\n"
                        f"请给出3-5条具体的补全建议，每条建议说明：（1）应该补充什么类型的知识？"
                        f"（2）建议找谁（什么背景的专家）来补充？（3）为什么这些盲区是高风险的？"
                    )}
                ], stream=False, temperature=0.3, max_tokens=1024)
                narrative = extract_assistant_content(llm_result) if isinstance(llm_result, dict) else ""
            except Exception:
                narrative = "（LLM 解读生成失败，请手动查看盲区统计）"

    # 生成 Markdown 报告
    report_name = f"gap_analysis_{uuid.uuid4().hex[:8]}.md"
    report_path = WORKSPACE / report_name
    lines = [
        f"# 知识盲区检测报告",
        f"",
        f"- 总条目数：**{total_rows}**",
        f"- 检测列数：**{len(col_stats)}**",
        f"- 盲区列数（填充率<50%）：**{len(gaps)}**",
        f"- 生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"",
        f"## 一、各列填充率",
        f"",
        f"| 列名 | 填充率 | 已填充 | 总计 |",
        f"|------|--------|--------|------|",
    ]
    for col, st in sorted(col_stats.items(), key=lambda x: x[1]["filled"] / max(x[1]["total"], 1)):
        rate = st["filled"] / max(st["total"], 1) * 100
        lines.append(f"| {col} | {rate:.0f}% | {st['filled']} | {st['total']} |")
    lines.append("")

    if gaps:
        lines.append("## 二、盲区清单（按严重度排序）")
        lines.append("")
        for g in gaps:
            sev = "🔴 高危" if g["severity"] == "high" else "🟡 中危"
            lines.append(f"- {sev} **{g['column']}**：填充率 {g['fill_rate']}%（{g['filled']}/{g['total']}）")
        lines.append("")

    if narrative:
        lines.append("## 三、补全建议")
        lines.append("")
        lines.append(narrative)

    report_path.write_text("\n".join(lines), encoding="utf-8")

    # 持久化到 pipeline，供 Step3 修订上下文使用
    if pipeline_id:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    sd["step2_gap_analysis_report"] = report_name
                    sd["step2_gap_analysis_url"] = f"/downloads/{report_name}"
                    sd["step2_gap_analysis_summary"] = {
                        "total_rows": total_rows,
                        "gaps": [{"column": g["column"], "fill_rate": g["fill_rate"], "severity": g["severity"]} for g in gaps[:5]],
                    }
                    save_pipelines(pipelines)
                    break

    return jsonify({
        "status": "ok",
        "skill_name": info.get("name", skill_id),
        "skill_id": skill_id,
        "total_rows": total_rows,
        "columns_analyzed": len(col_stats),
        "gaps_found": len(gaps),
        "gaps": gaps,
        "column_stats": {k: {"fill_rate": round(v["filled"] / max(v["total"], 1) * 100, 1)} for k, v in col_stats.items()},
        "narrative": narrative,
        "report_name": report_name,
        "download_url": f"/downloads/{report_name}",
    })


# ─── 新 Skill: 知识保鲜度审计 ─────────────────────────────────

def _execute_freshness_audit():
    """知识保鲜度审计：检测知识库中可能过时的规则和被案例突破的知识点。"""
    skill_id = request.form.get("skill_id", "knowledge-freshness-audit")
    info = SKILL_REGISTRY.get(skill_id, {})
    pipeline_id = request.form.get("pipeline_id", "")
    model_name = request.form.get("model", "")
    excel_file = request.files.get("excel")

    if not pipeline_id and not excel_file:
        return jsonify({"status": "error", "error": "请提供 pipeline_id 或上传知识 Excel 文件"})

    input_path = None
    tacit_annotations = []
    if pipeline_id:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.get("step_data", {})
                    tacit_annotations = sd.get("step4_tacit_annotations") or []
                    resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="compile")
                    if resolved:
                        input_path = str(resolved)
                    break
    if excel_file:
        input_path = save_upload(excel_file, prefix="freshness_audit")

    if not input_path or not os.path.exists(input_path):
        return jsonify({"status": "error", "error": "未找到知识 Excel 文件"})

    # 读取知识条目
    from excel_to_skill import read_excel_knowledge
    records, version_info = read_excel_knowledge(input_path)

    if not records:
        return jsonify({"status": "error", "error": "知识文件中无有效条目"})

    # 程序化统计
    total = len(records)
    high_conf = sum(1 for r in records if str(r.get("置信度", "")).strip() == "高")
    low_conf = sum(1 for r in records if str(r.get("置信度", "")).strip() == "低")
    with_source = sum(1 for r in records if str(r.get("来源文档", "")).strip())
    with_contributor = sum(1 for r in records if str(r.get("贡献专家", "")).strip())

    # 构建保鲜审计提示
    stale_indicators = []
    if total > 0 and high_conf / total < 0.3:
        stale_indicators.append(f"高置信度条目仅占 {round(high_conf/total*100)}%——大量知识缺乏充分验证")
    if total > 0 and with_source / total < 0.5:
        stale_indicators.append(f"仅 {round(with_source/total*100)}% 条目标注了来源——知识溯源困难")
    if total > 0 and with_contributor / total < 0.3:
        stale_indicators.append(f"仅 {round(with_contributor/total*100)}% 条目有贡献专家署名——知识责任人不清")
    if tacit_annotations:
        stale_indicators.append(f"已有 {len(tacit_annotations)} 条隐性注释——这些注释可能暗示某些知识已被实际突破")

    # LLM 深度审计
    narrative = ""
    if model_name:
        model_cfg = get_model_by_name(model_name)
        if not model_cfg:
            models_list = load_llm_config()
            model_cfg = models_list[0] if models_list else None
        if model_cfg:
            item_sample = []
            for r in records[:15]:
                item_sample.append(
                    f"- [{r.get('知识分类', '未分类')}] {str(r.get('知识描述', ''))[:120]}"
                    f"（置信度：{r.get('置信度', '未标')}）"
                )
            annotations_text = ""
            if tacit_annotations:
                annotations_text = "## 专家隐性注释（可能暗示知识已被突破）\n"
                for a in tacit_annotations[:10]:
                    annotations_text += f"- [{a.get('action', '')}] {a.get('question', '')}\n  专家回答：{a.get('answer', '')}\n"

            try:
                llm_result = call_llm_with_retry(model_cfg, [
                    {"role": "system", "content": (
                        "你是一位银行知识管理专家，正在审计知识库的保鲜度。"
                        "请识别以下知识条目中可能已过时、需要更新或被案例突破的内容。"
                        "重点关注：（1）依赖特定时间窗口的规则（如'近三年'）是否仍在有效期内"
                        "（2）依赖特定监管政策的规则是否仍适用（3）被专家隐性注释质疑的知识。"
                    )},
                    {"role": "user", "content": (
                        f"知识条目样本（共{total}条，以下为前15条）：\n"
                        + "\n".join(item_sample) + "\n\n"
                        f"保鲜风险指标：\n" + "\n".join(f"- {s}" for s in stale_indicators) + "\n\n"
                        + annotations_text + "\n\n"
                        f"请给出3-5条保鲜建议，标注哪些类型的知识最需要更新，以及建议的更新优先级。"
                    )}
                ], stream=False, temperature=0.3, max_tokens=1536)
                narrative = extract_assistant_content(llm_result) if isinstance(llm_result, dict) else ""
            except Exception:
                narrative = "（LLM 深度审计生成失败，请参考统计数据）"

    # 生成报告
    report_name = f"freshness_audit_{uuid.uuid4().hex[:8]}.md"
    report_path = WORKSPACE / report_name
    lines = [
        f"# 知识保鲜度审计报告",
        f"",
        f"- 知识条目总数：**{total}**",
        f"- 高置信度占比：**{round(high_conf/total*100) if total else 0}%**",
        f"- 来源覆盖率：**{round(with_source/total*100) if total else 0}%**",
        f"- 专家署名率：**{round(with_contributor/total*100) if total else 0}%**",
        f"- 隐性注释数：**{len(tacit_annotations)}**",
        f"- 生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"",
        f"## 保鲜风险指标",
        f"",
    ]
    for s in stale_indicators:
        lines.append(f"- ⚠️ {s}")
    if not stale_indicators:
        lines.append("- ✅ 当前未检测到明显保鲜风险")

    if narrative:
        lines.append("")
        lines.append("## 深度审计与保鲜建议")
        lines.append("")
        lines.append(narrative)

    report_path.write_text("\n".join(lines), encoding="utf-8")

    return jsonify({
        "status": "ok",
        "skill_name": info.get("name", skill_id),
        "skill_id": skill_id,
        "total_items": total,
        "high_confidence_pct": round(high_conf / total * 100, 1) if total else 0,
        "source_coverage_pct": round(with_source / total * 100, 1) if total else 0,
        "stale_indicators": stale_indicators,
        "tacit_annotations_count": len(tacit_annotations),
        "narrative": narrative,
        "report_name": report_name,
        "download_url": f"/downloads/{report_name}",
    })


# ─── Step 3/4 pipeline outputs ────────────────────────────────────


@app.route("/api/step3/revision_context", methods=["GET"])
def api_step3_revision_context():
    """返回 Step3 修订上下文：模式发现 + 盲区检测的关键发现，供专家修订时参考。"""
    pipeline_id = request.args.get("pipeline_id", "")
    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    with _pipelines_lock:
        pipelines = load_pipelines()
        sd = {}
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {}) or {}
                break

    ctx = {"status": "ok", "insights": [], "warnings": []}

    # 模式发现洞察
    pm = sd.get("step2_pattern_mining_summary")
    if pm:
        top = pm.get("top_signals") or []
        if top:
            ctx["insights"].append({
                "source": "跨案例模式发现",
                "icon": "🔬",
                "text": f"跨 {pm.get('case_count', '?')} 个案例发现 {len(top)} 个高频信号",
                "details": top,
            })
        blind = pm.get("blind_spots") or []
        if blind:
            ctx["warnings"].append({
                "source": "系统性风险盲区",
                "icon": "⚠️",
                "details": blind,
            })

    # 盲区检测洞察
    ga = sd.get("step2_gap_analysis_summary")
    if ga:
        gaps = ga.get("gaps") or []
        high_gaps = [g for g in gaps if g.get("severity") == "high"]
        if high_gaps:
            ctx["warnings"].append({
                "source": "知识盲区检测",
                "icon": "🎯",
                "text": f"发现 {len(high_gaps)} 个高危盲区列",
                "details": [f"{g['column']}（填充率 {g['fill_rate']}%）" for g in high_gaps],
            })
        ctx["insights"].append({
            "source": "知识盲区检测",
            "icon": "📊",
            "text": f"共 {ga.get('total_rows', '?')} 条知识，{len(gaps)} 列填充不足",
        })

    # 隐性注释提醒
    ta = sd.get("step4_tacit_annotations") or []
    if ta:
        ctx["insights"].append({
            "source": "隐性注释",
            "icon": "💡",
            "text": f"已有 {len(ta)} 条专家隐性注释可用于修订参考",
        })

    return jsonify(ctx)


@app.route("/api/step3/prev_output", methods=["GET"])
def api_step3_prev_output():
    """获取Step2知识萃取的输出件，供Step3知识修订使用"""
    pipeline_id = request.args.get("pipeline_id", "")
    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {})
                # Step2 outputs an Excel file with extracted knowledge
                step2_file = sd.get("step2_output_file", "")
                step2_download = sd.get("step2_download_url", "")
                step2_md_file = sd.get("step2_md_file", "")
                step2_md_download = sd.get("step2_md_download_url", "")
                if not is_step2_preextract_filename(step2_file):
                    return jsonify({
                        "status": "ok",
                        "has_output": False,
                        "hint": "未找到有效萃取 Excel（preextract_*.xlsx），请先完成知识萃取（Step2）",
                    })

                # Read the Step2 Excel for structure info and row count
                fields_info = []
                excel_file_to_read = step2_file
                file_path = safe_workspace_path(WORKSPACE, excel_file_to_read, must_exist=True)
                if file_path:
                    try:
                        with _safe_workbook(str(file_path)) as wb:
                            for ws_name in wb.sheetnames:
                                ws = wb[ws_name]
                                headers = []
                                if ws.max_row >= 1 and ws.max_column >= 1:
                                    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                                        headers.append(str(cell.value) if cell.value else "")
                                row_count = ws.max_row - 1 if ws.max_row > 1 else 0
                                fields_info.append({"sheet": ws_name, "headers": headers, "rows": row_count})
                    except Exception:
                        pass

                return jsonify({
                    "status": "ok", "has_output": True,
                    "file_name": step2_file,
                    "download_url": step2_download or ("/downloads/" + step2_file),
                    "markdown_file": step2_md_file,
                    "markdown_download_url": step2_md_download or (f"/downloads/{step2_md_file}" if step2_md_file else ""),
                    "fields_info": fields_info,
                    "scenario": sd.get("skill_extract_scenario", ""),
                    "style": sd.get("skill_extract_style", ""),
                    "extracted_count": sd.get("step2_extracted_count", 0),
                    "step1_file": sd.get("step1_output_file", ""),
                })
    return jsonify({"status": "ok", "has_output": False})


@app.route("/api/step4/prev_output", methods=["GET"])
def api_step4_prev_output():
    """获取知识对齐稿（final_*.xlsx），供智能转化等下游使用"""
    pipeline_id = request.args.get("pipeline_id", "")
    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {})
                file_path_obj, source_key = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="compile")
                if not file_path_obj:
                    return jsonify({"status": "ok", "has_output": False})
                file_path = str(file_path_obj)
                out_name = file_path_obj.name

                try:
                    fields_info = []
                    with _safe_workbook(file_path) as wb:
                        for ws_name in wb.sheetnames:
                            ws = wb[ws_name]
                            headers = []
                            if ws.max_row >= 1 and ws.max_column >= 1:
                                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                                    headers.append(str(cell.value) if cell.value else "")
                            row_count = ws.max_row - 1 if ws.max_row > 1 else 0
                            fields_info.append({"sheet": ws_name, "headers": headers, "rows": row_count})
                    if source_key == "step4_final_file":
                        dl_url = sd.get("step4_download_url") or ("/downloads/" + out_name)
                        md_name = sd.get("step4_md_file", "")
                        md_url = sd.get("step4_md_download_url") or (f"/downloads/{md_name}" if md_name else "")
                        align_style = sd.get("step4_final_style", "")
                        align_count = sd.get("step4_final_count", 0)
                    elif source_key == "step3_revision_file":
                        dl_url = sd.get("step3_download_url") or ("/downloads/" + out_name)
                        md_name = sd.get("step3_md_file", "")
                        md_url = sd.get("step3_md_download_url") or (f"/downloads/{md_name}" if md_name else "")
                        align_style = sd.get("step3_revision_style", "")
                        align_count = sd.get("step3_revision_count", 0)
                    else:
                        dl_url = sd.get("step2_download_url") or ("/downloads/" + out_name)
                        md_name = sd.get("step2_md_file", "")
                        md_url = sd.get("step2_md_download_url") or (f"/downloads/{md_name}" if md_name else "")
                        align_style = sd.get("skill_extract_style", "")
                        align_count = sd.get("step2_extracted_count", 0)
                    return jsonify({
                        "status": "ok", "has_output": True,
                        "file_name": out_name,
                        "download_url": dl_url,
                        "markdown_file": md_name,
                        "markdown_download_url": md_url,
                        "fields_info": fields_info,
                        "revision_style": align_style,
                        "revision_count": align_count,
                        "source": source_key,
                    })
                except Exception:
                    return jsonify({"status": "ok", "has_output": False})
    return jsonify({"status": "ok", "has_output": False})


@app.route("/api/step4/finalize", methods=["POST"])
def api_step4_finalize():
    """知识对齐：基于 Step3 修订稿（或 Step2 萃取稿）+专家意见，生成最终稿"""
    pipeline_id = request.form.get("pipeline_id", "")
    expert_text = request.form.get("expert_text", "")
    expert_cached_file = os.path.basename(request.form.get("expert_cached_file", "").strip())
    style = _normalize_revision_style(request.form.get("style", "标准修订"))
    style_rule = REVISION_STYLE_RULES[style]
    model_name = request.form.get("model", "")

    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})
    _debug_log(
        "H4",
        "app_server.py:api_step4_finalize",
        "step4 finalize start",
        {"has_pipeline_id": bool(pipeline_id), "style": style, "has_expert_text": bool(expert_text)},
    )

    # 支持文件上传
    if not expert_text:
        expert_file = request.files.get("expert_file")
        if expert_file and expert_file.filename:
            expert_text = extract_text_from_file(expert_file)
    if not expert_text and expert_cached_file:
        try:
            cached_path = WORKSPACE / expert_cached_file
            if cached_path.exists():
                expert_text = extract_text_from_path(str(cached_path))
        except Exception:
            pass

    if not expert_text:
        expert_text = ""

    # 对齐输入：已有 final 则在其上再对齐；否则用 Step2 萃取稿（四步法不要求 revision_*.xlsx）
    with _pipelines_lock:
        pipelines = load_pipelines()
        source_file_path = None
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {})
                resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="align")
                if resolved:
                    source_file_path = str(resolved)
                break

    if not source_file_path:
        return jsonify({
            "status": "error",
            "error": "未找到可对齐的知识稿。请先完成「知识萃取」生成 preextract_*.xlsx",
        })

    uploaded_material = _load_align_expert_upload_text(
        uploaded_file=request.files.get("expert_file"),
        cached_file_name=expert_cached_file,
    )
    if not expert_text and uploaded_material:
        expert_text = uploaded_material

    if _should_pass_through_preextract(expert_text, uploaded_material):
        payload = _align_no_opinion_success_payload(
            pipeline_id, source_file_path, expert_text, style, style_rule
        )
        payload["align_mode"] = "pass_through"
        return jsonify(payload)

    try:
        from workbook_layout import build_revision_context, layout_prompt_rules, normalize_revision_notes

        excel_context_str, layout_map = build_revision_context(source_file_path)
    except Exception as e:
        return jsonify({"status": "error", "error": f"读取待对齐稿失败: {str(e)}"})

    llm_expert_text = (expert_text or "").strip()
    if uploaded_material.strip() and uploaded_material not in llm_expert_text:
        llm_expert_text = (
            f"{llm_expert_text}\n\n## 上传材料\n{uploaded_material[:12000]}"
            if llm_expert_text
            else uploaded_material[:12000]
        )

    prompt = f"""你是一位知识管理专家，正在将「专家意见」解析为结构化修订 JSON（不是主动改写知识稿）。

## 当前知识稿（含 excel_row 物理行号，仅供定位引用）
{excel_context_str[:6500]}

## 专家意见（唯一修订依据）
{llm_expert_text}

## 修订风格: {style}
硬规则：{style_rule["prompt_hint"]}

{layout_prompt_rules()}
{_alignment_llm_guard_rules()}

## 输出要求
请严格输出 JSON 数组，每条修订包含：
sheet, row（excel_row）, col（1-based）, action, old_value, new_value, note

```json
[
  {{"sheet": "Sheet1", "row": 5, "col": 3, "action": "modify", "old_value": "原内容", "new_value": "新内容", "note": "说明"}}
]
```"""

    # 调用LLM
    try:
        model = get_model_by_name(model_name)
        if not model:
            models_list = load_llm_config()
            if models_list:
                model = models_list[0]
            else:
                return jsonify({"status": "error", "error": "无可用 LLM 模型"})

        llm_result = call_llm_with_retry(
            model,
            messages=[
                {
                    "role": "system",
                    "content": "你是知识管理专家，仅将专家意见中明确提出的修订解析为 JSON 数组。无明确修订时输出 []。禁止根据知识稿自行编造修订。仅输出 JSON 数组。",
                },
                {"role": "user", "content": prompt},
            ],
            stream=False,
            temperature=style_rule["temperature"],
        )
        llm_text = extract_assistant_content(llm_result) if isinstance(llm_result, dict) else ""
        llm_text = _extract_json_from_text(llm_text)

        expert_notes = json.loads(llm_text)
        if not isinstance(expert_notes, list):
            expert_notes = [expert_notes]
        expert_notes, note_stats = _apply_revision_style_rules(expert_notes, style)
        expert_notes, row_stats = normalize_revision_notes(expert_notes, layout_map)
        note_stats["row_adjusted"] = row_stats.get("adjusted", 0)
        note_stats["row_skipped_header"] = row_stats.get("skipped_header", 0)

    except LlmApiError as e:
        return jsonify({"status": "error", "error": str(e)})
    except json.JSONDecodeError as e:
        return jsonify({"status": "error", "error": f"LLM输出解析失败: {str(e)}", "raw_output": llm_text[:500]})
    except Exception as e:
        return jsonify({"status": "error", "error": f"LLM调用失败: {str(e)}"})

    if not expert_notes and not _should_pass_through_preextract(expert_text, uploaded_material):
        return jsonify({
            "status": "ok",
            "notes": [],
            "total": 0,
            "no_opinion": False,
            "align_mode": "llm_revision",
            "message": "未从专家意见/上传材料中解析出可执行的修订条目，请补充更明确的修改说明（如行号、列、修改内容）。",
            "style": style,
            "style_rule": {
                "mode": style,
                "max_actions": note_stats["max_actions"],
                "raw_count": note_stats["raw_count"],
                "processed_count": note_stats["processed_count"],
            },
        })

    # 调用 revision_processor 生成最终稿
    try:
        from revision_processor import process_workbook
        from datetime import datetime

        output_name = f"final_{pipeline_id[:8]}_{datetime.now().strftime('%H%M%S')}.xlsx"
        output_path = os.path.join(WORKSPACE, output_name)

        revision_count = process_workbook(source_file_path, expert_notes, output_path, layouts=layout_map)
        md_name, md_url = _maybe_generate_markdown_artifact(
            pipeline_id,
            output_name,
            md_prefix="final",
            title=f"Step3 知识对齐 · {pipeline_id[:8]}",
        )

        # 保存到pipeline step_data
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    p.setdefault("step_data", {})["step4_final_file"] = output_name
                    p.setdefault("step_data", {})["step4_download_url"] = "/downloads/" + output_name
                    p.setdefault("step_data", {})["step4_final_notes"] = expert_text[:500]
                    p.setdefault("step_data", {})["step4_final_style"] = style
                    p.setdefault("step_data", {})["step4_final_count"] = revision_count
                    if md_name:
                        p.setdefault("step_data", {})["step4_md_file"] = md_name
                        p.setdefault("step_data", {})["step4_md_download_url"] = md_url
                    else:
                        p.setdefault("step_data", {}).pop("step4_md_file", None)
                        p.setdefault("step_data", {}).pop("step4_md_download_url", None)
                    save_pipelines(pipelines)
                    break

        skill_info = SKILL_REGISTRY.get("knowledge-revision", {})
        return jsonify({
            "status": "ok",
            "revision_count": revision_count,
            "output_file": output_name,
            "download_name": output_name,
            "download_url": "/downloads/" + output_name,
            "markdown_file": md_name,
            "markdown_download_url": md_url,
            "expert_notes_json": expert_notes,
            "skill_name": skill_info.get("name", "知识对齐"),
            "skill_id": "knowledge-revision",
            "style": style,
            "style_rule": {
                "mode": style,
                "max_actions": note_stats["max_actions"],
                "raw_count": note_stats["raw_count"],
                "processed_count": note_stats["processed_count"],
            },
        })
    except Exception as e:
        _debug_log(
            "H4",
            "app_server.py:api_step4_finalize",
            "step4 failure",
            {"error": str(e)[:300]},
        )
        return jsonify({"status": "error", "error": f"最终稿生成失败: {str(e)}"})


def _load_align_expert_upload_text(
    *,
    uploaded_file=None,
    cached_file_name: str = "",
) -> str:
    text_parts = []
    if uploaded_file and getattr(uploaded_file, "filename", None):
        try:
            text_parts.append(extract_text_from_file(uploaded_file).strip())
        except Exception:
            pass
    if cached_file_name:
        try:
            cached_path = WORKSPACE / os.path.basename(cached_file_name)
            if cached_path.exists():
                text_parts.append(extract_text_from_path(str(cached_path)).strip())
        except Exception:
            pass
    return "\n\n".join(p for p in text_parts if p).strip()


def _build_align_preview_from_text(
    pipeline_id: str,
    expert_text: str,
    style: str,
    model_name: str = "",
    *,
    uploaded_material_text: str = "",
):
    """根据专家意见文本生成对齐建议并缓存，供交互式审核与对话式修订复用。"""
    style = _normalize_revision_style(style)
    style_rule = REVISION_STYLE_RULES[style]
    with _pipelines_lock:
        pipelines = load_pipelines()
        source_file_path = None
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {})
                resolved, _src = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="align")
                if resolved:
                    source_file_path = str(resolved)
                break

    if not source_file_path:
        return {"status": "error", "error": "未找到可对齐的知识稿。请先完成「知识萃取」"}

    if _should_pass_through_preextract(expert_text, uploaded_material_text):
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    sd["_align_preview_notes"] = []
                    sd["_align_preview_style"] = style
                    sd["_align_source_file"] = os.path.basename(source_file_path)
                    save_pipelines(pipelines)
                    break
        payload = _align_no_opinion_success_payload(
            pipeline_id, source_file_path, expert_text, style, style_rule
        )
        payload["align_mode"] = "pass_through"
        return payload

    try:
        from workbook_layout import build_revision_context, layout_prompt_rules, normalize_revision_notes
        excel_context_str, layout_map = build_revision_context(source_file_path)
    except Exception as e:
        return {"status": "error", "error": f"读取待对齐稿失败: {str(e)}"}

    source_cells = {}
    try:
        with _safe_workbook(source_file_path) as wb:
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is not None:
                            key = f"{ws.title}:{cell.row}:{cell.column}"
                            source_cells[key] = str(cell.value)
    except Exception:
        pass

    llm_expert_text = (expert_text or "").strip()
    if uploaded_material_text.strip():
        if llm_expert_text:
            llm_expert_text = (
                f"{llm_expert_text}\n\n## 上传材料（会议纪要/访谈记录等）\n{uploaded_material_text[:12000]}"
            )
        else:
            llm_expert_text = uploaded_material_text[:12000]

    prompt = f"""你是一位知识管理专家，正在将「专家意见」解析为结构化修订 JSON（不是主动改写知识稿）。

## 当前知识稿（含 excel_row 物理行号，仅供定位引用）
{excel_context_str[:6500]}

## 专家意见（唯一修订依据）
{llm_expert_text}

## 修订风格: {style}
硬规则：{style_rule["prompt_hint"]}

{layout_prompt_rules()}
{_alignment_llm_guard_rules()}

## 输出要求
请严格输出 JSON 数组，每条修订包含：
sheet, row（excel_row）, col（1-based）, action, old_value, new_value, note

```json
[
  {{"sheet": "Sheet1", "row": 5, "col": 3, "action": "modify", "old_value": "原内容", "new_value": "新内容", "note": "说明"}}
]
```"""

    try:
        model = get_model_by_name(model_name)
        if not model:
            models_list = load_llm_config()
            if models_list:
                model = models_list[0]
            else:
                return {"status": "error", "error": "无可用 LLM 模型"}

        llm_result = call_llm_with_retry(
            model,
            messages=[
                {
                    "role": "system",
                    "content": "你是知识管理专家，仅将专家意见中明确提出的修订解析为 JSON 数组。无明确修订时输出 []。禁止根据知识稿自行编造修订。仅输出 JSON 数组。",
                },
                {"role": "user", "content": prompt},
            ],
            stream=False,
            temperature=style_rule["temperature"],
        )
        llm_text = extract_assistant_content(llm_result) if isinstance(llm_result, dict) else ""
        llm_text = _extract_json_from_text(llm_text)

        expert_notes = json.loads(llm_text)
        if not isinstance(expert_notes, list):
            expert_notes = [expert_notes]
        expert_notes, note_stats = _apply_revision_style_rules(expert_notes, style)
        expert_notes, row_stats = normalize_revision_notes(expert_notes, layout_map)
        note_stats["row_adjusted"] = row_stats.get("adjusted", 0)
        note_stats["row_skipped_header"] = row_stats.get("skipped_header", 0)
    except LlmApiError as e:
        return {"status": "error", "error": str(e)}
    except json.JSONDecodeError as e:
        return {"status": "error", "error": f"LLM输出解析失败: {str(e)}", "raw_output": llm_text[:500]}
    except Exception as e:
        return {"status": "error", "error": f"LLM调用失败: {str(e)}"}

    if not expert_notes:
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    sd["_align_preview_notes"] = []
                    sd["_align_preview_style"] = style
                    sd["_align_source_file"] = os.path.basename(source_file_path)
                    save_pipelines(pipelines)
                    break
        if _should_pass_through_preextract(expert_text, uploaded_material_text):
            payload = _align_no_opinion_success_payload(
                pipeline_id, source_file_path, expert_text, style, style_rule
            )
            payload["align_mode"] = "pass_through"
            return payload
        return {
            "status": "ok",
            "notes": [],
            "total": 0,
            "no_opinion": False,
            "auto_finalized": False,
            "align_mode": "llm_revision",
            "message": "未从专家意见/上传材料中解析出可执行的修订条目，请补充更明确的修改说明（建议注明行号、列名与修改内容）。",
            "style": style,
            "style_rule": {
                "mode": style,
                "max_actions": note_stats["max_actions"],
                "raw_count": note_stats["raw_count"],
                "processed_count": note_stats["processed_count"],
            },
        }

    for i, note in enumerate(expert_notes):
        note["id"] = i
        cell_key = f"{note.get('sheet', '')}:{note.get('row', '')}:{note.get('col', '')}"
        if not note.get("old_value") and cell_key in source_cells:
            note["old_value"] = source_cells[cell_key]

    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.setdefault("step_data", {})
                sd["_align_preview_notes"] = expert_notes
                sd["_align_preview_style"] = style
                sd["_align_source_file"] = os.path.basename(source_file_path)
                save_pipelines(pipelines)
                break

    return {
        "status": "ok",
        "notes": expert_notes,
        "total": len(expert_notes),
        "align_mode": "llm_revision",
        "style": style,
        "style_rule": {
            "mode": style,
            "max_actions": note_stats["max_actions"],
            "raw_count": note_stats["raw_count"],
            "processed_count": note_stats["processed_count"],
        },
    }


@app.route("/api/step4/align_preview", methods=["POST"])
def api_step4_align_preview():
    """交互式对齐 Phase 1：返回 AI 对齐建议列表，但不实际修改 Excel。"""
    pipeline_id = request.form.get("pipeline_id", "")
    expert_text = request.form.get("expert_text", "")
    expert_cached_file = os.path.basename(request.form.get("expert_cached_file", "").strip())
    style = _normalize_revision_style(request.form.get("style", "标准修订"))
    style_rule = REVISION_STYLE_RULES[style]
    model_name = request.form.get("model", "")

    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})
    # region agent log
    _agent_debug_log(
        "run-1",
        "H1",
        "app_server.py:api_step4_align_preview:entry",
        "align_preview request received",
        {
            "has_pipeline_id": bool(pipeline_id),
            "style": style,
            "has_expert_text": bool(expert_text),
            "has_cached_file": bool(expert_cached_file),
            "model_name": model_name or "",
        },
    )
    # endregion

    if not expert_text:
        expert_file = request.files.get("expert_file")
        if expert_file and expert_file.filename:
            expert_text = extract_text_from_file(expert_file)
    if not expert_text and expert_cached_file:
        try:
            cached_path = WORKSPACE / expert_cached_file
            if cached_path.exists():
                expert_text = extract_text_from_path(str(cached_path))
        except Exception:
            pass

    if not expert_text:
        expert_text = ""

    uploaded_material = _load_align_expert_upload_text(
        uploaded_file=request.files.get("expert_file"),
        cached_file_name=expert_cached_file,
    )
    preview = _build_align_preview_from_text(
        pipeline_id,
        expert_text,
        style,
        model_name,
        uploaded_material_text=uploaded_material,
    )
    return jsonify(preview)


@app.route("/api/step4/align_chat", methods=["POST"])
def api_step4_align_chat():
    """对话式知识对齐：按聊天轮次累积专家意见，返回模型回复与可审核修订建议。"""
    pipeline_id = request.form.get("pipeline_id", "")
    message = (request.form.get("message", "") or request.form.get("expert_text", "")).strip()
    expert_cached_file = os.path.basename(request.form.get("expert_cached_file", "").strip())
    style = _normalize_revision_style(request.form.get("style", "标准修订"))
    model_name = request.form.get("model", "")

    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    uploaded_material = _load_align_expert_upload_text(
        uploaded_file=request.files.get("expert_file"),
        cached_file_name=expert_cached_file,
    )
    if not message and uploaded_material:
        message = uploaded_material[:8000]

    user_history = []
    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.setdefault("step_data", {})
                history = sd.get("_align_chat_history", [])
                if isinstance(history, list):
                    user_history = [h for h in history if isinstance(h, dict)]
                break

    if message or _uploaded_expert_material_is_substantive(uploaded_material):
        user_history.append({
            "role": "user",
            "content": (message or uploaded_material[:2000]).strip(),
            "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    # 仅保留最近 12 轮，避免上下文无限增长
    user_history = user_history[-24:]
    user_turns = [h for h in user_history if h.get("role") == "user" and str(h.get("content", "")).strip()]
    check_text = "\n\n".join(str(item.get("content", "")).strip() for item in user_turns[-12:])
    merged_expert_text = "\n\n".join(
        [f"第{i + 1}轮专家意见：\n{str(item.get('content', '')).strip()}" for i, item in enumerate(user_turns[-12:])]
    )
    pass_through = _should_pass_through_preextract(check_text, uploaded_material)
    expert_for_build = check_text if pass_through else (merged_expert_text or check_text)

    preview_result = _build_align_preview_from_text(
        pipeline_id,
        expert_for_build,
        style,
        model_name,
        uploaded_material_text=uploaded_material,
    )
    if preview_result.get("status") != "ok":
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    sd["_align_chat_history"] = user_history
                    save_pipelines(pipelines)
                    break
        return jsonify(preview_result)

    notes = preview_result.get("notes", []) or []
    if notes:
        assistant_reply = (
            f"已结合当前与历史意见，生成 {len(notes)} 条修订建议。"
            "请在下方逐条采纳/驳回/编辑后生成对齐稿。"
        )
    elif preview_result.get("auto_finalized"):
        assistant_reply = preview_result.get("message") or "已自动确认当前稿为对齐稿（无修订）。"
    else:
        assistant_reply = preview_result.get("message") or "本轮未识别到可执行修订。你可以继续补充更具体的修改点。"

    user_history.append({
        "role": "assistant",
        "content": assistant_reply,
        "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    user_history = user_history[-24:]

    with _pipelines_lock:
        pipelines = load_pipelines()
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.setdefault("step_data", {})
                sd["_align_chat_history"] = user_history
                save_pipelines(pipelines)
                break

    preview_result["assistant_reply"] = assistant_reply
    preview_result["chat_history"] = user_history
    return jsonify(preview_result)


@app.route("/api/step4/apply_notes", methods=["POST"])
def api_step4_apply_notes():
    """交互式对齐 Phase 2：按用户选择的建议子集生成 final_*.xlsx。"""
    data = request.get_json(force=True)
    pipeline_id = data.get("pipeline_id", "")
    accepted_ids = set(data.get("accepted_ids", []))
    raw_edited_notes = data.get("edited_notes", [])
    tacit_annotations = data.get("tacit_annotations", [])  # 隐性注释 — 专家修订背后的经验分享
    # region agent log
    _agent_debug_log(
        "run-1",
        "H2",
        "app_server.py:api_step4_apply_notes:entry",
        "apply_notes request received",
        {
            "has_pipeline_id": bool(pipeline_id),
            "accepted_count": len(accepted_ids),
            "accepted_id_types": sorted(list({type(v).__name__ for v in accepted_ids})),
            "edited_count_raw": len(raw_edited_notes) if isinstance(raw_edited_notes, list) else -1,
        },
    )
    # endregion
    edited_notes = {int(n["id"]): n for n in raw_edited_notes if "id" in n}

    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})
    if not accepted_ids:
        return jsonify({"status": "error", "error": "请至少采纳一条对齐建议"})

    with _pipelines_lock:
        pipelines = load_pipelines()
        source_file_path = None
        cached_notes = None
        cached_style = "标准修订"
        for p in pipelines:
            if p["id"] == pipeline_id:
                sd = p.get("step_data", {})
                cached_notes = sd.get("_align_preview_notes")
                cached_style = sd.get("_align_preview_style", "标准修订")
                src_name = sd.get("_align_source_file", "")
                if src_name:
                    resolved = safe_workspace_path(WORKSPACE, src_name, must_exist=True)
                    if resolved:
                        source_file_path = str(resolved)
                if not source_file_path:
                    resolved2, _ = resolve_knowledge_workbook_path(WORKSPACE, sd, purpose="align")
                    if resolved2:
                        source_file_path = str(resolved2)
                break

    if not cached_notes:
        return jsonify({"status": "error", "error": "未找到预览建议缓存，请重新执行「生成对齐建议」"})
    if not source_file_path:
        return jsonify({"status": "error", "error": "未找到源知识稿"})

    final_notes = []
    for note in cached_notes:
        nid = note.get("id")
        if nid not in accepted_ids:
            continue
        if nid in edited_notes:
            merged = {**note, **edited_notes[nid]}
            final_notes.append(merged)
        else:
            final_notes.append(note)
    # region agent log
    _agent_debug_log(
        "run-1",
        "H3",
        "app_server.py:api_step4_apply_notes:selection",
        "apply_notes selection materialized",
        {
            "cached_notes_count": len(cached_notes),
            "accepted_ids_count": len(accepted_ids),
            "edited_notes_count": len(edited_notes),
            "final_notes_count": len(final_notes),
        },
    )
    # endregion

    try:
        from workbook_layout import build_revision_context, normalize_revision_notes
        _, layout_map = build_revision_context(source_file_path)
        from revision_processor import process_workbook
        from datetime import datetime

        output_name = f"final_{pipeline_id[:8]}_{datetime.now().strftime('%H%M%S')}.xlsx"
        output_path = os.path.join(WORKSPACE, output_name)
        revision_count = process_workbook(source_file_path, final_notes, output_path, layouts=layout_map)
        md_name, md_url = _maybe_generate_markdown_artifact(
            pipeline_id,
            output_name,
            md_prefix="final",
            title=f"Step3 知识对齐 · {pipeline_id[:8]}",
        )

        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    sd = p.setdefault("step_data", {})
                    sd["step4_final_file"] = output_name
                    sd["step4_download_url"] = "/downloads/" + output_name
                    sd["step4_final_notes"] = json.dumps(
                        [{"action": n.get("action"), "note": n.get("note", "")} for n in final_notes],
                        ensure_ascii=False,
                    )[:500]
                    if tacit_annotations:
                        sd["step4_tacit_annotations"] = [
                            {"note_id": a.get("note_id"), "action": a.get("action"),
                             "question": a.get("question"), "answer": a.get("answer")}
                            for a in tacit_annotations if a.get("answer")
                        ]
                    sd["step4_final_style"] = cached_style
                    sd["step4_final_count"] = revision_count
                    if md_name:
                        sd["step4_md_file"] = md_name
                        sd["step4_md_download_url"] = md_url
                    else:
                        sd.pop("step4_md_file", None)
                        sd.pop("step4_md_download_url", None)
                    sd.pop("_align_preview_notes", None)
                    sd.pop("_align_preview_style", None)
                    sd.pop("_align_source_file", None)
                    sd.pop("_align_chat_history", None)
                    save_pipelines(pipelines)
                    break
        # region agent log
        _agent_debug_log(
            "run-1",
            "H4",
            "app_server.py:api_step4_apply_notes:success",
            "apply_notes generated final file",
            {
                "output_name": output_name,
                "revision_count": revision_count,
                "accepted_count": len(final_notes),
            },
        )
        # endregion

        return jsonify({
            "status": "ok",
            "revision_count": revision_count,
            "accepted_count": len(final_notes),
            "total_suggested": len(cached_notes),
            "output_file": output_name,
            "download_name": output_name,
            "download_url": "/downloads/" + output_name,
            "markdown_file": md_name,
            "markdown_download_url": md_url,
        })
    except Exception as e:
        return jsonify({"status": "error", "error": f"生成对齐稿失败: {str(e)}"})


@app.route("/api/step4/confirm_as_is", methods=["POST"])
def api_step4_confirm_as_is():
    """专家无修订意见时，将当前对齐输入稿直接确认为 final_*.xlsx。"""
    data = request.get_json(force=True) or {}
    pipeline_id = data.get("pipeline_id", "")
    if not pipeline_id:
        return jsonify({"status": "error", "error": "缺少 pipeline_id"})

    source_file_path, _ = _resolve_align_source_for_pipeline(pipeline_id)
    if not source_file_path:
        return jsonify({"status": "error", "error": "未找到可对齐的知识稿"})

    try:
        style = "标准修订"
        with _pipelines_lock:
            pipelines = load_pipelines()
            for p in pipelines:
                if p["id"] == pipeline_id:
                    style = p.get("step_data", {}).get("_align_preview_style") or style
                    break
        output_name, md_name, md_url = _publish_final_from_source(pipeline_id, source_file_path, style=style)
        return jsonify({
            "status": "ok",
            "revision_count": 0,
            "accepted_count": 0,
            "total_suggested": 0,
            "output_file": output_name,
            "download_name": output_name,
            "download_url": "/downloads/" + output_name,
            "markdown_file": md_name,
            "markdown_download_url": md_url,
            "message": "已确认当前稿为对齐稿（无修订）",
        })
    except Exception as e:
        return jsonify({"status": "error", "error": f"确认对齐稿失败: {str(e)}"})


# ─── Main ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Tacit Knowledge Extraction Web App")
    parser.add_argument("--host", default="0.0.0.0", help="Host to bind")
    parser.add_argument("--port", type=int, default=5000, help="Port to bind")
    parser.add_argument("--debug", action="store_true", help="Enable debug mode")
    args = parser.parse_args()

    print(f"Starting server at http://{args.host}:{args.port}")
    print(f"Workspace: {WORKSPACE}")
    print(f"Frontend dir: {FRONTEND_DIR}")
    _vendor_checks = [
        FRONTEND_DIR / "vendor" / "luckysheet" / "plugins" / "js" / "plugin.js",
        FRONTEND_DIR / "vendor" / "luckysheet" / "luckysheet.umd.js",
    ]
    for p in _vendor_checks:
        if p.exists():
            print(f"  [OK] Excel editor asset: {p.relative_to(FRONTEND_DIR)}")
        else:
            print(f"  [WARN] Missing {p} — run: cd frontend && npm install luckysheet@2.1.13 jquery@3.6.4 --no-save && node ../scripts/copy-frontend-vendor.js")
    app.run(host=args.host, port=args.port, debug=args.debug)


if __name__ == "__main__":
    main()
